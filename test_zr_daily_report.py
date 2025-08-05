import unittest
from datetime import datetime, date
import os
import json
import mysql.connector
import csv
from openpyxl import load_workbook
from ZR_Daily_Report import (
    validate_file_name, 
    generate_excel_with_chart, 
    validate_inventory_value, 
    connect_to_database, 
    fetch_inventory_data_from_db,
    read_devices_from_csv,
    get_device_id_by_no,
    generate_query_for_device,
    validate_date
)

class TestZRDailyReport(unittest.TestCase):
    """综合测试ZR Daily Report的所有功能"""
    
    def setUp(self):
        """测试前准备"""
        # 读取配置文件
        with open('query_config.json', 'r', encoding='utf-8') as f:
            self.query_config = json.load(f)
            self.db_config_data = {'db_config': self.query_config.get('db_config', {})}

    def test_file_name_validation(self):
        """测试文件名验证功能"""
        print("\n开始测试文件名验证...")
        valid_filename = "AW46抗磨液压油_TW24011700700016_202401.xlsx"
        invalid_filenames = [
            "AW46抗磨液压油-TW24011700700016-202401.xlsx",  # 错误连接符
            "AW46抗磨液压油_TW24011700700016_2024.xlsx",    # 错误月份格式
            "AW46抗磨液压油_TW24011700700016.xlsx",         # 缺少月份
            "AW46抗磨液压油_TW24011700700016_202401.xls"    # 错误扩展名
        ]
        
        # 测试有效文件名
        device_code = validate_file_name(valid_filename)
        self.assertEqual(device_code, "TW24011700700016")
        print(f"验证通过: {valid_filename}，设备编码: {device_code}")
        
        # 测试无效文件名
        for filename in invalid_filenames:
            with self.assertRaises(SystemExit, msg=f"文件名应该失败: {filename}"):
                validate_file_name(filename)
                print(f"验证失败: {filename}")

    def test_database_connection(self):
        """测试数据库连接功能"""
        print("\n开始测试数据库连接...")
        
        db_config = self.db_config_data.get('db_config', {})
        connection = connect_to_database(db_config)
        
        # 检查连接是否成功建立
        self.assertIsNotNone(connection)
        self.assertTrue(connection.is_connected())
        
        # 测试执行简单查询
        cursor = connection.cursor()
        cursor.execute("SELECT VERSION()")
        version = cursor.fetchone()
        print(f"数据库版本: {version[0]}")
        
        # 关闭连接
        cursor.close()
        connection.close()
        print("数据库连接功能测试通过")

    def test_get_device_id_by_no(self):
        """测试根据设备编号查询设备ID功能"""
        print("\n开始测试根据设备编号查询设备ID功能...")
        
        db_config = self.db_config_data.get('db_config', {})
        connection = connect_to_database(db_config)
        
        # 从配置中获取SQL模板
        sql_templates = self.query_config.get('sql_templates', {})
        device_query_template = sql_templates.get('device_id_query', "SELECT id FROM oil.t_device WHERE device_no = %s ORDER BY create_time DESC LIMIT 1")
        
        # 测试查询一个不存在的设备编号
        device_id = get_device_id_by_no(connection, "NON_EXISTENT_DEVICE", device_query_template)
        self.assertIsNone(device_id)
        
        connection.close()
        print("根据设备编号查询设备ID功能测试通过（验证了函数能正确处理不存在的设备）")

    def test_fetch_inventory_data_from_db(self):
        """测试从数据库获取库存数据功能"""
        print("\n开始测试从数据库获取库存数据功能...")
        
        db_config = self.db_config_data.get('db_config', {})
        
        # 连接数据库并获取数据
        connection = connect_to_database(db_config)
        
        # 使用一个简单的测试查询
        test_query = "SELECT 1 as id, NOW() as '加注时间', '测试油品' as '油品名称', 0.95 as '原油剩余比例'"
        data, columns, raw_data = fetch_inventory_data_from_db(connection, test_query)
        connection.close()
        
        # 验证返回的数据结构
        self.assertIsInstance(data, list)
        self.assertIsInstance(columns, list)
        self.assertIsInstance(raw_data, list)
        
        print(f"从数据库获取库存数据功能测试通过，共获取到 {len(data)} 条处理后的数据")

    def test_read_devices_from_csv(self):
        """测试读取设备信息CSV文件功能"""
        print("\n开始测试读取设备信息CSV文件功能...")
        
        # 创建临时设备信息文件用于测试 (UTF-8编码)
        test_devices = [
            {"device_no": "MO24111301600008", "start_date": "2025-07-01", "end_date": "2025-07-31"},
            {"device_no": "TW24011700700016", "start_date": "2025-07-01", "end_date": "2025-07-31"}
        ]
        
        # 写入临时CSV文件 (UTF-8编码)
        with open('temp_devices.csv', 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=["device_no", "start_date", "end_date"])
            writer.writeheader()
            writer.writerows(test_devices)
        
        try:
            # 读取设备信息
            devices = read_devices_from_csv('temp_devices.csv')
            
            # 验证结果
            self.assertIsInstance(devices, list)
            self.assertEqual(len(devices), 2)
            
            first_device = devices[0]
            self.assertIn('device_no', first_device)
            self.assertIn('start_date', first_device)
            self.assertIn('end_date', first_device)
            
        finally:
            # 清理临时文件
            if os.path.exists('temp_devices.csv'):
                os.remove('temp_devices.csv')
        
        print("读取设备信息CSV文件功能测试通过")

    def test_excel_generation(self):
        """测试Excel文件生成功能"""
        print("\n开始测试Excel文件生成...")
        
        # 模拟数据
        test_data = [
            (date(2024, 1, 1), 80.0),
            (date(2024, 1, 2), 75.0),
            (date(2024, 1, 3), 70.0)
        ]
        
        start_date = date(2024, 1, 1)
        end_date = date(2024, 1, 3)
        
        # 生成Excel文件
        output_dir = 'test_output'
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "test_excel_generation.xlsx")
        
        try:
            generate_excel_with_chart(
                test_data,
                output_file,
                "TW24011700700016",
                start_date,
                end_date
            )
            
            # 验证文件存在
            self.assertTrue(os.path.exists(output_file), f"Excel文件未生成: {output_file}")
            print(f"Excel文件已生成: {output_file}")
            
            # 验证Excel内容
            wb = load_workbook(output_file)
            ws = wb.active
            
            # 检查标题
            title_cell = ws['A1']
            self.assertIn("TW24011700700016每日库存余量变化趋势", title_cell.value)
            
            # 检查表头
            self.assertEqual(ws['A2'].value, "日期")
            self.assertEqual(ws['B2'].value, "库存百分比")
            
            # 验证数据行数（包括补全的数据）
            # 由于数据补全功能，应该有完整的日期范围数据
            expected_rows = (end_date - start_date).days + 1 + 2  # +2 for title and header rows
            actual_rows = ws.max_row
            self.assertEqual(actual_rows, expected_rows)
            
            print("Excel内容验证通过")
            
            # 验证图表是否存在
            self.assertTrue(len(ws._charts) > 0, "图表未添加到工作表")
            chart = ws._charts[0]
            # 获取图表标题文本
            chart_title = chart.title.tx.rich.p[0].r[0].t if chart.title and chart.title.tx and chart.title.tx.rich and chart.title.tx.rich.p and chart.title.tx.rich.p[0].r else None
            self.assertEqual(chart_title, "每日库存余量变化趋势")
            print("图表已成功添加")
            
        except Exception as e:
            print(f"Excel生成测试失败: {e}")
            raise
        finally:
            # 清理测试文件
            if os.path.exists(output_file):
                os.remove(output_file)

    def test_inventory_value_validation(self):
        """测试库存值验证功能"""
        print("\n开始测试库存值验证...")
        
        # 测试有效值
        valid_values = [0, 50, 100, 150, 1000]
        for value in valid_values:
            result = validate_inventory_value(value)
            self.assertEqual(result, value)
            print(f"验证通过 - 有效值: {value}")
        
        # 测试无效值
        invalid_values = [-1, -10, -100]
        for value in invalid_values:
            with self.assertRaises(ValueError, msg=f"值应为无效: {value}"):
                validate_inventory_value(value)
            print(f"验证通过 - 无效值正确抛出异常: {value}")

    def test_empty_data_handling(self):
        """测试空数据处理功能"""
        print("\n开始测试空数据处理...")
        
        # 测试空数据列表
        empty_data = []
        start_date = date(2024, 1, 1)
        end_date = date(2024, 1, 3)
        
        output_dir = 'test_output'
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, "test_empty_data.xlsx")
        
        # 空数据不会引发异常，而是会生成默认数据图表
        try:
            generate_excel_with_chart(
                empty_data,
                output_file,
                "TW24011700700016",
                start_date,
                end_date
            )
            
            # 验证文件存在
            self.assertTrue(os.path.exists(output_file), f"Excel文件未生成: {output_file}")
            print("空数据处理测试通过 - 成功生成默认数据图表")
            
        except Exception as e:
            print(f"空数据处理测试失败: {e}")
            raise
        finally:
            # 清理测试文件
            if os.path.exists(output_file):
                os.remove(output_file)

    def test_date_validation(self):
        """测试日期验证功能"""
        print("\n开始测试日期验证功能...")
        
        # 测试有效日期
        valid_dates = ["2025-07-01", "2025-12-31", "2024-02-29", "2025/07/01", "2025/7/1"]
        for date_str in valid_dates:
            self.assertTrue(validate_date(date_str), f"有效日期 {date_str} 应该通过验证")
            print(f"验证通过 - 有效日期: {date_str}")
        
        # 测试无效日期
        invalid_dates = ["2025-13-01", "2025-02-30", "07-01-2025", ""]
        for date_str in invalid_dates:
            self.assertFalse(validate_date(date_str), f"无效日期 {date_str} 应该验证失败")
            print(f"验证通过 - 无效日期正确识别: {date_str}")

    def test_config_file_reading(self):
        """测试配置文件读取功能"""
        print("\n开始测试配置文件读取功能...")
        
        # 测试数据库配置
        db_config = self.db_config_data.get('db_config', {})
        self.assertIn('host', db_config)
        self.assertIn('port', db_config)
        self.assertIn('user', db_config)
        self.assertIn('password', db_config)
        self.assertIn('database', db_config)
        
        # 测试SQL模板配置
        self.assertIn('sql_templates', self.query_config)
        sql_templates = self.query_config['sql_templates']
        self.assertIn('device_id_query', sql_templates)
        self.assertIn('inventory_query', sql_templates)
        
        print("配置文件读取功能测试通过")

if __name__ == '__main__':
    print("开始执行ZR Daily Report综合测试...")
    unittest.main(verbosity=2)