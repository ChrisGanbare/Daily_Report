"""
基础报表类模块
提供所有报表生成器的通用功能
"""
import os
from abc import ABC, abstractmethod
from datetime import datetime
from openpyxl.styles import Alignment, Font


class BaseReportGenerator(ABC):
    """
    基础报表生成器类
    提供所有报表生成器的通用功能，如文件操作、日期处理等
    """
    
    def __init__(self):
        """初始化基础报表生成器"""
        pass

    def _validate_date_range(self, start_date, end_date):
        """
        验证日期范围的有效性
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            bool: 日期范围是否有效
            
        Raises:
            ValueError: 日期范围无效时抛出异常
        """
        if start_date > end_date:
            raise ValueError("开始日期不能晚于结束日期")
        return True

    def _create_output_directory(self, output_path):
        """
        创建输出目录
        
        Args:
            output_path (str): 输出文件路径
            
        Returns:
            str: 输出目录路径
        """
        output_dir = os.path.dirname(output_path)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        return output_dir

    @abstractmethod
    def generate_report(self, data, output_file_path, **kwargs):
        """
        生成报表的抽象方法
        
        Args:
            data: 报表数据
            output_file_path (str): 输出文件路径
            **kwargs: 其他参数
            
        Returns:
            bool: 报表生成是否成功
        """
        pass

    def _format_date_for_display(self, date):
        """
        格式化日期用于显示
        
        Args:
            date: 日期对象
            
        Returns:
            str: 格式化后的日期字符串
        """
        if date:
            return date.strftime("%Y-%m-%d")
        return "未知日期"

    def _get_unique_filename(self, file_path):
        """
        生成唯一的文件名，如果文件已存在，则添加序号

        Args:
            file_path (str): 原始文件路径

        Returns:
            str: 唯一的文件路径
        """
        # 如果文件不存在，直接返回原路径
        if not os.path.exists(file_path):
            return file_path

        # 分离文件名和扩展名
        base_name, extension = os.path.splitext(file_path)
        counter = 1
        new_file_path = file_path

        # 循环查找未被占用的文件名
        while os.path.exists(new_file_path):
            new_file_path = f"{base_name}({counter}){extension}"
            counter += 1

        return new_file_path

    def _get_config_hint_text(self):
        """
        获取配置信息提示文本
        
        Returns:
            tuple: (基础提示文本, 配置信息文本)
        """
        hint_text_base = "提示：【设备桶数】已从配置文件自动读取，未配置的设备默认桶数为1；若设备数据变动，需同步维护test_data/device_config.csv配置文件，保持一次维护、多次复用的准确性；"
        
        config_info_text = ""
        try:
            from src.core.device_config_manager import DeviceConfigManager
            config_manager = DeviceConfigManager()
            config_info = config_manager.get_config_info()
            
            if config_info['file_exists']:
                config_path = config_info['config_file_abspath']
                maintenance_time = config_info.get('last_maintenance_time', '未知')
                maintenance_type = config_info.get('last_maintenance_type', '')
                
                config_info_text = f"配置文件位置: {config_path}   最近维护时间: {maintenance_time}"
                if maintenance_type:
                    config_info_text += f" ({maintenance_type})"
        except Exception:
            # 如果获取配置信息失败，忽略错误
            pass
        
        return hint_text_base, config_info_text

    def _extract_value_from_dict_or_scalar(self, data, default=0):
        """
        从字典或标量值中提取值
        
        Args:
            data: 可能是字典或标量值
            default: 默认值
            
        Returns:
            提取的值
        """
        if isinstance(data, dict):
            return data.get('value', default)
        return data if data is not None else default

    def _add_config_hint_row(self, ws, row_num, start_col=1, end_col=20):
        """
        添加配置信息提示行到工作表
        
        Args:
            ws: 工作表对象
            row_num: 行号
            start_col: 起始列
            end_col: 结束列
        """
        hint_text_base, config_info_text = self._get_config_hint_text()
        
        if config_info_text:
            hint_text = hint_text_base + "\n" + config_info_text
        else:
            hint_text = hint_text_base
        
        ws.append([hint_text])
        ws.merge_cells(start_row=row_num, start_column=start_col, end_row=row_num, end_column=end_col)
        hint_cell = ws.cell(row=row_num, column=start_col)
        hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        hint_cell.font = Font(size=9, color="FF4500", bold=True, italic=False)
        ws.row_dimensions[row_num].height = 35

    def _add_calculation_rules(self, ws, annotation_row, annotation_col, barrel_count, period_type="日"):
        """
        添加计算规则说明到工作表
        
        Args:
            ws: 工作表对象
            annotation_row: 起始行号
            annotation_col: 起始列号
            barrel_count: 桶数
            period_type: 周期类型（"日" 或 "月"）
        """
        from openpyxl.styles import Font
        
        # 添加标题
        ws.cell(row=annotation_row, column=annotation_col).value = "计算规则说明："
        ws.cell(row=annotation_row, column=annotation_col).font = Font(bold=True)
        
        # 将每个计算规则分别写入单独的行，并合并单元格以确保完整显示
        annotation_row += 1
        consumption_formula = f"库存消耗总量(L) = (前{period_type}库存 - 当{period_type}库存 + 当{period_type}加油（入库）量)"
        if barrel_count > 1:
            consumption_formula += f" * {barrel_count} (桶数)"
        ws.cell(row=annotation_row, column=annotation_col).value = consumption_formula
        ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)
        
        annotation_row += 1
        ws.cell(row=annotation_row, column=annotation_col).value = "中润亏损(L) = MAX(0, 库存消耗总量 - 订单累积总量)"
        ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)
        
        annotation_row += 1
        ws.cell(row=annotation_row, column=annotation_col).value = "客户亏损(L) = MAX(0, 订单累积总量 - 库存消耗总量)"
        ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)