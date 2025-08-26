import csv
import os
import time
from datetime import timedelta
from decimal import Decimal

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill

from .base_report import BaseReportGenerator


class InventoryReportGenerator(BaseReportGenerator):
    """库存报表生成器，专门负责生成设备库存报表和图表"""

    def __init__(self):
        """初始化库存报表生成器"""
        super().__init__()

    def generate_report(self, inventory_data, output_file_path, **kwargs):
        """
        生成库存报表的实现方法
        
        Args:
            inventory_data (list): 库存数据列表
            output_file_path (str): 输出文件路径
            **kwargs: 其他参数
            
        Returns:
            bool: 报表生成是否成功
        """
        device_code = kwargs.get('device_code')
        start_date = kwargs.get('start_date')
        end_date = kwargs.get('end_date')
        oil_name = kwargs.get('oil_name')
        chart_style = kwargs.get('chart_style')
        export_format = kwargs.get('export_format', 'xlsx')
        
        return self.generate_inventory_report_with_chart(
            inventory_data, output_file_path, device_code, start_date, end_date,
            oil_name, chart_style, export_format
        )

    def generate_inventory_report_with_chart(
        self,
        inventory_data,
        output_file_path,
        device_code,
        start_date,
        end_date,
        oil_name=None,
        chart_style=None,
        export_format="xlsx",
    ):
        """
        生成包含库存数据和趋势图表的Excel报告文件

        Args:
            inventory_data (list): 库存数据列表，格式为[(date, value), ...]
            output_file_path (str): 输出文件路径
            device_code (str): 设备编码
            start_date (date): 开始日期
            end_date (date): 结束日期
            oil_name (str): 油品名称
            chart_style (dict): 图表样式配置
            export_format (str): 导出格式，支持xlsx和csv
        """
        try:
            # 验证并清理数据
            cleaned_data = []
            invalid_records = []
            for date, value in inventory_data:
                try:
                    validated_value = self._validate_inventory_value(value)
                    if validated_value > 100:
                        print(f"提示：日期 {date} 的库存值 {validated_value}% 超过100%")
                    cleaned_data.append((date, validated_value))
                except ValueError as e:
                    invalid_records.append((date, value, str(e)))
                    print(f"警告：日期 {date} 的数据已跳过 - {str(e)}")

            if invalid_records:
                print("\n无效数据汇总：")
                for date, value, reason in invalid_records:
                    print(f"- {date}: {value} ({reason})")

            # 如果没有有效数据，尝试生成一个带有默认值的图表
            if not cleaned_data:
                print("警告：没有有效的库存数据可供处理，将生成默认数据图表")
                # 使用默认数据点，确保能生成图表
                cleaned_data = [(start_date, 0), (end_date, 0)]
                print(f"使用默认数据点: {cleaned_data}")

            # 初始化工作簿
            wb = Workbook()

            # 处理不同导出格式
            if export_format.lower() == "csv":
                with open(
                    output_file_path.replace(".xlsx", ".csv"), "w", newline=""
                ) as f:
                    writer = csv.writer(f)
                    writer.writerow(["日期", "库存百分比"])
                    writer.writerows(cleaned_data)
                print(
                    f"数据已导出为CSV格式：{output_file_path.replace('.xlsx', '.csv')}"
                )
                # 立即关闭工作簿
                wb.close()
                return

            # 补全数据
            data_dict = dict(cleaned_data)
            complete_data = []
            current_date = start_date

            # 处理空数据情况，避免索引错误
            if cleaned_data:
                last_inventory = next(iter(cleaned_data))[1]
            else:
                last_inventory = 0
                print("警告：没有有效的库存数据可供处理，将生成默认数据图表")
                # 使用默认数据点，确保能生成图表
                cleaned_data = [(start_date, 0), (end_date, 0)]
                print(f"使用默认数据点: {cleaned_data}")

            while current_date <= end_date:
                current_inventory = data_dict.get(current_date, last_inventory)
                complete_data.append([current_date, current_inventory])
                last_inventory = current_inventory
                current_date += timedelta(days=1)

            # Excel处理
            wb = Workbook()
            ws = wb.active
            ws.title = "库存数据"

            # 添加标题行
            oil_name_str = f" {oil_name} " if oil_name else " "
            title = f"{device_code}{oil_name_str}每日库存余量变化趋势({start_date} - {end_date})"
            ws.append([title])
            # 将合并单元格的宽度增加到18列
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=1, column=1).font = Font(size=14, bold=True)

            # 添加数据列标题
            ws.append(["日期", "库存百分比"])

            # 写入补全后的数据
            for row in complete_data:
                ws.append(row)

            # 调整列宽
            ws.column_dimensions["A"].width = 12  # 日期列宽度
            ws.column_dimensions["B"].width = 12  # 库存百分比列宽度

            # 创建图表
            chart = LineChart()
            chart.title = "每日库存余量变化趋势"
            chart.style = 13
            chart.y_axis.title = "库存百分比"
            chart.x_axis.title = "日期"

            # 设置图表显示数据标签

            # 添加这部分配置来调整x轴日期显示
            chart.x_axis.tickLblSkip = 3  # 每隔3个标签显示一个
            chart.x_axis.tickLblPos = "low"  # 将标签位置调整到底部
            chart.x_axis.textRotation = 0  # 将文本旋转角度设为0度（水平显示）

            # 设置数据范围
            data_range = Reference(
                ws, min_col=2, min_row=2, max_col=2, max_row=len(complete_data) + 2
            )
            dates = Reference(ws, min_col=1, min_row=3, max_row=len(complete_data) + 2)

            # 添加数据到图表
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(dates)

            # 应用图表样式
            if chart_style:
                marker_style = chart_style.get("marker_style", "circle")
                marker_size = chart_style.get("marker_size", 8)
                line_color = chart_style.get("line_color", "0000FF")
                line_width = chart_style.get("line_width", 2.5)

                series = chart.series[0]
                series.graphicalProperties = GraphicalProperties()
                series.graphicalProperties.line = LineProperties(
                    w=line_width * 12700, solidFill=line_color
                )
                series.marker = Marker(symbol=marker_style, size=marker_size)
            else:
                # 默认样式
                chart.series[0].marker = Marker(symbol="circle", size=8)

            # 恢复图表到初始大小
            chart.width = 30
            chart.height = 15

            # 添加图表到工作表，从E5开始绘制
            ws.add_chart(chart, "E5")

            try:
                wb.save(output_file_path)
                print(f"  库存余量图表已生成并保存为{export_format.upper()}格式")
            except PermissionError:
                print(
                    f"错误：无法保存文件 '{output_file_path}'，可能是文件正在被其他程序占用。"
                )
                print("请关闭相关文件后重试。")
                raise
            finally:
                # 确保工作簿被关闭
                if wb is not None:
                    try:
                        wb.close()
                    except:
                        pass
        except Exception as e:
            print("发生错误：")
            print(str(e))
            raise
        # finally 块已移至try-save-block内部，统一管理关闭逻辑

    def _validate_inventory_value(self, value):
        """
        验证库存值是否有效
        - 允许大于100的值
        - 不允许小于0的值
        - 不允许非数字值
        """
        try:
            float_value = float(value)
            if float_value < 0:
                raise ValueError("库存值不能为负数")
            return float_value
        except (ValueError, TypeError):
            raise ValueError(f"无效的库存值: {value}")
