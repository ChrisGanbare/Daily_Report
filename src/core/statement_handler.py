"""
客户对账单处理器 - 重新设计实现，确保与development-copy分支功能完全一致
"""

import logging
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import tempfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.services.customer_statement_service import CustomerStatementService

logger = logging.getLogger(__name__)


class StatementHandler:
    """客户对账单处理器 - 重新设计实现"""
    
    def __init__(self, template_path: Optional[str] = None):
        """
        初始化客户对账单处理器
        
        Args:
            template_path: Excel模板文件路径
        """
        self.template_path = template_path
        self.customer_statement_service = CustomerStatementService()
    
    def generate_customer_statement_from_template(self,
                                                 customer_data: Dict[str, Any],
                                                 devices_data: List[Dict[str, Any]],
                                                 output_dir: str,
                                                 start_date: str,
                                                 end_date: str) -> str:
        """
        从模板生成客户对账单 - 使用新的CustomerStatementService
        
        Args:
            customer_data: 客户基础信息
            devices_data: 设备数据列表
            output_dir: 输出目录
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            生成的Excel文件路径
        """
        logger.info(f"开始生成客户对账单: {customer_data.get('customer_name', '未知客户')}")
        
        try:
            # 验证日期范围（不超过1个月）
            self._validate_date_span(start_date, end_date)
            
            # 验证设备数据
            self._validate_devices_data(devices_data)
            
            # 提取设备编码列表
            device_codes = [device['device_code'] for device in devices_data]
            
            # 生成对账单
            output_path, warnings = self.customer_statement_service.generate_customer_statement(
                device_codes=device_codes,
                start_date=start_date,
                end_date=end_date,
                output_dir=output_dir
            )
            
            # 记录警告信息
            for warning in warnings:
                logger.warning(warning)
            
            logger.info(f"客户对账单生成成功: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"生成客户对账单失败: {e}", exc_info=True)
            raise
    
    def _validate_date_span(self, start_date: str, end_date: str) -> None:
        """
        验证日期范围不超过1个月
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Raises:
            ValueError: 日期范围超过1个月
        """
        from src.utils.date_utils import validate_date_span
        
        # 使用date_utils中的验证逻辑
        date_row = {"start_date": start_date, "end_date": end_date}
        if not validate_date_span(date_row):
            raise ValueError("客户对账单的日期范围不能超过1个月")
    
    def _validate_devices_data(self, devices_data: List[Dict[str, Any]]) -> None:
        """
        验证设备数据
        
        Args:
            devices_data: 设备数据列表
            
        Raises:
            ValueError: 设备数据验证失败
        """
        if not devices_data:
            raise ValueError("设备数据不能为空")
        
        required_fields = ['device_code', 'device_name', 'oil_name']
        for i, device in enumerate(devices_data):
            for field in required_fields:
                if field not in device or not device[field]:
                    raise ValueError(f"设备数据第{i+1}行缺少必要字段: {field}")
    
    def _update_homepage_sheet_minimal(self, 
                                      wb: Workbook, 
                                      customer_data: Dict[str, Any],
                                      devices_data: List[Dict[str, Any]],
                                      start_date: str,
                                      end_date: str) -> None:
        """
        最小化更新主页工作表 - 基于development-copy分支的核心逻辑
        
        Args:
            wb: Excel工作簿对象
            customer_data: 客户基础信息
            devices_data: 设备数据列表
            start_date: 开始日期
            end_date: 结束日期
        """
        # 获取或创建主页工作表
        if "主页" in wb.sheetnames:
            ws = wb["主页"]
            # 清空现有内容但保留格式
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.active
            ws.title = "主页"
        
        # 设置客户名称和日期范围
        customer_name = customer_data.get('customer_name', '未知客户')
        
        # 主标题
        ws.append(["客户对账单"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # 客户信息
        ws.append(["客户名称:", customer_name])
        ws.append(["日期范围:", f"{start_date} 至 {end_date}"])
        ws.append(["生成时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])  # 空行
        
        # 清理设备数据（移除空值）
        cleaned_devices_data = []
        for device in devices_data:
            cleaned_device = {}
            for key, value in device.items():
                if value is not None and value != "":
                    cleaned_device[key] = value
            cleaned_devices_data.append(cleaned_device)
        
        # 检查设备数量限制
        if len(cleaned_devices_data) > 100:
            logger.warning(f"设备数量超过100台限制: {len(cleaned_devices_data)}")
        
        # 设备数据表格标题
        headers = ["设备编号", "设备名称", "油品名称", "期初库存(L)", "期末库存(L)", 
                  "加油量(L)", "订单量(L)", "实际消耗(L)", "误差(L)"]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            header_cell = ws.cell(row=ws.max_row, column=col)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # 写入设备数据
        for device in cleaned_devices_data:
            row_data = [
                device.get('device_code', ''),
                device.get('device_name', ''),
                device.get('oil_name', ''),
                device.get('beginning_inventory', 0),
                device.get('ending_inventory', 0),
                device.get('refill_volume', 0),
                device.get('order_volume', 0),
                device.get('actual_consumption', 0),
                device.get('error', 0)
            ]
            ws.append(row_data)
        
        # 添加备注信息
        self._add_notes_to_homepage(ws, len(cleaned_devices_data))
        
        # 调整格式
        self._adjust_homepage_formatting(ws)
    
    def _add_notes_to_homepage(self, ws, device_count: int) -> None:
        """添加备注信息到主页"""
        # 空行
        ws.append([])
        ws.append([])
        
        # 备注标题
        ws.append(["备注:"])
        note_title_cell = ws.cell(row=ws.max_row, column=1)
        note_title_cell.font = Font(bold=True)
        
        # 备注内容
        notes = [
            f"1. 本对账单包含 {device_count} 台设备的数据",
            "2. 实际消耗 = (期初库存 - 期末库存 + 加油量)",
            "3. 误差 = 实际消耗 - 订单量",
            "4. 正误差表示中润亏损，负误差表示客户亏损",
            "5. 数据来源: 设备监控系统"
        ]
        
        for note in notes:
            ws.append([note])
    
    def _adjust_homepage_formatting(self, ws) -> None:
        """调整主页格式"""
        # 设置列宽
        column_widths = {
            'A': 12,  # 设备编号
            'B': 15,  # 设备名称
            'C': 15,  # 油品名称
            'D': 12,  # 期初库存
            'E': 12,  # 期末库存
            'F': 10,  # 加油量
            'G': 10,  # 订单量
            'H': 12,  # 实际消耗
            'I': 10   # 误差
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置数字格式
        for row in range(6, ws.max_row + 1):  # 从数据行开始
            for col in range(4, 10):  # D-I列（数字列）
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
    
    def _update_monthly_comparison_sheet(self, 
                                       wb: Workbook, 
                                       devices_data: List[Dict[str, Any]],
                                       start_date: str,
                                       end_date: str) -> None:
        """
        更新月度对比工作表 - 基于development-copy分支的核心逻辑
        
        Args:
            wb: Excel工作簿对象
            devices_data: 设备数据列表
            start_date: 开始日期
            end_date: 结束日期
        """
        # 创建或获取月度对比工作表
        if "月度对比" in wb.sheetnames:
            ws = wb["月度对比"]
            ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet("月度对比")
        
        # 设置标题
        ws.append(["月度消耗对比分析"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # 日期范围
        ws.append([f"日期范围: {start_date} 至 {end_date}"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
        ws.append([])  # 空行
        
        # 收集月度数据
        monthly_data = self._collect_monthly_data(devices_data, start_date, end_date)
        
        # 处理设备油品组合
        device_oil_combinations = self._process_device_oil_combinations(monthly_data)
        
        # 写入月份数据
        self._write_monthly_data_to_sheet(ws, device_oil_combinations, start_date, end_date)
        
        # 设置单元格格式
        self._format_monthly_comparison_cells(ws)
    
    def _collect_monthly_data(self, devices_data: List[Dict[str, Any]], start_date: str, end_date: str) -> List[Dict[str, Any]]:
        """收集月度数据"""
        monthly_data = []
        
        # 简化实现：基于设备数据生成月度数据
        for device in devices_data:
            # 这里可以添加更复杂的月度数据收集逻辑
            monthly_record = {
                'device_code': device.get('device_code'),
                'device_name': device.get('device_name'),
                'oil_name': device.get('oil_name'),
                'month': start_date[:7],  # 取年月部分
                'consumption': device.get('actual_consumption', 0),
                'order_volume': device.get('order_volume', 0),
                'error': device.get('error', 0)
            }
            monthly_data.append(monthly_record)
        
        return monthly_data
    
    def _process_device_oil_combinations(self, monthly_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """处理设备油品组合"""
        combinations = {}
        
        for record in monthly_data:
            key = f"{record['device_code']}_{record['oil_name']}"
            if key not in combinations:
                combinations[key] = {
                    'device_code': record['device_code'],
                    'device_name': record['device_name'],
                    'oil_name': record['oil_name'],
                    'months': {}
                }
            
            month = record['month']
            combinations[key]['months'][month] = {
                'consumption': record['consumption'],
                'order_volume': record['order_volume'],
                'error': record['error']
            }
        
        return combinations
    
    def _write_monthly_data_to_sheet(self, ws, device_oil_combinations: Dict[str, Any], start_date: str, end_date: str) -> None:
        """将月度数据写入工作表"""
        # 生成月份列表
        months = self._generate_month_list(start_date, end_date)
        
        # 表头
        headers = ["设备编号", "设备名称", "油品名称"] + months + ["平均消耗", "总误差"]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            header_cell = ws.cell(row=ws.max_row, column=col)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        # 写入数据行
        for combo_key, combo_data in device_oil_combinations.items():
            row_data = [
                combo_data['device_code'],
                combo_data['device_name'],
                combo_data['oil_name']
            ]
            
            # 添加月度数据
            total_consumption = 0
            total_error = 0
            month_count = 0
            
            for month in months:
                month_data = combo_data['months'].get(month, {})
                consumption = month_data.get('consumption', 0)
                error = month_data.get('error', 0)
                
                row_data.append(consumption)
                total_consumption += consumption
                total_error += error
                if consumption > 0:
                    month_count += 1
            
            # 计算平均值
            avg_consumption = total_consumption / month_count if month_count > 0 else 0
            row_data.extend([avg_consumption, total_error])
            
            ws.append(row_data)
    
    def _generate_month_list(self, start_date: str, end_date: str) -> List[str]:
        """生成月份列表"""
        # 简化实现：返回开始日期所在的月份
        return [start_date[:7]]  # 格式: YYYY-MM
    
    def _format_monthly_comparison_cells(self, ws) -> None:
        """设置月度对比单元格格式"""
        # 设置列宽
        column_widths = {'A': 12, 'B': 15, 'C': 15, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 12}
        
        for col, width in column_widths.items():
            if col in ws.column_dimensions:
                ws.column_dimensions[col].width = width
        
        # 设置数字格式
        for row in range(4, ws.max_row + 1):  # 从数据行开始
            for col in range(4, ws.max_column + 1):  # 从第4列开始（月份数据）
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'