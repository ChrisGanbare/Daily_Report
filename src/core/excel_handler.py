import os
import csv
from collections import defaultdict
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from decimal import Decimal


class ExcelHandler:
    """Excel处理类，负责所有Excel相关操作"""

    def __init__(self):
        """初始化Excel处理器"""
        pass

    def generate_excel_with_chart(self, data, output_file, device_code, start_date, end_date,
                                  chart_style=None, export_format='xlsx'):
        """
        根据数据生成折线图并保存为Excel文件
        
        Args:
            data (list): 库存数据列表
            output_file (str): 输出文件路径
            device_code (str): 设备编码
            start_date (date): 开始日期
            end_date (date): 结束日期
            chart_style (dict): 图表样式配置
            export_format (str): 导出格式
        """
        try:
            # 验证并清理数据
            cleaned_data = []
            invalid_records = []
            for date, value in data:
                try:
                    validated_value = self._validate_inventory_value(value)
                    if validated_value > 100:
                        print(f"提示：日期 {date} 的库存值 {validated_value}% 超过100%")
                    cleaned_data.append((date, validated_value))
                except ValueError as e:
                    invalid_records.append((date, value, str(e)))
                    print(f"警告：日期 {date} 的数据已跳过 - {str(e)}")

            if invalid_records:
                print("\n无效数据汇总：")
                for date, value, reason in invalid_records:
                    print(f"- {date}: {value} ({reason})")

            # 如果没有有效数据，尝试生成一个带有默认值的图表
            if not cleaned_data:
                print("警告：没有有效的库存数据可供处理，将生成默认数据图表")
                # 使用默认数据点，确保能生成图表
                cleaned_data = [(start_date, 0), (end_date, 0)]
                print(f"使用默认数据点: {cleaned_data}")

            # 处理不同导出格式
            if export_format.lower() == 'csv':
                with open(output_file.replace('.xlsx', '.csv'), 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['日期', '库存百分比'])
                    writer.writerows(cleaned_data)
                print(f"数据已导出为CSV格式：{output_file.replace('.xlsx', '.csv')}")
                return

            # 补全数据
            data_dict = dict(cleaned_data)
            complete_data = []
            current_date = start_date
            last_inventory = next(iter(cleaned_data))[1] if cleaned_data else 0

            while current_date <= end_date:
                current_inventory = data_dict.get(current_date, last_inventory)
                complete_data.append([current_date, current_inventory])
                last_inventory = current_inventory
                current_date += timedelta(days=1)

            # Excel处理
            wb = Workbook()
            ws = wb.active
            ws.title = "库存数据"

            # 添加标题行
            title = f"{device_code}每日库存余量变化趋势({start_date} - {end_date})"
            ws.append([title])
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            # 添加数据列标题
            ws.append(["日期", "库存百分比"])

            # 写入补全后的数据
            for row in complete_data:
                ws.append(row)

            # 创建图表
            chart = LineChart()
            chart.title = "每日库存余量变化趋势"
            chart.style = 13
            chart.y_axis.title = "库存百分比"
            chart.x_axis.title = "日期"

            # 添加这部分配置来调整x轴日期显示
            chart.x_axis.tickLblSkip = 3  # 每隔3个标签显示一个
            chart.x_axis.tickLblPos = "low"  # 将标签位置调整到底部
            chart.x_axis.textRotation = 0  # 将文本旋转角度设为0度（水平显示）

            # 设置数据范围
            data_range = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=len(complete_data) + 2)
            dates = Reference(ws, min_col=1, min_row=3, max_row=len(complete_data) + 2)

            # 添加数据到图表
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(dates)

            # 应用图表样式
            if chart_style:
                marker_style = chart_style.get('marker_style', 'circle')
                marker_size = chart_style.get('marker_size', 8)
                line_color = chart_style.get('line_color', '0000FF')
                line_width = chart_style.get('line_width', 2.5)

                series = chart.series[0]
                series.graphicalProperties = GraphicalProperties()
                series.graphicalProperties.line = LineProperties(w=line_width * 12700, solidFill=line_color)
                series.marker = Marker(symbol=marker_style, size=marker_size)
            else:
                # 默认样式
                chart.series[0].marker = Marker(symbol='circle', size=8)

            # 添加图表到工作表
            ws.add_chart(chart, "E5")

            try:
                wb.save(output_file)
                print(f"  库存余量图表已生成并保存为{export_format.upper()}格式")
            except PermissionError:
                print(f"错误：无法保存文件 '{output_file}'，可能是文件正在被其他程序占用。")
                print("请关闭相关文件后重试。")
                raise
        except Exception as e:
            print("发生错误：")
            print(str(e))
            raise

    def generate_enhanced_excel(self, all_devices_data, output_file, customer_name, start_date, end_date):
        """
        生成增强版Excel报表

        Args:
            all_devices_data: 所有设备的数据 [{device_code, oil_name, data, raw_data, columns}, ...]
            output_file: 输出文件路径
            customer_name: 客户名称
            start_date: 开始日期
            end_date: 结束日期
        """
        try:
            # 创建工作簿
            wb = Workbook()
            wb.remove(wb.active)  # 删除默认工作表

            # 按设备分组处理数据
            device_groups = defaultdict(list)
            for device_data in all_devices_data:
                device_groups[device_data['device_code']].append(device_data)

            # 调用各个工作表更新函数
            for sheet_name in wb.sheetnames:
                if sheet_name == "中润对账单":
                    self._update_statement_sheet(wb[sheet_name], all_devices_data, customer_name, start_date, end_date)
                elif sheet_name == "订单明细":
                    self._update_order_details_sheet(wb[sheet_name], all_devices_data)
                elif sheet_name == "每日用量明细":
                    self._update_daily_usage_sheet(wb[sheet_name], all_devices_data, start_date, end_date)
                elif sheet_name == "每月用量对比":
                    self._update_monthly_comparison_sheet(wb[sheet_name], all_devices_data, start_date, end_date)

            try:
                wb.save(output_file)
                print(f"增强版Excel报表已生成并保存为: {output_file}")
            except PermissionError:
                print(f"错误：无法保存文件 '{output_file}'，可能是文件正在被其他程序占用。")
                print("请关闭相关文件后重试。")
                raise

        except Exception as e:
            print("生成增强版Excel报表时发生错误：")
            print(str(e))
            raise

    def generate_enhanced_excel_from_template(self, all_devices_data, output_file, customer_name, start_date, end_date):
        """基于模板生成对账单Excel报表"""
        template_path = os.path.join(os.path.dirname(__file__), 'template', 'statement_template.xlsx')

        # 检查模板目录是否存在
        template_dir = os.path.dirname(template_path)
        if not os.path.exists(template_dir):
            try:
                os.makedirs(template_dir)
            except Exception as e:
                raise FileNotFoundError(f"无法创建模板目录: {template_dir}\n错误: {e}")
            raise FileNotFoundError(f"已创建模板目录，请将模板文件放入: {template_dir}")

        if not os.path.exists(template_path):
            raise FileNotFoundError(
                f"模板文件不存在: {template_path}\n"
                f"请确保文件 'statement_template.xlsx' 位于 '{template_dir}' 目录下"
            )

        try:
            wb = load_workbook(template_path)

            # 修改工作表名称以匹配本地模板
            sheets = {
                "对账单": self._update_statement_sheet,
                "订单记录": self._update_order_details_sheet,
                "日用量": self._update_daily_usage_sheet,
                "月用量": self._update_monthly_comparison_sheet
            }

            missing_sheets = [name for name in sheets if name not in wb.sheetnames]
            if missing_sheets:
                raise KeyError(
                    f"模板缺少以下工作表: {', '.join(missing_sheets)}\n"
                    f"请确保模板包含所有必需的工作表:\n"
                    f"- {chr(10).join(sheets.keys())}"
                )

            # 更新各工作表
            for name, update_func in sheets.items():
                try:
                    update_func(wb[name], all_devices_data, customer_name, start_date, end_date)
                except Exception as e:
                    raise RuntimeError(f"更新工作表 '{name}' 时出错: {e}")

            # 保存结果前处理图表相关警告
            for sheet in wb.worksheets:
                if sheet._charts:
                    for chart in sheet._charts:
                        if hasattr(chart, 'externalData'):
                            chart.externalData = None

            try:
                wb.save(output_file)
                print(f"已生成对账单: {output_file}")
            except PermissionError:
                raise PermissionError(f"无法保存文件，可能被其他程序占用: {output_file}")

        except Exception as e:
            print(f"生成对账单时发生错误: {str(e)}")
            raise

    def _validate_inventory_value(self, value):
        """
        验证库存值是否有效
        - 允许大于100的值
        - 不允许小于0的值
        - 不允许非数字值
        """
        try:
            float_value = float(value)
            if float_value < 0:
                raise ValueError("库存值不能为负数")
            return float_value
        except (ValueError, TypeError):
            raise ValueError(f"无效的库存值: {value}")

    def _update_statement_sheet(self, ws, all_devices_data, customer_name, start_date, end_date):
        """更新对账单工作表"""
        try:
            # 更新客户信息和日期范围
            ws['B2'] = customer_name
            ws['D2'] = f"{start_date.strftime('%Y-%m-%d')}至{end_date.strftime('%Y-%m-%d')}"

        except Exception as e:
            print(f"更新对账单工作表时出错: {e}")
            raise

    def _update_order_details_sheet(self, ws, all_devices_data):
        """更新订单明细工作表"""
        try:
            # 清除现有数据（保留表头）
            for row in range(4, ws.max_row + 1):
                ws.delete_rows(row)

            # 数据写入的起始行（根据模板格式调整）
            current_row = 4

            # 写入所有设备的订单数据
            for device_data in all_devices_data:
                if device_data['raw_data']:
                    for row_data in device_data['raw_data']:
                        for col, value in enumerate(row_data, 1):
                            ws.cell(row=current_row, column=col, value=value)
                        current_row += 1

        except Exception as e:
            print(f"更新订单明细工作表时出错: {e}")
            raise


    def _update_daily_usage_sheet(self, ws, all_devices_data, start_date, end_date):
        """更新每日用量明细工作表"""
        try:
            # 更新标题日期范围
            ws['A1'] = f"每日用量明细({start_date}至{end_date})"

            # 收集每日用量数据，统一使用 Decimal
            daily_usage = defaultdict(lambda: defaultdict(Decimal))
            oil_names = set()

            for device_data in all_devices_data:
                oil_name = device_data['oil_name']
                oil_names.add(oil_name)
                for date, value in device_data['data']:
                    # 将数值统一转换为 Decimal 类型
                    value = Decimal(str(value)) if value is not None else Decimal('0')
                    daily_usage[date][oil_name] += value

            # 写入数据（从第4行开始，保留表头）
            current_row = 4
            sorted_dates = sorted(daily_usage.keys())
            sorted_oils = sorted(oil_names)

            # 更新表头
            for col, oil_name in enumerate(sorted_oils, 2):
                ws.cell(row=3, column=col, value=oil_name)

            # 写入每日数据
            for date in sorted_dates:
                ws.cell(row=current_row, column=1, value=date)
                for col, oil_name in enumerate(sorted_oils, 2):
                    value = daily_usage[date].get(oil_name, Decimal('0'))
                    # 转换为浮点数以便Excel显示
                    ws.cell(row=current_row, column=col, value=float(value.quantize(Decimal('0.01'))))
                current_row += 1

            # 更新图表（如果模板中包含图表）
            if ws._charts:
                chart = ws._charts[0]
                data = Reference(ws, min_col=2, min_row=3, max_col=len(sorted_oils)+1,
                               max_row=current_row-1)
                cats = Reference(ws, min_col=1, min_row=4, max_row=current_row-1)
                chart.set_categories(cats)
                chart.series[0].values = data

        except Exception as e:
            print(f"更新每日用量明细工作表时出错: {e}")
            raise

    def _update_monthly_comparison_sheet(self, ws, all_devices_data, start_date, end_date):
        """更新每月用量对比工作表"""
        try:
            ws['A1'] = f"每月用量对比({start_date}至{end_date})"

            # 收集月度数据
            monthly_stats = defaultdict(lambda: defaultdict(float))
            for device_data in all_devices_data:
                oil_name = device_data['oil_name']
                for date, value in device_data['data']:
                    month = date.strftime('%Y-%m')
                    monthly_stats[month][oil_name] += value

            # 写入数据
            current_row = 4
            for month, oil_stats in sorted(monthly_stats.items()):
                ws.cell(row=current_row, column=1, value=month)
                for col, (oil_name, value) in enumerate(sorted(oil_stats.items()), 2):
                    ws.cell(row=current_row, column=col, value=round(value, 2))
                current_row += 1

        except Exception as e:
            print(f"更新每月用量对比工作表时出错: {e}")
            raise