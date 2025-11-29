import os
from datetime import timedelta, datetime # 引入 datetime 类
from collections import defaultdict

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

from .base_report import BaseReportGenerator


class DailyConsumptionErrorReportGenerator(BaseReportGenerator):
    """每日消耗误差报表生成器，专门负责生成设备每日消耗误差报表和图表。
    
    报表基于油桶的原油剩余量计算真实的消耗量，并与订单累积量进行对比，
    以发现消耗误差。计算公式：
    真实消耗量 = 前一天结束库存 - 当天结束库存 + 当天加油量
    """

    def __init__(self):
        """初始化每日消耗误差报表生成器"""
        super().__init__()

    def generate_report(self, inventory_data, error_data, output_file_path, **kwargs):
        """
        生成每日消耗误差报表的实现方法
        报表基于油桶的原油剩余量计算真实的消耗量，并与订单累积量进行对比，
        以发现消耗误差。计算公式：
        真实消耗量 = 前一天结束库存 - 当天结束库存 + 当天加油量

        Args:
            inventory_data (list): 库存数据列表
            error_data (dict): 误差数据字典
            output_file_path (str): 输出文件路径# D:/Daily_Report/src/core/data_manager.py (错误的代码)
    if sorted_dates:
        first_day_data = sorted(daily_data[sorted_dates[0]], key=lambda x: x['order_time'])
        if first_day_data:
            previous_day_end_inventory = first_day_data[0]['avai_oil'] # 使用第一天的数据初始化“前一天”的库存
            **kwargs: 其他参数
            
        Returns:
            bool: 报表生成是否成功
        """
        device_code = kwargs.get('device_code')
        start_date = kwargs.get('start_date')
        end_date = kwargs.get('end_date')
        oil_name = kwargs.get('oil_name')
        chart_style = kwargs.get('chart_style')
        barrel_count = int(kwargs.get('barrel_count', 1))
        export_format = kwargs.get('export_format', 'xlsx')
        
        return self.generate_daily_consumption_error_report_with_chart(
            inventory_data, error_data, output_file_path, device_code, start_date, end_date,
            oil_name, chart_style, barrel_count, export_format
        )

    def generate_daily_consumption_error_report_with_chart(
        self,
        inventory_data,
        error_data,
        output_file_path,
        device_code,
        start_date,
        end_date,
        oil_name=None,
        chart_style=None,
        barrel_count=1,
        export_format="xlsx",
    ):
        """
        生成包含库存数据和误差分析的Excel报告文件

        报表基于油桶的原油剩余量计算真实的消耗量，并与订单累积量进行对比，
        以发现消耗误差。计算公式：
        真实消耗量 = 前一天结束库存 - 当天结束库存 + 当天加油量

        Args:
            inventory_data (list): 库存数据列表，格式为[(date, value), ...]
            error_data (dict): 误差数据字典
            output_file_path (str): 输出文件路径
            device_code (str): 设备编码
            start_date (date): 开始日期
            end_date (date): 结束日期
            oil_name (str): 油品名称
            chart_style (dict): 图表样式配置
            barrel_count (int): 油桶数量
            export_format (str): 导出格式，支持xlsx
        """
        try:
            # 确保输出文件路径不重复，如果重复则添加序号
            output_file_path = self._get_unique_filename(output_file_path)
            
            # 验证并清理库存数据
            cleaned_inventory_data = []
            invalid_records = []
            for date, value in inventory_data:
                try:
                    validated_value = self._validate_inventory_value(value)
                    cleaned_inventory_data.append((date, validated_value))
                except ValueError as e:
                    invalid_records.append((date, value, str(e)))
                    print(f"警告：日期 {date} 的数据已跳过 - {str(e)}")

            if invalid_records:
                print("\n无效数据汇总：")
                for date, value, reason in invalid_records:
                    print(f"- {date}: {value} ({reason})")

            # 如果没有有效数据，尝试生成一个带有默认值的图表
            if not cleaned_inventory_data:
                print("警告：没有有效的原油剩余量数据可供处理，将生成默认数据图表")
                # 使用默认数据点，确保能生成图表
                cleaned_inventory_data = [(start_date, 0), (end_date, 0)]
                print(f"使用默认数据点: {cleaned_inventory_data}")

            # 初始化工作簿
            wb = Workbook()

            # 补全库存数据
            # 优先使用 error_data 中的 daily_end_inventory，因为它为所有日期生成了数据
            # 包括没有订单的日期，这样可以避免"向前泄漏"的问题
            daily_end_inventory = error_data.get('daily_end_inventory', {})
            complete_inventory_data = []
            current_date = start_date

            # 创建 data_dict 用于 fallback
            data_dict = dict(cleaned_inventory_data)
            
            while current_date <= end_date:
                # 优先使用 error_data 中的 daily_end_inventory
                if current_date in daily_end_inventory:
                    current_inventory = daily_end_inventory[current_date]
                else:
                    # 如果 error_data 中没有该日期的数据，fallback 到 inventory_data
                    if current_date in data_dict:
                        current_inventory = data_dict[current_date]
                    else:
                        # 如果都没有，使用上一个日期的值
                        if complete_inventory_data:
                            current_inventory = complete_inventory_data[-1][1]
                        else:
                            current_inventory = 0
                
                complete_inventory_data.append([current_date, current_inventory])
                current_date += timedelta(days=1)

            # Excel处理
            ws = wb.active
            ws.title = "消耗误差分析"

            # 添加标题行
            oil_name_str = f" {oil_name} " if oil_name else " "
            title = f"{device_code}{oil_name_str}每日消耗误差分析({start_date} - {end_date})"
            ws.append([title])
            # 将合并单元格的宽度增加到18列
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", wrap_text=True)
            ws.cell(row=1, column=1).font = Font(size=14, bold=True)

            # 添加配置信息提示行
            hint_text_base = "提示：【设备桶数】已从配置文件自动读取，未配置的设备默认桶数为1；若设备数据变动，需同步维护test_data/device_config.csv配置文件，保持一次维护、多次复用的准确性；"
            
            # 获取配置信息
            config_info_text = ""
            try:
                from src.core.device_config_manager import DeviceConfigManager
                config_manager = DeviceConfigManager()
                config_info = config_manager.get_config_info()
                
                if config_info['file_exists']:
                    config_path = config_info['config_file_abspath']
                    maintenance_time = config_info.get('last_maintenance_time', '未知')
                    maintenance_type = config_info.get('last_maintenance_type', '')
                    
                    config_info_text = f"配置文件位置: {config_path}   最近维护时间: {maintenance_time}"
                    if maintenance_type:
                        config_info_text += f" ({maintenance_type})"
            except Exception:
                # 如果获取配置信息失败，忽略错误
                pass
            
            # 组合提示文本，如果有配置信息则在单元格内换行显示
            if config_info_text:
                hint_text = hint_text_base + "\n" + config_info_text
            else:
                hint_text = hint_text_base
            ws.append([hint_text])
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
            hint_cell = ws.cell(row=2, column=1)
            hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            hint_cell.font = Font(size=9, color="FF4500", bold=True, italic=False)
            # 设置行高为两行提示文本的紧凑高度（固定值，减少上下空白）
            ws.row_dimensions[2].height = 35

            # 添加数据列标题（现在在第3行）
            ws.append(["日期", "原油剩余量(L)", "订单累积总量(L)", "库存消耗总量(L)","中润亏损(L)", "客户亏损(L)"])

            # 准备误差数据
            daily_order_totals = error_data.get('daily_order_totals', {})  # 每日订单累计总量数据
            daily_shortage_errors = error_data.get('daily_shortage_errors', {})  # 每日中润亏损数据
            daily_excess_errors = error_data.get('daily_excess_errors', {})  # 每日客户亏损数据
            daily_consumption = error_data.get('daily_consumption', {})  # 获取每日消耗量数据

            # 写入补全后的数据
            for row in complete_inventory_data:
                date = row[0]
                inventory_value = row[1]

                order_total = daily_order_totals.get(date, 0)
                # 处理可能为字典格式的误差数据
                shortage_data = daily_shortage_errors.get(date, 0)
                excess_data = daily_excess_errors.get(date, 0)
                consumption_data = daily_consumption.get(date, 0)  # 获取每日消耗量
                
                # 如果是字典格式，提取value字段
                if isinstance(shortage_data, dict):
                    shortage_error = shortage_data.get('value', 0)
                else:
                    shortage_error = shortage_data
                    
                if isinstance(excess_data, dict):
                    excess_error = excess_data.get('value', 0)
                else:
                    excess_error = excess_data
                    
                if isinstance(consumption_data, dict):
                    consumption_value = consumption_data.get('value', 0)
                else:
                    consumption_value = consumption_data
                    
                ws.append([date, inventory_value, order_total, consumption_value, shortage_error, excess_error])

            # 调整列宽
            ws.column_dimensions["A"].width = 12  # 日期列宽度
            ws.column_dimensions["B"].width = 15  # 原油剩余量列宽度
            ws.column_dimensions["C"].width = 15  # 订单累积总量列宽度
            ws.column_dimensions["D"].width = 15  # 库存消耗总量列宽度
            ws.column_dimensions["E"].width = 12  # 中润亏损列宽度
            ws.column_dimensions["F"].width = 12  # 客户亏损列宽度

            # 创建图表
            chart = LineChart()
            chart.title = "每日消耗误差分析"
            chart.style = 13
            chart.y_axis.title = "值 (L)"
            chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(noFill=True))
            chart.x_axis.title = "日期"
            chart.x_axis.number_format = 'yyyy-mm-dd'

            # 设置图表显示数据标签
            chart.x_axis.tickLblSkip = 3  # 每隔3个标签显示一个
            chart.x_axis.tickLblPos = "low"  # 将标签位置调整到底部
            chart.x_axis.textRotation = 0  # 将文本旋转角度设为0度（水平显示）

            # 设置数据范围（数据列标题在第3行，数据从第4行开始）
            dates = Reference(ws, min_col=1, min_row=4, max_row=len(complete_inventory_data) + 3)
            data_range = Reference(ws, min_col=2, min_row=3, max_col=4, max_row=len(complete_inventory_data) + 3)

            # 添加数据到图表
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(dates)

            # 为不同数据系列设置不同的颜色
            # 原油剩余量 - 蓝色
            chart.series[0].graphicalProperties = GraphicalProperties()
            chart.series[0].graphicalProperties.line = LineProperties(w=2.5 * 12700, solidFill="0000FF")
            chart.series[0].marker = Marker(symbol="circle", size=8)
            
            # 订单累积总量 - 绿色
            chart.series[1].graphicalProperties = GraphicalProperties()
            chart.series[1].graphicalProperties.line = LineProperties(w=2.5 * 12700, solidFill="00FF00")
            chart.series[1].marker = Marker(symbol="circle", size=8)
            
            # 库存消耗总量 - 紫色
            chart.series[2].graphicalProperties = GraphicalProperties()
            chart.series[2].graphicalProperties.line = LineProperties(w=2.5 * 12700, solidFill="800080")
            chart.series[2].marker = Marker(symbol="circle", size=8)

            # 恢复图表到初始大小
            chart.width = 30
            chart.height = 15

            # 添加图表到工作表，从G5开始绘制
            ws.add_chart(chart, "G5")
            
            # 在H34单元格开始添加计算规则说明，确保位于图表下方
            annotation_row = 34
            annotation_col = 8  # H列
            
            # 添加标题
            ws.cell(row=annotation_row, column=annotation_col).value = "计算规则说明："
            ws.cell(row=annotation_row, column=annotation_col).font = Font(bold=True)
            
            # 将每个计算规则分别写入单独的行，并合并单元格以确保完整显示
            annotation_row += 1
            consumption_formula = "库存消耗总量(L) = (前日库存 - 当日库存 + 当日加油（入库）量)"
            if barrel_count > 1:
                consumption_formula += f" * {barrel_count} (桶数)"
            ws.cell(row=annotation_row, column=annotation_col).value = consumption_formula
            ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7) # 合并8个单元格
            
            annotation_row += 1
            ws.cell(row=annotation_row, column=annotation_col).value = "中润亏损(L) = MAX(0, 库存消耗总量 - 订单累积总量)"
            ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)
            
            annotation_row += 1
            ws.cell(row=annotation_row, column=annotation_col).value = "客户亏损(L) = MAX(0, 订单累积总量 - 库存消耗总量)"
            ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)
            
        except Exception as e:
            print("发生错误：")
            print(str(e))
            import traceback
            traceback.print_exc()
            raise

        # 将保存和关闭操作移到主try块之外，以确保文件句柄被正确释放
        try:
            wb.save(output_file_path)
            print(f"  每日消耗误差图表已生成并保存为XLSX格式")
            return True
        except PermissionError:
            print(f"错误：无法保存文件 '{output_file_path}'，可能是文件正在被其他程序占用。")
            print("请关闭相关文件后重试。")
            raise
        finally:
            # 确保工作簿总能被关闭
            if 'wb' in locals() and wb is not None:
                try:
                    wb.close()
                except Exception as close_exc:
                    print(f"关闭工作簿时发生错误: {close_exc}")

    def _validate_inventory_value(self, value):
        """
        验证原油剩余量值是否有效
        - 不允许小于0的值
        - 不允许非数字值
        """
        try:
            float_value = float(value)
            if float_value < 0:
                raise ValueError("原油剩余量不能为负数")
            return float_value
        except (ValueError, TypeError):
            raise ValueError(f"无效的原油剩余量值: {value}")



class MonthlyConsumptionErrorReportGenerator(BaseReportGenerator):
    """每月消耗误差报表生成器，专门负责生成设备每月消耗误差报表和图表。
    
    报表基于油桶的原油剩余量计算真实的消耗量，并与订单累积量进行对比，
    以发现消耗误差。计算公式：
    真实消耗量 = 前一月结束库存 - 当月结束库存 + 当月加油量
    """

    def __init__(self):
        """初始化每月消耗误差报表生成器"""
        super().__init__()

    def generate_report(self, inventory_data, error_data, output_file_path, **kwargs):
        """
        生成每月消耗误差报表的实现方法
        
        报表基于油桶的原油剩余量计算真实的消耗量，并与订单累积量进行对比，
        以发现消耗误差。计算公式：
        真实消耗量 = 前一月结束库存 - 当月结束库存 + 当月加油量

        Args:
            inventory_data (list): 库存数据列表
            error_data (dict): 误差数据字典
            output_file_path (str): 输出文件路径
            **kwargs: 其他参数
            
        Returns:
            bool: 报表生成是否成功
        """
        device_code = kwargs.get('device_code')
        start_date = kwargs.get('start_date')
        end_date = kwargs.get('end_date')
        oil_name = kwargs.get('oil_name')
        chart_style = kwargs.get('chart_style')
        barrel_count = int(kwargs.get('barrel_count', 1))
        export_format = kwargs.get('export_format', 'xlsx')
        
        return self.generate_monthly_consumption_error_report_with_chart(
            inventory_data, error_data, output_file_path, device_code, start_date, end_date,
            oil_name, chart_style, barrel_count, export_format
        )

    def generate_monthly_consumption_error_report_with_chart(
        self,
        inventory_data,
        error_data,
        output_file_path,
        device_code,
        start_date,
        end_date,
        oil_name=None,
        chart_style=None,
        barrel_count=1,
        export_format="xlsx",
    ):
        """
        生成包含库存数据和误差分析的Excel报告文件
        
        报表基于油桶的原油剩余量计算真实的消耗量，并与订单累积量进行对比，
        以发现消耗误差。计算公式：
        真实消耗量 = 前一月结束库存 - 当月结束库存 + 当月加油量

        Args:
            inventory_data (list): 库存数据列表，格式为[(date, value), ...]
            error_data (dict): 误差数据字典
            output_file_path (str): 输出文件路径
            device_code (str): 设备编码
            start_date (date): 开始日期
            end_date (date): 结束日期
            oil_name (str): 油品名称
            chart_style (dict): 图表样式配置
            barrel_count (int): 油桶数量
            export_format (str): 导出格式，支持xlsx
        """
        try:
            # 确保输出文件路径不重复，如果重复则添加序号
            output_file_path = self._get_unique_filename(output_file_path)
            
            # 验证并清理库存数据
            cleaned_inventory_data = []
            invalid_records = []
            for date, value in inventory_data:
                try:
                    validated_value = self._validate_inventory_value(value)
                    cleaned_inventory_data.append((date, validated_value))
                except ValueError as e:
                    invalid_records.append((date, value, str(e)))
                    print(f"警告：日期 {date} 的数据已跳过 - {str(e)}")

            if invalid_records:
                print("\n无效数据汇总：")
                for date, value, reason in invalid_records:
                    print(f"- {date}: {value} ({reason})")

            # 准备月度数据
            complete_inventory_data = []
            
            # 直接使用error_data中的月份数据，而不是生成日期范围
            # 这样可以确保日期格式完全匹配
            monthly_order_totals = error_data.get('monthly_order_totals', {})
            monthly_shortage_errors = error_data.get('monthly_shortage_errors', {})
            monthly_excess_errors = error_data.get('monthly_excess_errors', {})
            monthly_consumption = error_data.get('monthly_consumption', {})
            monthly_end_inventory = error_data.get('monthly_end_inventory', {})
            
            # 获取所有唯一的月份标签并排序
            all_months = set()
            all_months.update(monthly_order_totals.keys())
            all_months.update(monthly_shortage_errors.keys())
            all_months.update(monthly_excess_errors.keys())
            all_months.update(monthly_consumption.keys())
            all_months.update(monthly_end_inventory.keys())
            
            # 排序月份
            sorted_months = sorted(list(all_months))
            
            # 限制只显示最多12个月的数据
            if len(sorted_months) > 12:
                sorted_months = sorted_months[-12:]
            
            # 为每个唯一月份创建数据点
            for month in sorted_months:
                complete_inventory_data.append([month, 0])

            # Excel处理
            wb = Workbook()
            ws = wb.active
            ws.title = "消耗误差分析"

            # 添加标题行
            oil_name_str = f" {oil_name} " if oil_name else " "
            title = f"{device_code}{oil_name_str}每月消耗误差分析({start_date} - {end_date})"
            ws.append([title])
            # 将合并单元格的宽度增加到18列
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", wrap_text=True)  # 修复换行参数
            ws.cell(row=1, column=1).font = Font(size=14, bold=True)

            # 添加配置信息提示行
            hint_text_base = "提示：【设备桶数】已从配置文件自动读取，未配置的设备默认桶数为1；若设备数据变动，需同步维护test_data/device_config.csv配置文件，保持一次维护、多次复用的准确性；"
            
            # 获取配置信息
            config_info_text = ""
            try:
                from src.core.device_config_manager import DeviceConfigManager
                config_manager = DeviceConfigManager()
                config_info = config_manager.get_config_info()
                
                if config_info['file_exists']:
                    config_path = config_info['config_file_abspath']
                    maintenance_time = config_info.get('last_maintenance_time', '未知')
                    maintenance_type = config_info.get('last_maintenance_type', '')
                    
                    config_info_text = f"配置文件位置: {config_path}   最近维护时间: {maintenance_time}"
                    if maintenance_type:
                        config_info_text += f" ({maintenance_type})"
            except Exception:
                # 如果获取配置信息失败，忽略错误
                pass
            
            # 组合提示文本，如果有配置信息则在单元格内换行显示
            if config_info_text:
                hint_text = hint_text_base + "\n" + config_info_text
            else:
                hint_text = hint_text_base
            ws.append([hint_text])
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
            hint_cell = ws.cell(row=2, column=1)
            hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            hint_cell.font = Font(size=9, color="FF4500", bold=True, italic=False)
            # 设置行高为两行提示文本的紧凑高度（固定值，减少上下空白）
            ws.row_dimensions[2].height = 35

            # 添加数据列标题（现在在第3行）
            ws.append(["月份", "订单累积总量(L)", "库存消耗总量(L)","中润亏损(L)", "客户亏损(L)"])

            # 写入补全后的数据
            for row in complete_inventory_data:
                month_str = row[0]
                
                # 从误差数据中获取实际值，如果不存在则使用默认值0
                order_total = monthly_order_totals.get(month_str, 0)
                
                # 处理可能为字典格式的误差数据
                shortage_data = monthly_shortage_errors.get(month_str, 0)
                excess_data = monthly_excess_errors.get(month_str, 0)
                consumption_data = monthly_consumption.get(month_str, 0)
                
                # 如果是字典格式，提取value字段
                if isinstance(shortage_data, dict):
                    shortage_error = shortage_data.get('value', 0)
                else:
                    shortage_error = shortage_data
                    
                if isinstance(excess_data, dict):
                    excess_error = excess_data.get('value', 0)
                else:
                    excess_error = excess_data
                    
                if isinstance(consumption_data, dict):
                    consumption_value = consumption_data.get('value', 0)
                else:
                    consumption_value = consumption_data
                    
                ws.append([month_str, order_total, consumption_value, shortage_error, excess_error])

            # 调整列宽
            ws.column_dimensions["A"].width = 12  # 月份列宽度
            ws.column_dimensions["B"].width = 15  # 订单累积总量列宽度
            ws.column_dimensions["C"].width = 15  # 库存消耗总量列宽度
            ws.column_dimensions["D"].width = 12  # 中润亏损列宽度
            ws.column_dimensions["E"].width = 12  # 客户亏损列宽度

            # 创建图表
            chart = LineChart()
            chart.title = "每月消耗误差分析"
            chart.style = 13
            chart.y_axis.title = "值(L)"
            chart.y_axis.titleLayout = None  # 垂直显示Y轴标题
            chart.x_axis.title = "月份"

            # 设置图表显示数据标签
            # 添加这部分配置来调整x轴日期显示
            chart.x_axis.tickLblSkip = 1  # 每个标签都显示
            chart.x_axis.tickLblPos = "low"  # 将标签位置调整到底部
            chart.x_axis.textRotation = 0  # 将文本旋转角度设为0度（水平显示）

            # 设置数据范围（数据列标题在第3行，数据从第4行开始）
            # 包含"订单累积总量"和"库存消耗总量"两列
            data_range = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=len(complete_inventory_data) + 3)
            months = Reference(ws, min_col=1, min_row=4, max_row=len(complete_inventory_data) + 3)

            # 添加数据到图表
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(months)

            # 为不同数据系列设置不同的颜色
            # 订单累积总量 - 绿色
            chart.series[0].graphicalProperties = GraphicalProperties()
            chart.series[0].graphicalProperties.line = LineProperties(w=2.5 * 12700, solidFill="00FF00")
            chart.series[0].marker = Marker(symbol="circle", size=8)
            
            # 库存消耗总量 - 紫色
            chart.series[1].graphicalProperties = GraphicalProperties()
            chart.series[1].graphicalProperties.line = LineProperties(w=2.5 * 12700, solidFill="800080")
            chart.series[1].marker = Marker(symbol="circle", size=8)

            # 恢复图表到初始大小
            chart.width = 30
            chart.height = 15

            # 添加图表到工作表，从F5开始绘制
            ws.add_chart(chart, "F5")
            
            # 在H34单元格开始添加计算规则说明，确保位于图表下方
            annotation_row = 34
            annotation_col = 8  # H列
            
            # 添加标题
            ws.cell(row=annotation_row, column=annotation_col).value = "计算规则说明："
            ws.cell(row=annotation_row, column=annotation_col).font = Font(bold=True)
            
            # 将每个计算规则分别写入单独的行，并合并单元格以确保完整显示
            annotation_row += 1
            consumption_formula = "库存消耗总量(L) = (前月库存 - 当月库存 + 当月加油（入库）量)"
            if barrel_count > 1:
                consumption_formula += f" * {barrel_count} (桶数)"
            ws.cell(row=annotation_row, column=annotation_col).value = consumption_formula
            ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)  # 合并8个单元格
            
            annotation_row += 1
            ws.cell(row=annotation_row, column=annotation_col).value = "中润亏损(L) = MAX(0, 库存消耗总量 - 订单累积总量)"
            ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)
            
            annotation_row += 1
            ws.cell(row=annotation_row, column=annotation_col).value = "客户亏损(L) = MAX(0, 订单累积总量 - 库存消耗总量)"
            ws.merge_cells(start_row=annotation_row, start_column=annotation_col, end_row=annotation_row, end_column=annotation_col + 7)

        except Exception as e:
            print("发生错误：")
            print(str(e))
            import traceback
            traceback.print_exc()
            raise

        # 将保存和关闭操作移到主try块之外，确保文件句柄被正确释放
        try:
            wb.save(output_file_path)
            print(f"  每月消耗误差图表已生成并保存为XLSX格式")
        except PermissionError:
            print(f"错误：无法保存文件 '{output_file_path}'，可能是文件正在被其他程序占用。")
            print("请关闭相关文件后重试。")
            raise
        finally:
            # 确保工作簿总能被关闭
            if 'wb' in locals() and wb is not None:
                try:
                    wb.close()
                    # 添加短暂延迟确保文件被系统完全释放（Windows系统需要）
                    import time
                    time.sleep(0.1)
                except Exception as close_exc:
                    print(f"关闭工作簿时发生错误: {close_exc}")
        
        return True


    def _validate_inventory_value(self, value):
        """
        验证原油剩余量值是否有效
        - 不允许小于0的值
        - 不允许非数字值
        """
        try:
            float_value = float(value)
            if float_value < 0:
                raise ValueError("原油剩余量不能为负数")
            return float_value
        except (ValueError, TypeError):
            raise ValueError(f"无效的原油剩余量值: {value}")


class ConsumptionErrorSummaryGenerator(BaseReportGenerator):
    """
    消耗误差汇总报表生成器。
    负责查询指定日期范围内所有存在误差的设备，并生成一份包含Excel公式的汇总报表。
    """

    def __init__(self):
        """初始化消耗误差汇总报表生成器"""
        super().__init__()

    def generate_report(self, summary_data, output_file_path, **kwargs):
        """
        生成消耗误差汇总报表。

        Args:
            summary_data (list): 从数据库直接查询出的汇总数据列表。
            output_file_path (str): 输出文件路径。
            **kwargs: 其他参数，如 start_date, end_date。
        """
        start_date = kwargs.get('start_date')
        end_date = kwargs.get('end_date')

        # --- 定义样式 ---
        # 标题字体
        title_font = Font(size=16, bold=True, name='Calibri')
        # 表头字体和填充
        header_font = Font(bold=True, color="FFFFFF", name='Calibri')
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        # 数据行交替填充
        even_row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        # 提示信息字体
        hint_font = Font(size=9, color="FF4500", name='Calibri', bold=True, italic=False)
        # 误差解读提示字体
        explanation_font = Font(size=9, color="0070C0", name='Calibri', bold=True)
        # 边框样式
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # 警告样式
        warning_font = Font(color="9C0006")
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        try:
            # 确保输出文件路径不重复
            output_file_path = self._get_unique_filename(output_file_path)

            wb = Workbook()
            ws = wb.active
            ws.title = "设备误差汇总" # 主Sheet

            # 添加数据列标题
            headers = [ "序号",
                "设备编码", "客户名称", "设备桶数", "订单总量(L)",
                "[单桶]库存消耗(L)", "库存消耗总量(L)", "误差值总数(L)", "平均每日误差(L)", "误差百分比(%)",
                "累计离线时长(小时)", "备注"
            ]

            # --- 调整：先写入所有行内容，再设置格式 ---
            # 1. 准备所有行内容
            title_text = f"安卓设备消耗误差汇总报表 ({start_date} - {end_date})"
            hint_text_1_base = "1. 提示：D列【设备桶数】已从配置文件自动读取，未配置的设备默认桶数为1；若设备数据变动，需同步维护test_data/device_config.csv配置文件，保持一次维护、多次复用的准确性；"
            
            # 添加配置信息到第一行提示文本末尾
            config_info_text = ""
            try:
                from src.core.device_config_manager import DeviceConfigManager
                config_manager = DeviceConfigManager()
                config_info = config_manager.get_config_info()
                
                if config_info['file_exists']:
                    config_path = config_info['config_file_abspath']
                    maintenance_time = config_info.get('last_maintenance_time', '未知')
                    maintenance_type = config_info.get('last_maintenance_type', '')
                    
                    # 合并到第一行，使用多个空格分隔，保持视觉美观
                    config_info_text = f"配置文件位置: {config_path}   最近维护时间: {maintenance_time}；"
                    if maintenance_type:
                        config_info_text += f" ({maintenance_type})"
            except Exception:
                # 如果获取配置信息失败，忽略错误
                pass
            
            # 将配置信息合并到第一行提示文本
            hint_text_1 = hint_text_1_base + config_info_text
            
            explanation_text = "2. 误差百分比解读：正数(%)表示`库存消耗 > 订单总量`，可能为公司亏损；负数(%)表示`库存消耗 < 订单总量`，可能为客户亏损。"
            hint_text = f"{hint_text_1}\n{explanation_text}"

            # 2. 一次性写入
            ws.append([title_text])
            ws.append([hint_text])
            ws.append(headers)

            # 3. 设置标题行格式
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
            title_cell = ws.cell(row=1, column=1)
            title_cell.alignment = Alignment(horizontal="center")
            title_cell.font = title_font

            # 4. 设置提示行格式
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
            hint_cell = ws.cell(row=2, column=1)
            hint_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            hint_cell.font = hint_font
            # 恢复默认行高（添加配置信息之前的高度）
            ws.row_dimensions[2].height = 30

            # 5. 设置表头行格式
            for cell in ws[3]: # 表头现在是第3行
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            data_start_row = 4 # 数据从第4行开始
            # 写入数据和公式
            for row_idx, device_data in enumerate(summary_data, start=data_start_row):
                row_num = row_idx # 当前行号
                # A列: 序号
                seq_cell = ws.cell(row=row_num, column=1, value=row_idx - data_start_row + 1)
                seq_cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.cell(row=row_num, column=2, value=device_data.get('device_code'))
                ws.cell(row=row_num, column=3, value=device_data.get('customer_name'))
                # D列: 设备桶数 - 从summary_data中直接读取（已从配置文件加载）
                barrel_count = device_data.get('barrel_count', 1)
                ws.cell(row=row_num, column=4, value=barrel_count)
                ws.cell(row=row_num, column=5, value=device_data.get('total_order_volume'))
                # F列: 单桶库存消耗（包含期初-期末+原油剩余量增量，单桶值）
                single_barrel_consumption = device_data.get('total_inventory_consumption_single_barrel', device_data.get('total_inventory_consumption', 0))
                ws.cell(row=row_num, column=6, value=single_barrel_consumption)

                # --- 写入Excel公式 ---
                # G列: 库存消耗总量 = 单桶库存消耗 * 桶数（业务逻辑3：不包含入库订单的oil_val）
                ws.cell(row=row_num, column=7, value=f"=D{row_num}*F{row_num}")
                # H列: 误差值总数 = G - E
                ws.cell(row=row_num, column=8, value=f"=G{row_num}-E{row_num}")
                # I列: 平均每日误差 = H / (查询天数)
                days_in_range = device_data.get('days_in_range', 1)
                avg_error_cell = ws.cell(row=row_num, column=9, value=f"=H{row_num}/{days_in_range}")
                avg_error_cell.number_format = '0.00'
                # J列: 误差百分比 = H / E
                ws.cell(row=row_num, column=10, value=f'=IF(E{row_num}=0, 0, H{row_num}/E{row_num})')
                ws.cell(row=row_num, column=10).number_format = '0.00%' # 设置为百分比格式

                # --- 数据验证和备注生成 ---
                remarks = []
                
                # 检查数据计算问题
                total_order_volume = device_data.get('total_order_volume')
                consumption = device_data.get('total_inventory_consumption_single_barrel', 0)
                
                if total_order_volume is None:
                    remarks.append("⚠️ 订单总量为NULL，可能该设备在指定日期范围内没有订单数据")
                elif total_order_volume == 0:
                    remarks.append("⚠️ 订单总量为0，该设备在指定日期范围内没有消耗订单")
                
                if consumption is None:
                    remarks.append("⚠️ 库存消耗计算为NULL，期初/期末库存数据可能异常")
                elif abs(consumption) > 10000:  # 异常大的值
                    remarks.append(f"⚠️ 库存消耗值异常大({consumption:.2f}L)，请检查期初/期末库存数据")
                
                # --- 处理离线时长和备注 ---
                # from datetime import datetime # 已在文件顶部引入，此处不再重复
                offline_events = device_data.get('offline_events', [])
                total_offline_hours = 0
                # 注意：remarks已在上面初始化用于数据验证，这里继续追加离线事件备注，不再重新定义

                # 将日期字符串转换为datetime对象
                start_date_dt = datetime.strptime(start_date, '%Y-%m-%d')
                end_date_dt = datetime.strptime(end_date + ' 23:59:59', '%Y-%m-%d %H:%M:%S')
                
                # 获取当前实际时间，用于精确计算“离线至今”
                current_real_time = datetime.now()

                for event in offline_events:
                    create_time = event.get('create_time')
                    biz_type = event.get('biz_type')
                    recovery_time = event.get('recovery_time')

                    # 格式化备注文本
                    create_time_str = create_time.strftime('%Y-%m-%d %H:%M')
                    
                    # biz_type = 2 代表已恢复
                    if biz_type == 2 and recovery_time:
                        recovery_time_str = recovery_time.strftime('%Y-%m-%d %H:%M')
                        remarks.append(f"{create_time_str}离线至{recovery_time_str}恢复")
                        # 计算交集时长
                        overlap_start = max(create_time, start_date_dt)
                        overlap_end = min(recovery_time, end_date_dt)
                        if overlap_end > overlap_start:
                            total_offline_hours += (overlap_end - overlap_start).total_seconds() / 3600
                    # biz_type = 1 代表未恢复
                    elif biz_type == 1:
                        # 确定离线事件的有效结束时间，不超过当前实际时间，也不超过查询结束时间
                        effective_end_time_for_event = min(current_real_time, end_date_dt)
                        
                        # 根据有效结束时间生成备注
                        remark_end_time_str = ""
                        if effective_end_time_for_event == current_real_time:
                            remark_end_time_str = f"当前时间({current_real_time.strftime('%Y-%m-%d %H:%M')})"
                        else: # effective_end_time_for_event == end_date_dt
                            remark_end_time_str = f"查询结束({end_date_dt.strftime('%Y-%m-%d %H:%M')})"
                        
                        remarks.append(f"{create_time_str}离线至{remark_end_time_str}")
                        
                        # 计算交集时长
                        overlap_start = max(create_time, start_date_dt)
                        overlap_end = effective_end_time_for_event # 使用有效结束时间进行计算
                        if overlap_end > overlap_start:
                            total_offline_hours += (overlap_end - overlap_start).total_seconds() / 3600


                # K列: 累计离线时长
                ws.cell(row=row_num, column=11, value=round(total_offline_hours, 2))

                # --- 应用样式 ---
                is_high_error = False
                try:
                    # 尝试计算误差百分比的绝对值
                    total_order = float(device_data.get('total_order_volume', 0))
                    total_consumption = float(device_data.get('total_inventory_consumption', 0))
                    if total_order != 0:
                        # 注意：这里用的是单桶消耗，需要乘以桶数才能和Excel公式对应，但我们只做判断，所以暂时忽略桶数影响
                        error_percentage_abs = abs((total_consumption - total_order) / total_order)
                        if error_percentage_abs > 0.05: # 误差超过5%
                            is_high_error = True
                except (ValueError, TypeError):
                    pass # 如果数据转换失败，则不应用高亮
                
                # L列: 备注
                remark_cell = ws.cell(row=row_num, column=12)
                if remarks:
                    remark_cell.value = "\n".join(remarks)
                    remark_cell.font = warning_font
                    remark_cell.alignment = Alignment(vertical='top', wrap_text=True) # 重新启用自动换行

                # 应用行样式
                for col_num in range(1, 13):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    if (row_num - data_start_row) % 2 == 1: # 斑马纹
                        cell.fill = even_row_fill
                
                # 设置数据行固定高度，防止备注过长自动扩展
                ws.row_dimensions[row_num].height = 15 # 设置一个适合单行文本的固定高度

                # 如果误差高，只高亮误差百分比列
                if is_high_error:
                    ws.cell(row=row_num, column=10).fill = warning_fill
                        
            # --- 应用筛选功能 ---
            # 设置筛选范围，从表头行开始，到数据最后一行结束
            ws.auto_filter.ref = f"A3:L{ws.max_row}"

            # --- 创建并设置 "非单桶设备编码" Sheet（保留作为参考） ---
            ws_update = wb.create_sheet("非单桶设备编码")
            
            # 设置标题和提示
            ws_update.append(["非单桶设备编码及桶数参考表"])
            ws_update.merge_cells('A1:C1')
            ws_update['A1'].font = title_font
            ws_update['A1'].alignment = Alignment(horizontal="center")

            ws_update.append(["注意：\n 1.此Sheet仅作为概览，显示报表中已从配置文件自动读取的设备编码和桶数（仅显示桶数>1的设备）。\n 2.若设备数据变动，需同步维护test_data/device_config.csv配置文件，保持一次维护、多次复用的准确性。"])
            ws_update.merge_cells('A2:C2')
            ws_update['A2'].font = hint_font
            ws_update['A2'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws_update.row_dimensions[2].height = 50  # 设置行高以确保文本完整显示，支持换行

            # 设置表头
            header_row = 3
            update_headers = ["设备编码", "桶数", "备注"]
            ws_update.append(update_headers)
            # 设置表头格式
            for col_idx, header in enumerate(update_headers, start=1):
                cell = ws_update.cell(row=header_row, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border

            # 从summary_data中提取设备编码和桶数，只写入桶数>1的设备
            reference_row = header_row + 1
            for device_data in summary_data:
                device_code = device_data.get('device_code')
                barrel_count = device_data.get('barrel_count', 1)
                # 只显示桶数>1的设备
                if device_code and barrel_count > 1:
                    ws_update.cell(row=reference_row, column=1, value=device_code)
                    ws_update.cell(row=reference_row, column=2, value=barrel_count)
                    ws_update.cell(row=reference_row, column=3, value="已配置非默认桶数")
                    reference_row += 1

            # 增加列宽度，确保提示文本能够完整显示
            ws_update.column_dimensions['A'].width = 50  # 增加A列宽度以显示完整提示文本
            ws_update.column_dimensions['B'].width = 12
            ws_update.column_dimensions['C'].width = 35

            # 调整列宽
            column_widths = {'A': 8, 'B': 20, 'C': 25, 'D': 12, 'E': 15, 'F': 20, 'G': 18, 'H': 15, 'I': 18, 'J': 15, 'K': 20, 'L': 30}
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            # 保存文件
            wb.save(output_file_path)
            print(f"误差汇总报表已生成并保存: {output_file_path}")
            return True

        except Exception as e:
            print(f"生成误差汇总报表时发生错误: {e}")
            import traceback
            traceback.print_exc()
            return False
        finally:
            if 'wb' in locals() and wb:
                wb.close()
