"""
报表服务模块
负责处理所有与报表生成相关的业务逻辑
"""
import tempfile
from pathlib import Path
from typing import List, Dict, Any, Tuple
from datetime import datetime, date, timedelta
import zipfile
import os
from collections import defaultdict
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.repositories.device_repository import DeviceRepository
from src.logger import logger
from src.utils.date_utils import parse_date, validate_error_summary_date_span


class ReportService:
    """
    封装报表生成的核心业务逻辑
    """

    def __init__(self, device_repo: DeviceRepository):
        self.device_repo = device_repo
        self._load_barrel_overrides()

    def _load_barrel_overrides(self):
        """使用pandas加载人工校准的桶数配置文件，更健壮"""
        self.barrel_overrides = {}
        override_file = Path(__file__).parent.parent.parent / "config" / "barrel_count_override.csv"
        
        logger.info(f"正在尝试加载桶数校准文件: {override_file.resolve()}")
        
        if not override_file.exists():
            logger.warning(f"桶数校准文件未找到。")
            return

        try:
            df = pd.read_csv(override_file, comment='#')
            df.columns = [col.strip() for col in df.columns]
            df = df.dropna(subset=['device_code', 'barrel_count'])
            df['device_code'] = df['device_code'].astype(str).str.strip()
            df = df[pd.to_numeric(df['barrel_count'], errors='coerce').notna()]
            df['barrel_count'] = df['barrel_count'].astype(int)
            self.barrel_overrides = pd.Series(df.barrel_count.values, index=df.device_code).to_dict()
            logger.info(f"成功加载 {len(self.barrel_overrides)} 条桶数校准记录。")
        except Exception as e:
            logger.error(f"使用pandas读取桶数校准文件时出错: {e}", exc_info=True)  # 修复缩进

    # --- Main Dispatch Method ---
    async def generate_report(
        self,
        report_type: str,
        devices: List[str], # 接收设备编码列表
        start_date: str,
        end_date: str,
    ) -> Tuple[Path, List[str]]:
        logger.info(f"收到报表生成请求: 类型='{report_type}', 设备数={len(devices)}")

        if report_type == "daily_consumption":
            return await self._generate_daily_consumption_report(
                devices, start_date, end_date
            )
        elif report_type == "error_summary":
            return await self._generate_error_summary_report(
                devices, start_date, end_date
            )
        elif report_type == "inventory":
            return await self._generate_inventory_report(
                devices, start_date, end_date
            )
        elif report_type == "statement":
            return await self._generate_customer_statement_report(
                devices, start_date, end_date
            )
        elif report_type == "refueling":
            return await self._generate_refueling_details_report(
                devices, start_date, end_date
            )
        else:
            raise NotImplementedError(f"报表类型 '{report_type}' 的生成逻辑尚未实现。")

    # --- Specific Report Generation Logic ---

    async def _generate_daily_consumption_report(
        self,
        device_codes: List[str],
        start_date_str: str,
        end_date_str: str,
    ) -> Tuple[Path, List[str]]:
        device_codes_tuple = tuple(sorted(device_codes))
        raw_report_data = await self.device_repo.get_daily_consumption_raw_data(
            device_codes_tuple, start_date_str, end_date_str
        )
        if not raw_report_data:
            raise ValueError("在指定日期范围内所有选定设备均未找到任何消耗数据。")

        data_by_device = defaultdict(list)
        for row in raw_report_data:
            data_by_device[row['device_code']].append(row)

        report_files = []
        warnings = []
        for device_code in device_codes:
            device_data = data_by_device.get(device_code)
            if not device_data:
                warnings.append(f"设备 {device_code} 没有可用于计算的数据，已跳过。")
                continue
            try:
                barrel_count = self.barrel_overrides.get(device_code, 1)
                logger.info(f"设备 {device_code} 使用桶数: {barrel_count}")
                final_report_data = self._calculate_final_data(device_data, barrel_count, "daily")
                
                if not final_report_data:
                    warnings.append(f"设备 {device_code} 计算后数据为空，已跳过。")
                    continue

                customer_name = final_report_data[0].get('customer_name', '未知客户')
                oil_name = final_report_data[0].get('oil_name', '未知油品')
                
                report_path = self._generate_daily_detail_excel(
                    final_report_data, device_code, customer_name, oil_name, start_date_str, end_date_str, barrel_count
                )
                if report_path:
                    report_files.append(report_path)
            except Exception as e:
                warnings.append(f"为设备 {device_code} 生成Excel文件失败: {e}")
                logger.error(f"为设备 {device_code} 生成Excel文件失败: {e}", exc_info=True)
        
        if not report_files:
            raise ValueError("所有选定设备均未能成功生成报表。")

        if len(report_files) == 1:
            return report_files[0], warnings

        zip_path = Path(tempfile.gettempdir()) / f"ZR_Daily_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file_path in report_files:
                if file_path:
                    zipf.write(file_path, os.path.basename(file_path))
                    os.remove(file_path)
        return zip_path, warnings



    def _calculate_final_data(self, device_raw_data: List[Dict[str, Any]], barrel_count: int, period: str) -> List[Dict[str, Any]]:
        order_volume_key = 'daily_order_volume' if period == 'daily' else 'monthly_order_volume'
        inventory_key = 'end_of_day_inventory' if period == 'daily' else 'end_of_month_inventory'
        prev_inventory_key = 'prev_day_inventory' if period == 'daily' else 'prev_month_inventory'
        refill_key = 'daily_refill' if period == 'daily' else 'monthly_refill'
        final_data = []
        for row in device_raw_data:
            prev_inventory = row.get(prev_inventory_key) or 0
            end_inventory = row.get(inventory_key) or 0
            refill = row.get(refill_key) or 0
            inventory_consumption = max(0, (prev_inventory - end_inventory) + refill)
            real_consumption = inventory_consumption * barrel_count
            order_volume = row.get(order_volume_key) or 0
            error = real_consumption - order_volume
            day_data = row.copy()
            day_data['real_consumption'] = real_consumption
            day_data['error'] = error
            final_data.append(day_data)
        return final_data

    def _generate_daily_detail_excel(self, report_data: List[Dict[str, Any]], device_code: str, customer_name: str, oil_name: str, start_date_str: str, end_date_str: str, barrel_count: int) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.title = "消耗误差分析"
        oil_name_str = f" {oil_name} " if oil_name and pd.notna(oil_name) else " "
        title = f"{device_code}{oil_name_str}每日消耗误差分析({start_date_str} - {end_date_str})"
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", wrap_text=True)
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        headers = ["日期", "原油剩余量(L)", "订单累积总量(L)", "库存消耗总量(L)", "中润亏损(L)", "客户亏损(L)"]
        ws.append(headers)
        for row in report_data:
            shortage_error = max(0, row['error'])
            excess_error = max(0, -row['error'])
            ws.append([row['report_date'].isoformat(), row['end_of_day_inventory'], row['daily_order_volume'], row['real_consumption'], shortage_error, excess_error])
        column_widths = {'A': 12, 'B': 15, 'C': 18, 'D': 18, 'E': 15, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        chart = LineChart()
        chart.title = "每日消耗误差分析"
        chart.style = 13
        chart.y_axis.title = "值 (L)"
        chart.y_axis.majorGridlines = ChartLines(spPr=GraphicalProperties(noFill=True))
        chart.x_axis.title = "日期"
        chart.x_axis.number_format = 'yyyy-mm-dd'
        chart.x_axis.tickLblSkip = 3
        chart.x_axis.tickLblPos = "low"
        chart.x_axis.textRotation = 0
        chart.width = 30
        chart.height = 15
        dates = Reference(ws, min_col=1, min_row=3, max_row=len(report_data) + 2)
        data = Reference(ws, min_col=2, min_row=2, max_col=4, max_row=len(report_data) + 2)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(dates)
        s1 = chart.series[0]
        s1.graphicalProperties.line.solidFill = "0000FF"
        s1.marker = Marker(symbol="circle", size=8)
        s2 = chart.series[1]
        s2.graphicalProperties.line.solidFill = "00FF00"
        s2.marker = Marker(symbol="circle", size=8)
        s3 = chart.series[2]
        s3.graphicalProperties.line.solidFill = "800080"
        s3.marker = Marker(symbol="circle", size=8)
        ws.add_chart(chart, "H5")
        annotation_row = 34
        ws.cell(row=annotation_row, column=8).value = "计算规则说明："
        ws.cell(row=annotation_row, column=8).font = Font(bold=True)
        consumption_formula = f"库存消耗总量(L) = (前日库存 - 当日库存 + 当日加油量) * {barrel_count} (桶数)"
        ws.cell(row=annotation_row + 1, column=8).value = consumption_formula
        ws.merge_cells(start_row=annotation_row + 1, start_column=8, end_row=annotation_row + 1, end_column=15)
        ws.cell(row=annotation_row + 2, column=8).value = "中润亏损(L) = MAX(0, 库存消耗总量 - 订单累积总量)"
        ws.merge_cells(start_row=annotation_row + 2, start_column=8, end_row=annotation_row + 2, end_column=15)
        ws.cell(row=annotation_row + 3, column=8).value = "客户亏损(L) = MAX(0, 订单累积总量 - 库存消耗总量)"
        ws.merge_cells(start_row=annotation_row + 3, start_column=8, end_row=annotation_row + 3, end_column=15)
        final_filename = f"{customer_name}_{device_code}_{start_date_str}_to_{end_date_str}_每日消耗误差报表.xlsx"
        output_path = Path(tempfile.gettempdir()) / final_filename
        wb.save(output_path)
        logger.info(f"Excel报表已生成: {output_path}")
        return output_path



    async def _generate_inventory_report(self, device_codes: List[str], start_date_str: str, end_date_str: str) -> Tuple[Path, List[str]]:
        """
        生成库存报表 - 基于development-copy分支的业务逻辑
        使用严格的数据验证和包含图表的Excel生成方式
        """
        logger.info(f"开始生成库存报表: 设备={device_codes}, 日期范围={start_date_str} 至 {end_date_str}")
        
        # 验证库存报表的日期范围不超过1个月
        date_row = {"start_date": start_date_str, "end_date": end_date_str}
        if not self._validate_inventory_date_span(date_row):
            raise ValueError("库存报表的日期范围不能超过1个月")
        
        device_codes_tuple = tuple(sorted(device_codes))
        raw_report_data = await self.device_repo.get_daily_consumption_raw_data(
            device_codes_tuple, start_date_str, end_date_str
        )
        
        if not raw_report_data:
            raise ValueError("在指定日期范围内所有选定设备均未找到任何库存数据。")

        # 按设备分组数据
        data_by_device = defaultdict(list)
        for row in raw_report_data:
            data_by_device[row['device_code']].append(row)

        report_files = []
        warnings = []
        
        for device_code in device_codes:
            device_data = data_by_device.get(device_code)
            if not device_data:
                warnings.append(f"设备 {device_code} 没有库存数据，已跳过。")
                continue
                
            try:
                # 基于development-copy分支逻辑：使用严格的数据验证
                cleaned_data = []
                invalid_records = []
                
                for i, row in enumerate(device_data):
                    try:
                        # 检查必需字段是否存在
                        if 'report_date' not in row or 'end_of_day_inventory' not in row:
                            logger.warning(f"设备 {device_code} 第{i+1}条数据缺少必需字段，已跳过")
                            continue
                            
                        # 验证库存值 - 使用development-copy分支的严格验证
                        inventory_value = self._validate_inventory_value_strict(row.get('end_of_day_inventory'))
                        
                        # 验证日期格式
                        report_date = self._validate_date_format(row.get('report_date'))
                        
                        # 检查库存值是否超过1000L（与原分支一致）
                        if inventory_value > 1000:
                            logger.info(f"提示：设备 {device_code} 日期 {report_date} 的库存值 {inventory_value} 超过1000L")
                        
                        # 创建包含所有必要字段的数据字典 - 适配daily_consumption_raw_query返回的字段
                        data_record = {
                            'create_time': report_date,
                            'report_date': report_date,
                            'end_of_day_inventory': inventory_value,
                            'oil_name': row.get('oil_name'),
                            'customer_name': row.get('customer_name'),
                            'order_number': '',  # 每日消耗数据没有订单号
                            'fill_time': '',      # 每日消耗数据没有加注时间
                            'oil_type_id': '',   # 每日消耗数据没有油品序号
                            'water_oil_ratio': 0, # 每日消耗数据没有水油比
                            'fill_value': 0,     # 每日消耗数据没有加注值
                            'crude_oil_remaining': inventory_value,  # 使用库存值作为原油剩余量
                            'crude_oil_ratio': 0  # 每日消耗数据没有原油剩余比例
                        }
                        cleaned_data.append(data_record)
                        
                    except ValueError as e:
                        invalid_records.append((row.get('report_date'), row.get('end_of_day_inventory'), str(e)))
                        logger.warning(f"设备 {device_code} 第{i+1}条数据验证失败: {e}")
                        continue
                    except Exception as e:
                        logger.warning(f"设备 {device_code} 第{i+1}条数据验证失败: {e}")
                        continue
                
                # 记录无效数据汇总
                if invalid_records:
                    logger.info(f"设备 {device_code} 无效数据汇总：")
                    for date, value, reason in invalid_records:
                        logger.info(f"- {date}: {value} ({reason})")
                
                # 如果没有有效数据，使用默认数据点
                if not cleaned_data:
                    logger.warning(f"设备 {device_code} 没有有效的库存数据，将生成默认数据图表")
                    try:
                        start_date = self._validate_date_format(start_date_str)
                        end_date = self._validate_date_format(end_date_str)
                        
                        # 从device_data中获取客户名称和油品名称，确保有正确的默认值
                        # 由于device_data可能包含多条记录，我们取第一条有效记录
                        oil_name_default = '未知油品'
                        customer_name_default = '未知客户'
                        
                        if device_data:
                            # 遍历device_data寻找包含客户名称和油品名称的记录
                            for record in device_data:
                                if record.get('customer_name') and record.get('oil_name'):
                                    oil_name_default = record.get('oil_name', '未知油品')
                                    customer_name_default = record.get('customer_name', '未知客户')
                                    break
                            # 如果遍历后仍然没有找到，使用第一条记录的get方法带默认值
                            if oil_name_default == '未知油品' and customer_name_default == '未知客户':
                                oil_name_default = device_data[0].get('oil_name', '未知油品')
                                customer_name_default = device_data[0].get('customer_name', '未知客户')
                        
                        # 创建包含所有必要字段的默认数据字典
                        default_data1 = {
                            'create_time': start_date,
                            'report_date': start_date,
                            'end_of_day_inventory': 0,
                            'oil_name': oil_name_default,
                            'customer_name': customer_name_default,
                            'order_number': '',
                            'fill_time': '',
                            'oil_type_id': '',
                            'water_oil_ratio': 0,
                            'fill_value': 0,
                            'crude_oil_remaining': 0,
                            'crude_oil_ratio': 0
                        }
                        default_data2 = {
                            'create_time': end_date,
                            'report_date': end_date,
                            'end_of_day_inventory': 0,
                            'oil_name': oil_name_default,
                            'customer_name': customer_name_default,
                            'order_number': '',
                            'fill_time': '',
                            'oil_type_id': '',
                            'water_oil_ratio': 0,
                            'fill_value': 0,
                            'crude_oil_remaining': 0,
                            'crude_oil_ratio': 0
                        }
                        cleaned_data = [default_data1, default_data2]
                    except Exception:
                        warnings.append(f"设备 {device_code} 所有数据验证失败，已跳过。")
                        continue
                
                # 获取油品名称和客户名称（确保有默认值）- 从cleaned_data中获取
                oil_name = cleaned_data[0].get('oil_name', '未知油品') if cleaned_data else '未知油品'
                customer_name = cleaned_data[0].get('customer_name', '未知客户') if cleaned_data else '未知客户'
                
                # 确保customer_name不为None，避免文件名中出现"None"
                if customer_name is None:
                    customer_name = '未知客户'
                if oil_name is None:
                    oil_name = '未知油品'
                
                # 生成包含图表的Excel报表
                # 将字符串日期转换为datetime对象
                start_date_dt = self._validate_date_format(start_date_str)
                end_date_dt = self._validate_date_format(end_date_str)
                report_path = self._generate_inventory_excel_with_chart(
                    cleaned_data, device_code, oil_name, customer_name, start_date_dt, end_date_dt
                )
                if report_path:
                    report_files.append(report_path)
                    
            except Exception as e:
                warnings.append(f"为设备 {device_code} 生成库存报表失败: {e}")
                logger.error(f"为设备 {device_code} 生成库存报表失败: {e}", exc_info=True)
        
        if not report_files:
            raise ValueError("所有选定设备均未能成功生成库存报表。")

        if len(report_files) == 1:
            return report_files[0], warnings

        # 多设备时打包为zip
        zip_path = Path(tempfile.gettempdir()) / f"ZR_Inventory_Reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for file_path in report_files:
                if file_path:
                    zipf.write(file_path, os.path.basename(file_path))
                    os.remove(file_path)
        return zip_path, warnings

    def _validate_inventory_value(self, value) -> float:
        """
        验证库存值 - 基于重构前的简单验证逻辑
        """
        if value is None:
            return 0.0
            
        try:
            # 转换为浮点数
            float_value = float(value)
            
            # 检查是否为负数
            if float_value < 0:
                logger.warning(f"库存值为负数: {float_value}，已设置为0")
                return 0.0
                
            return float_value
            
        except (ValueError, TypeError):
            logger.warning(f"库存值格式无效: {value}，已设置为0")
            return 0.0

    def _validate_inventory_value_strict(self, value) -> float:
        """
        验证库存值 - 基于development-copy分支的严格验证逻辑
        - 允许大于100的值
        - 不允许小于0的值
        - 不允许非数字值
        """
        try:
            float_value = float(value)
            if float_value < 0:
                raise ValueError("库存值不能为负数")
            return float_value
        except (ValueError, TypeError):
            raise ValueError(f"无效的库存值: {value}")

    def _validate_date_format(self, date_value):
        """
        验证日期格式 - 支持多种日期格式
        基于原分支实现，支持datetime.date、datetime.datetime和字符串类型
        """
        if date_value is None:
            raise ValueError("日期值为空")
            
        try:
            # 如果是datetime.date对象，直接返回（原分支的处理方式）
            if isinstance(date_value, date):
                return date_value
            # 如果是datetime.datetime对象，提取日期部分
            elif isinstance(date_value, datetime):
                return date_value.date()
            # 如果是字符串，尝试解析
            elif isinstance(date_value, str):
                return parse_date(date_value)
            else:
                raise ValueError(f"不支持的日期格式: {type(date_value)}")
                
        except Exception as e:
            raise ValueError(f"日期格式验证失败: {date_value}, 错误: {e}")

    def _validate_inventory_date_span(self, date_row: Dict[str, str]) -> bool:
        """
        验证库存报表的日期范围不超过1个月
        
        Args:
            date_row (dict): 包含'start_date'和'end_date'字段的字典
            
        Returns:
            bool: 验证通过返回True，否则返回False
        """
        try:
            start_date = self._validate_date_format(date_row["start_date"])
            end_date = self._validate_date_format(date_row["end_date"])
        except ValueError as e:
            logger.error(f"库存报表日期格式错误: {date_row}。{e}")
            return False

        # 验证开始日期不能晚于结束日期
        if start_date > end_date:
            logger.error(f"库存报表日期逻辑错误: 开始日期 {date_row['start_date']} 不能晚于结束日期 {date_row['end_date']}")
            return False

        # 计算总天数
        total_days = (end_date - start_date).days
        
        # 计算年份差和月份差
        year_diff = end_date.year - start_date.year
        month_diff = end_date.month - start_date.month
        
        # 总月份差
        total_month_diff = year_diff * 12 + month_diff
        
        # 检查是否在1个月范围内（范围 <= 1个月）
        if total_month_diff > 1 or (total_month_diff == 1 and total_days > 31):  # 允许最多1个月，但总天数不超过31天
            logger.error(f"库存报表日期跨度错误: 从 {date_row['start_date']} 到 {date_row['end_date']} 的日期范围超过1个月")
            return False
        
        return True

    def _generate_inventory_excel(self, report_data: List[Dict[str, Any]], device_code: str, start_date_str: str, end_date_str: str) -> Path:
        """
        生成库存Excel报表 - 基于重构前的简单格式
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "库存报表"
        
        # 设置标题
        title = f"设备 {device_code} 库存报表({start_date_str} - {end_date_str})"
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        # 表头
        headers = ["日期", "当日库存(L)", "客户名称", "油品名称"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # 填充数据
        for row in report_data:
            ws.append([
                row['report_date'].strftime('%Y-%m-%d'),
                row['end_of_day_inventory'],
                row.get('customer_name', '未知客户'),
                row.get('oil_name', '未知油品')
            ])
        
        # 设置列宽
        column_widths = {'A': 12, 'B': 15, 'C': 20, 'D': 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 生成文件名
        customer_name = report_data[0].get('customer_name', '未知客户') if report_data else '未知客户'
        final_filename = f"{customer_name}_{device_code}_{start_date_str}_to_{end_date_str}_库存报表.xlsx"
        output_path = Path(tempfile.gettempdir()) / final_filename
        
        wb.save(output_path)
        logger.info(f"库存Excel报表已生成: {output_path}")
        return output_path

    async def _generate_error_summary_report(self, device_codes: List[str], start_date_str: str, end_date_str: str) -> Tuple[Path, List[str]]:
        # 验证日期范围不超过2个月
        date_row = {"start_date": start_date_str, "end_date": end_date_str}
        if not validate_error_summary_date_span(date_row):
            raise ValueError("误差汇总报表的日期范围不能超过2个月")
        
        device_codes_tuple = tuple(sorted(device_codes))
        raw_report_data = await self.device_repo.get_error_summary_raw_data(
            device_codes_tuple, start_date_str, end_date_str
        )
        if not raw_report_data:
            raise ValueError("在指定日期范围内所有选定设备均未找到任何误差汇总数据。")

        # 查询离线事件数据
        offline_events = await self._get_offline_events(device_codes_tuple, start_date_str, end_date_str)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "误差汇总报表"
        
        # 设置标题
        title = f"安卓设备消耗误差汇总报表({start_date_str} - {end_date_str})"
        ws.append([title])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=1).font = Font(size=16, bold=True, name="微软雅黑")
        
        # 添加计算公式说明在主标题下第二行
        ws.append(["计算公式说明：单桶库存消耗 = 期末库存 - 期初库存 + 加油量，库存消耗总量 = 设备桶数 × 单桶库存消耗，误差值总数 = 库存消耗总量 - 订单总量，误差百分比 = 误差值总数 / 订单总量"])  
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
        ws.cell(row=2, column=1).font = Font(size=10, name="微软雅黑")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")
        
        # 调整表头，将误差类型列移到误差百分比列前面，删除是否有离线事件列
        headers = ["序号", "设备编码", "客户名称", "设备桶数", "订单总量(L)", "单桶库存消耗(L)", 
                  "库存消耗总量(L)", "误差值总数(L)", "误差类型", "误差百分比(%)", "离线时长(小时)", "备注"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[3]:  # 表头行现在是第3行
            cell.font = Font(bold=True, size=12, name="微软雅黑")
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 准备数据
        processed_data = []
        for row in raw_report_data:
            # 修正计算真实消耗量：期初库存 - 期末库存 + 加油量
            end_inventory = row.get('end_of_day_inventory') or 0
            prev_inventory = row.get('prev_day_inventory') or 0
            daily_refill = row.get('daily_refill') or 0
            real_consumption = prev_inventory - end_inventory + daily_refill  # 修正这里！
            
            # 添加计算后的字段
            processed_row = row.copy()
            processed_row['real_consumption'] = real_consumption
            processed_data.append(processed_row)
        
        # 按设备编码和日期分组
        grouped_data = defaultdict(list)
        for row in processed_data:
            key = (row['device_code'], row.get('customer_name', '未知客户'), row.get('oil_name', '未知油品'))
            grouped_data[key].append(row)
        
        # 填充数据行
        row_index = 4  # 数据行从第4行开始
        index = 1
        
        for (device_code, customer_name, oil_name), rows in grouped_data.items():
            # 计算汇总数据
            total_order_volume = sum(r.get('daily_order_volume', 0) for r in rows)
            
            # 计算单桶库存消耗（使用最后一条记录的真实消耗量）并确保≥0
            single_barrel_consumption = max(0, rows[-1].get('real_consumption', 0))
            
            # 获取设备桶数（从校准文件中获取，默认为1）
            barrel_count = self.barrel_overrides.get(device_code, 1)
            
            # 计算库存消耗总量 = 设备桶数 × 单桶库存消耗
            total_inventory_consumption = barrel_count * single_barrel_consumption
            
            # 计算误差值总数 = 库存消耗总量 - 订单总量
            total_error = total_inventory_consumption - total_order_volume
            
            # 计算误差百分比
            error_percentage = (total_error / total_order_volume * 100) if total_order_volume > 0 else 0
            
            # 获取该设备的离线事件
            device_offline_events = offline_events.get(device_code, [])
            
            # 计算离线时长和生成备注
            total_offline_hours = 0
            offline_remarks = []
            
            for event in device_offline_events:
                # 计算离线时长
                start_time = event['create_time']
                recovery_time = event.get('recovery_time')
                
                # 如果恢复时间为None，表示设备仍在离线状态，使用当前时间计算
                if recovery_time is None:
                    end_time = datetime.now()
                    end_str = f"至今({end_time.strftime('%Y-%m-%d %H:%M')})"
                else:
                    end_time = recovery_time
                    end_str = end_time.strftime("%Y-%m-%d %H:%M")
                
                duration_hours = (end_time - start_time).total_seconds() / 3600
                total_offline_hours += duration_hours
                
                # 格式化事件时间并添加到备注
                start_str = start_time.strftime("%Y-%m-%d %H:%M")
                offline_remarks.append(f"{start_str} - {end_str}")
            
            # 生成完整备注，使用换行符分隔离线事件
            remarks = []
            if offline_remarks:
                remarks.append("离线事件：\n" + "\n".join(offline_remarks))
            
            if abs(error_percentage) > 10:
                remarks.append("误差超过10%，请重点关注")
            elif abs(error_percentage) > 5:
                remarks.append("误差超过5%，建议核查")
            
            remark_text = "\n".join(remarks) if remarks else "正常"
            
            # 确定误差类型
            if total_error > 0:
                error_type = "客户亏损"
            elif total_error < 0:
                error_type = "中润亏损"
            else:
                error_type = "无误差"
            
            # 填充数据行，调整列顺序
            ws.cell(row=row_index, column=1).value = index  # 序号
            ws.cell(row=row_index, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_index, column=1).font = Font(size=11, name="微软雅黑")
            
            ws.cell(row=row_index, column=2).value = device_code  # 设备编码
            ws.cell(row=row_index, column=2).font = Font(size=11, name="微软雅黑")
            
            ws.cell(row=row_index, column=3).value = customer_name  # 客户名称
            ws.cell(row=row_index, column=3).font = Font(size=11, name="微软雅黑")
            
            ws.cell(row=row_index, column=4).value = barrel_count  # 设备桶数
            ws.cell(row=row_index, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=row_index, column=4).font = Font(size=11, name="微软雅黑")
            
            ws.cell(row=row_index, column=5).value = total_order_volume  # 订单总量
            ws.cell(row=row_index, column=5).alignment = Alignment(horizontal="right")
            ws.cell(row=row_index, column=5).font = Font(size=11, name="微软雅黑")
            
            ws.cell(row=row_index, column=6).value = single_barrel_consumption  # 单桶库存消耗
            ws.cell(row=row_index, column=6).alignment = Alignment(horizontal="right")
            ws.cell(row=row_index, column=6).font = Font(size=11, name="微软雅黑")
            
            # 库存消耗总量使用公式
            ws.cell(row=row_index, column=7).value = f"=D{row_index}*F{row_index}"  # 库存消耗总量
            ws.cell(row=row_index, column=7).alignment = Alignment(horizontal="right")
            ws.cell(row=row_index, column=7).font = Font(size=11, name="微软雅黑")
            
            # 误差值总数使用公式
            ws.cell(row=row_index, column=8).value = f"=G{row_index}-E{row_index}"  # 误差值总数
            ws.cell(row=row_index, column=8).alignment = Alignment(horizontal="right")
            ws.cell(row=row_index, column=8).font = Font(size=11, name="微软雅黑")
            
            # 误差类型列移到误差百分比前面
            ws.cell(row=row_index, column=9).value = error_type  # 误差类型
            ws.cell(row=row_index, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=row_index, column=9).font = Font(size=11, name="微软雅黑")
            
            # 误差百分比使用公式
            ws.cell(row=row_index, column=10).value = f"=H{row_index}/E{row_index}"  # 误差百分比
            ws.cell(row=row_index, column=10).number_format = '0.00%'
            ws.cell(row=row_index, column=10).alignment = Alignment(horizontal="right")
            ws.cell(row=row_index, column=10).font = Font(size=11, name="微软雅黑")
            
            # 仅误差百分比列标记背景色，不再整行标记
            if abs(error_percentage) > 5:
                ws.cell(row=row_index, column=10).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            ws.cell(row=row_index, column=11).value = round(total_offline_hours, 2)  # 离线时长
            ws.cell(row=row_index, column=11).alignment = Alignment(horizontal="center")
            ws.cell(row=row_index, column=11).font = Font(size=11, name="微软雅黑")
            
            ws.cell(row=row_index, column=12).value = remark_text  # 备注
            ws.cell(row=row_index, column=12).alignment = Alignment(wrap_text=True, vertical="top")  # 自动换行，顶部对齐
            ws.cell(row=row_index, column=12).font = Font(size=10, name="微软雅黑")
            
            # 设置固定行高，避免备注内容影响整体布局
            ws.row_dimensions[row_index].height = 30
            
            # 斑马纹样式（跳过已标记背景色的单元格）
            if row_index % 2 == 0:
                for col in range(1, 13):
                    if col != 10:  # 跳过误差百分比列
                        cell = ws.cell(row=row_index, column=col)
                        if not cell.fill.start_color.rgb or cell.fill.start_color.rgb == "00000000":
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            row_index += 1
            index += 1
        
        # 设置列宽
        column_widths = [6, 15, 20, 10, 12, 15, 15, 12, 12, 12, 12, 30]
        for i, width in enumerate(column_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = width
        
        # 不再添加统计信息行
        
        # 不再添加单独的计算公式说明区域
        
        final_filename = f"安卓设备消耗误差汇总_{start_date_str}_to_{end_date_str}.xlsx"
        output_path = Path(tempfile.gettempdir()) / final_filename
        wb.save(output_path)
        logger.info(f"误差汇总报表已生成: {output_path}")
        return output_path, []

    async def _get_offline_events(self, device_codes, start_date, end_date):
        """
        获取指定设备和日期范围内的离线事件
        """
        # 移除不存在的settings导入，直接使用SQL查询
        from sqlalchemy import text
        from sqlalchemy.ext.asyncio import AsyncSession
        from src.database import get_db_session
        import json
        from pathlib import Path
        
        offline_events = defaultdict(list)
        
        try:
            # 直接读取查询配置文件
            config_path = Path(__file__).parent.parent.parent / "config" / "query_config.json"
            with open(config_path, 'r', encoding='utf-8') as f:
                sql_templates = json.load(f)
            
            # 获取离线事件查询模板
            sql_templates_config = sql_templates.get("sql_templates", {})
            query_template = sql_templates_config.get("error_summary_offline_query", "")
            
            if not query_template:
                logger.warning("未找到离线事件查询模板")
                return offline_events
            
            async with get_db_session() as session:
                params = {
                    "start_date_param_full": f"{start_date} 00:00:00",
                    "end_date_param_full": f"{end_date} 23:59:59",
                    "device_codes": tuple(device_codes)
                }
                
                statement = text(query_template)
                result = await session.execute(statement, params)
                # 正确转换SQLAlchemy Row对象为字典
                events = [dict(zip(result.keys(), row)) for row in result.fetchall()]
                
                # 按设备编码分组（SQL查询已通过IN条件筛选，直接分组即可）
                for event in events:
                    device_code = event['device_code']
                    offline_events[device_code].append(event)
                
            logger.info(f"获取到 {sum(len(events) for events in offline_events.values())} 条离线事件记录")
        except Exception as e:
            logger.error(f"获取离线事件时出错: {e}", exc_info=True)
        
        return output_path

    async def _generate_customer_statement_report(
        self,
        device_codes: List[str],
        start_date_str: str,
        end_date_str: str,
    ) -> Tuple[Path, List[str]]:
        """
        生成客户对账单 - 基于模板格式生成，包含3个sheet
        
        业务逻辑：
        - 不需要期初库存、期末库存、加油量等
        - 只需要：序号、油品名称/型号、设备编码、本月总升数计量（L）、备注
        - 本月总升数计量 = 订单量（daily_order_volume）的汇总
        
        Args:
            device_codes: 设备编码列表
            start_date_str: 开始日期
            end_date_str: 结束日期
            
        Returns:
            (生成的Excel文件路径, 警告信息列表)
        """
        logger.info(f"开始生成客户对账单: 设备={device_codes}, 日期范围={start_date_str} 至 {end_date_str}")
        
        warnings = []
        
        try:
            # 1. 验证日期范围（不超过1个月）
            from src.utils.date_utils import validate_date_span
            date_row = {"start_date": start_date_str, "end_date": end_date_str}
            if not validate_date_span(date_row):
                raise ValueError("客户对账单的日期范围不能超过1个月")
            
            # 2. 获取设备数据
            device_codes_tuple = tuple(sorted(device_codes))
            raw_report_data = await self.device_repo.get_daily_consumption_raw_data(
                device_codes_tuple, start_date_str, end_date_str
            )
            
            if not raw_report_data:
                raise ValueError("在指定日期范围内所有选定设备均未找到任何数据。")
            
            # 3. 按设备和油品分组，汇总订单量
            # 数据结构: {(device_code, oil_name): {'customer_name': ..., 'total_order_volume': ..., 'device_name': ...}}
            device_oil_data = defaultdict(lambda: {
                'customer_name': '',
                'device_name': '',
                'oil_name': '',
                'total_order_volume': 0.0
            })
            
            for row in raw_report_data:
                device_code = row.get('device_code', '')
                oil_name = row.get('oil_name', '未知油品')
                key = (device_code, oil_name)
                
                if not device_oil_data[key]['customer_name']:
                    device_oil_data[key]['customer_name'] = row.get('customer_name', '未知客户')
                if not device_oil_data[key]['device_name']:
                    device_oil_data[key]['device_name'] = row.get('device_name', '未知设备')
                if not device_oil_data[key]['oil_name']:
                    device_oil_data[key]['oil_name'] = oil_name
                
                # 汇总订单量
                order_volume = row.get('daily_order_volume', 0) or 0
                device_oil_data[key]['total_order_volume'] += float(order_volume)
            
            # 4. 按客户分组设备
            customer_groups = defaultdict(list)
            for (device_code, oil_name), data in device_oil_data.items():
                customer_name = data['customer_name']
                customer_groups[customer_name].append({
                    'device_code': device_code,
                    'device_name': data['device_name'],
                    'oil_name': data['oil_name'],
                    'total_order_volume': data['total_order_volume']
                })
            
            if not customer_groups:
                raise ValueError("所有选定设备均未能成功处理数据。")
            
            # 5. 生成客户对账单（使用模板）
            output_path = self._generate_customer_statement_from_template(
                customer_groups, start_date_str, end_date_str
            )
            
            logger.info(f"客户对账单已生成: {output_path}")
            
            return output_path, warnings
            
        except Exception as e:
            error_msg = f"生成客户对账单失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ValueError(error_msg)
    
    def _generate_customer_statement_from_template(
        self,
        customer_groups: Dict[str, List[Dict[str, Any]]],
        start_date_str: str,
        end_date_str: str
    ) -> Path:
        """
        使用模板文件生成客户对账单
        
        模板包含3个sheet：
        1. 中润对账单 - 主对账单，包含序号、油品名称/型号、设备编码、本月总升数计量（L）、备注
        2. 每日用量明细 - 每日用量明细
        3. 每月用量对比 - 月度对比数据
        
        Args:
            customer_groups: 按客户分组的设备数据
            start_date_str: 开始日期
            end_date_str: 结束日期
            
        Returns:
            生成的Excel文件路径
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment
        from datetime import datetime
        import tempfile
        from pathlib import Path
        
        # 加载模板文件
        template_path = Path(__file__).parent.parent.parent / "template" / "statement_template.xlsx"
        if not template_path.exists():
            raise FileNotFoundError(f"客户对账单模板文件未找到: {template_path}")
        
        wb = load_workbook(template_path)
        
        # 为每个客户生成对账单
        for customer_name, devices in customer_groups.items():
            # 更新"中润对账单"sheet
            if "中润对账单" in wb.sheetnames:
                ws_main = wb["中润对账单"]
                self._update_main_statement_sheet(ws_main, customer_name, devices, start_date_str, end_date_str)
            
            # 更新"每日用量明细"sheet
            if "每日用量明细" in wb.sheetnames:
                ws_daily = wb["每日用量明细"]
                self._update_daily_detail_sheet(ws_daily, devices, start_date_str, end_date_str)
            
            # 更新"每月用量对比"sheet
            if "每月用量对比" in wb.sheetnames:
                ws_monthly = wb["每月用量对比"]
                self._update_monthly_comparison_sheet(ws_monthly, devices, start_date_str, end_date_str)
        
        # 生成文件名
        if len(customer_groups) == 1:
            customer_name = next(iter(customer_groups.keys()))
            safe_customer_name = customer_name.replace("/", "_").replace("\\", "_")
            filename = f"{safe_customer_name}_客户对账单_{start_date_str}_to_{end_date_str}.xlsx"
        else:
            filename = f"多客户对账单汇总_{start_date_str}_to_{end_date_str}.xlsx"
        
        output_path = Path(tempfile.gettempdir()) / filename
        wb.save(output_path)
        
        return output_path
    
    def _update_main_statement_sheet(
        self,
        ws,
        customer_name: str,
        devices: List[Dict[str, Any]],
        start_date_str: str,
        end_date_str: str
    ) -> None:
        """
        更新"中润对账单"sheet
        
        格式：
        行5: 客户名称、月份
        行8: 表头（序号、油品名称/型号、设备编码、本月总升数计量（L）、备注）
        行9开始: 数据行
        """
        # 更新客户名称和月份（行5）
        # 格式: 客户名称：XXX    月份：X月份（YYYY.M.D-YYYY.M.D）
        from datetime import datetime
        
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        month_str = f"{start_date.month}月份（{start_date.year}.{start_date.month}.{start_date.day}-{end_date.year}.{end_date.month}.{end_date.day}）"
        
        # 行5: 客户名称和月份
        ws.cell(row=5, column=1).value = f"客户名称：{customer_name}"
        ws.cell(row=5, column=7).value = f"月份：{month_str}"
        
        # 清空数据行（从行9开始）
        data_start_row = 9
        if ws.max_row >= data_start_row:
            ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)
        
        # 写入数据行
        for idx, device in enumerate(devices, start=1):
            row = data_start_row + idx - 1
            ws.cell(row=row, column=1).value = idx  # 序号
            ws.cell(row=row, column=2).value = device.get('oil_name', '未知油品')  # 油品名称/型号
            ws.cell(row=row, column=5).value = device.get('device_code', '')  # 设备编码
            ws.cell(row=row, column=6).value = device.get('total_order_volume', 0)  # 本月总升数计量（L）
            ws.cell(row=row, column=8).value = device.get('remark', '')  # 备注（可选）
        
        # 添加"以下空白"行（如果有空行）
        if len(devices) < 20:  # 假设最多20行数据
            ws.cell(row=data_start_row + len(devices), column=2).value = "以下空白"
    
    def _update_daily_detail_sheet(
        self,
        ws,
        devices: List[Dict[str, Any]],
        start_date_str: str,
        end_date_str: str
    ) -> None:
        """
        更新"每日用量明细"sheet
        
        格式：
        行2-3: 初始日期、结束日期
        行5开始: 每日用量数据
        """
        from datetime import datetime
        
        # 更新日期范围（行2-3）
        ws.cell(row=2, column=7).value = "初始日期"
        ws.cell(row=2, column=8).value = "结束日期"
        ws.cell(row=3, column=7).value = f"{start_date_str} 00:00:00"
        ws.cell(row=3, column=8).value = f"{end_date_str} 00:00:00"
        
        # TODO: 填充每日用量明细数据
        # 这里需要根据实际业务需求填充每日数据
        logger.info("每日用量明细sheet已更新日期范围，详细数据填充待实现")
    
    def _update_monthly_comparison_sheet(
        self,
        ws,
        devices: List[Dict[str, Any]],
        start_date_str: str,
        end_date_str: str
    ) -> None:
        """
        更新"每月用量对比"sheet
        
        格式：
        包含月度对比数据
        """
        # TODO: 填充每月用量对比数据
        # 这里需要根据实际业务需求填充月度对比数据
        logger.info("每月用量对比sheet更新待实现")
    
    def _group_devices_by_customer(self, customer_devices_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """按客户分组设备"""
        customer_groups = {}
        
        for device_data in customer_devices_data:
            customer_name = device_data.get("customer_name", "未知客户")
            
            if customer_name not in customer_groups:
                customer_groups[customer_name] = []
            
            customer_groups[customer_name].append(device_data)
        
        return customer_groups
    
    def _generate_customer_statement_excel(self, 
                                         customer_groups: Dict[str, List[Dict[str, Any]]],
                                         start_date_str: str,
                                         end_date_str: str) -> Path:
        """生成客户对账单Excel文件"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import tempfile
        from pathlib import Path
        
        wb = Workbook()
        
        # 为每个客户生成对账单
        for customer_name, devices in customer_groups.items():
            # 创建客户工作表
            if customer_name == "未知客户":
                ws_name = "未知客户对账单"
            else:
                ws_name = f"{customer_name}对账单"
            
            if len(ws_name) > 31:  # Excel工作表名称最大31字符
                ws_name = ws_name[:31]
            
            ws = wb.create_sheet(ws_name)
            
            # 设置主页内容
            self._setup_customer_statement_sheet(ws, customer_name, devices, start_date_str, end_date_str)
        
        # 删除默认的工作表
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # 生成文件名
        if len(customer_groups) == 1:
            customer_name = next(iter(customer_groups.keys()))
            safe_customer_name = customer_name.replace("/", "_").replace("\\", "_")
            filename = f"{safe_customer_name}_客户对账单_{start_date_str}_to_{end_date_str}.xlsx"
        else:
            filename = f"多客户对账单汇总_{start_date_str}_to_{end_date_str}.xlsx"
        
        output_path = Path(tempfile.gettempdir()) / filename
        wb.save(output_path)
        
        return output_path
    
    def _setup_customer_statement_sheet(self, ws, customer_name: str, devices: List[Dict[str, Any]], 
                                        start_date_str: str, end_date_str: str) -> None:
        """设置客户对账单工作表"""
        # 主标题
        ws.append([f"{customer_name} 客户对账单"])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # 客户信息
        ws.append(["客户名称:", customer_name])
        ws.append(["日期范围:", f"{start_date_str} 至 {end_date_str}"])
        ws.append(["生成时间:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])  # 空行
        
        # 设备数据表格标题
        headers = ["设备编号", "设备名称", "油品名称", "期初库存(L)", "期末库存(L)", 
                  "加油量(L)", "订单量(L)", "实际消耗(L)", "误差(L)"]
        ws.append(headers)
        
        # 设置表头样式
        for col in range(1, len(headers) + 1):
            header_cell = ws.cell(row=ws.max_row, column=col)
            header_cell.font = Font(bold=True)
            header_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # 写入设备数据
        for device in devices:
            row_data = [
                device.get('device_code', ''),
                device.get('device_name', ''),
                device.get('oil_name', ''),
                device.get('beginning_inventory', 0),
                device.get('ending_inventory', 0),
                device.get('refill_volume', 0),
                device.get('order_volume', 0),
                device.get('actual_consumption', 0),
                device.get('error', 0)
            ]
            ws.append(row_data)
        
        # 添加备注信息
        self._add_notes_to_statement(ws, len(devices))
        
        # 调整格式
        self._adjust_statement_formatting(ws)
    
    def _add_notes_to_statement(self, ws, device_count: int) -> None:
        """添加备注信息到对账单"""
        # 空行
        ws.append([])
        ws.append([])
        
        # 备注标题
        ws.append(["备注:"])
        note_title_cell = ws.cell(row=ws.max_row, column=1)
        note_title_cell.font = Font(bold=True)
        
        # 备注内容
        notes = [
            f"1. 本对账单包含 {device_count} 台设备的数据",
            "2. 实际消耗 = (期初库存 - 期末库存 + 加油量)",
            "3. 误差 = 实际消耗 - 订单量",
            "4. 正误差表示中润亏损，负误差表示客户亏损",
            "5. 数据来源: 设备监控系统"
        ]
        
        for note in notes:
            ws.append([note])
    
    def _adjust_statement_formatting(self, ws) -> None:
        """调整对账单格式"""
        # 设置列宽
        column_widths = {
            'A': 12,  # 设备编号
            'B': 15,  # 设备名称
            'C': 15,  # 油品名称
            'D': 12,  # 期初库存
            'E': 12,  # 期末库存
            'F': 10,  # 加油量
            'G': 10,  # 订单量
            'H': 12,  # 实际消耗
            'I': 10   # 误差
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 设置数字格式
        for row in range(6, ws.max_row + 1):  # 从数据行开始
            for col in range(4, 10):  # D-I列（数字列）
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
    
    def _group_failed_devices_by_customer(self, failed_devices: List[str], 
                                         customer_devices_data: List[Dict[str, Any]]) -> Dict[str, List[str]]:
        """按客户分组失败设备"""
        # 创建设备到客户的映射
        device_to_customer = {}
        for device_data in customer_devices_data:
            device_to_customer[device_data['device_code']] = device_data['customer_name']
        
        # 分组失败设备
        failed_by_customer = {}
        for device_code in failed_devices:
            customer_name = device_to_customer.get(device_code, "未知客户")
            if customer_name not in failed_by_customer:
                failed_by_customer[customer_name] = []
            failed_by_customer[customer_name].append(device_code)
        
        return failed_by_customer

    def _generate_inventory_excel_with_chart(self, inventory_data, device_code, oil_name, customer_name, start_date, end_date):
        """
        生成库存报表Excel文件（基于development-copy分支逻辑，确保与原分支效果一致）
        
        Args:
            inventory_data: 库存数据列表
            device_code: 设备编码
            oil_name: 油品名称
            customer_name: 客户名称
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            Excel文件路径
        """
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.axis import DateAxis
        from openpyxl.chart.marker import Marker
        from openpyxl.chart.shapes import GraphicalProperties
        from openpyxl.drawing.line import LineProperties
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        import tempfile
        from pathlib import Path
        from datetime import datetime, timedelta
        
        # 数据补全逻辑（与原分支保持一致）
        def complete_data(data_list):
            """补全日期范围内的数据，与原分支逻辑一致"""
            if not data_list:
                return []
                
            # 将数据转换为字典格式
            data_dict = {}
            for record in data_list:
                date = record.get('report_date')
                if isinstance(date, datetime):
                    date_key = date.date()
                else:
                    date_key = date
                data_dict[date_key] = record.get('end_of_day_inventory', 0)
            
            # 补全日期范围内的数据
            complete_data_list = []
            current_date = start_date.date() if isinstance(start_date, datetime) else start_date
            end_date_dt = end_date.date() if isinstance(end_date, datetime) else end_date
            last_inventory = 0
            
            # 如果有数据，使用第一条数据的库存值作为默认值
            if data_list:
                last_inventory = data_list[0].get('end_of_day_inventory', 0)
            
            while current_date <= end_date_dt:
                current_inventory = data_dict.get(current_date, last_inventory)
                complete_data_list.append({
                    'report_date': current_date,
                    'end_of_day_inventory': current_inventory,
                    'customer_name': customer_name,
                    'oil_name': oil_name
                })
                last_inventory = current_inventory
                current_date += timedelta(days=1)
            
            return complete_data_list
        
        # 补全数据
        complete_inventory_data = complete_data(inventory_data)
        
        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = "库存数据"
        
        # 设置报表标题（与原分支完全一致）
        oil_name_str = f" {oil_name} " if oil_name else " "
        title = f"{device_code}{oil_name_str}每日库存余量变化趋势({start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')})"
        ws.append([title])
        # 将合并单元格的宽度增加到20列（与原分支一致）
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=1).font = Font(size=14, bold=True)
        
        # 添加数据列标题（与原分支完全一致）
        ws.append(["日期", "原油剩余量(L)"])
        
        # 写入补全后的数据（与原分支完全一致）
        for record in complete_inventory_data:
            ws.append([
                record['report_date'].strftime('%Y-%m-%d') if isinstance(record['report_date'], datetime) else record['report_date'],
                record['end_of_day_inventory']
            ])
        
        # 调整列宽（与原分支完全一致）
        ws.column_dimensions["A"].width = 12  # 日期列宽度
        ws.column_dimensions["B"].width = 12  # 原油剩余量(L)列宽度
        
        # 创建图表（与原分支完全一致）
        if len(complete_inventory_data) > 1:
            chart = LineChart()
            chart.title = "每日库存余量变化趋势"
            chart.style = 13
            chart.y_axis.title = "原油剩余量(L)"
            chart.x_axis.title = "日期"
            
            # 设置图表显示数据标签
            chart.x_axis.tickLblSkip = 3  # 每隔3个标签显示一个（与原分支一致）
            chart.x_axis.tickLblPos = "low"  # 将标签位置调整到底部（与原分支一致）
            chart.x_axis.textRotation = 0  # 将文本旋转角度设为0度（水平显示，与原分支一致）
            
            # 设置数据范围（与原分支完全一致）
            data_range = Reference(
                ws, min_col=2, min_row=2, max_col=2, max_row=len(complete_inventory_data) + 2
            )
            dates = Reference(ws, min_col=1, min_row=3, max_row=len(complete_inventory_data) + 2)
            
            # 添加数据到图表（与原分支完全一致）
            chart.add_data(data_range, titles_from_data=True)
            chart.set_categories(dates)
            
            # 应用图表样式（与原分支完全一致）
            chart.series[0].marker = Marker(symbol="circle", size=8)
            
            # 恢复图表到初始大小（与原分支完全一致）
            chart.width = 30
            chart.height = 15
            
            # 添加图表到工作表，从E5开始绘制（与原分支完全一致）
            ws.add_chart(chart, "E5")
        
        # 生成文件名（与原分支完全一致）
        start_date_str = start_date.strftime("%Y%m%d")
        end_date_str = end_date.strftime("%Y%m%d")
        
        # 确保customer_name不为None，避免文件名中出现"None"
        if customer_name is None:
            customer_name = '未知客户'
        if oil_name is None:
            oil_name = '未知油品'
            
        filename = f"{customer_name}_{device_code}_{start_date_str}_to_{end_date_str}_库存报表.xlsx"
        output_path = Path(tempfile.gettempdir()) / filename
        
        # 保存文件
        wb.save(output_path)
        logger.info(f"库存报表已生成: {output_path}")
        
        return output_path

    async def _generate_refueling_details_report(
        self,
        device_codes: List[str],
        start_date_str: str,
        end_date_str: str,
    ) -> Tuple[Path, List[str]]:
        """
        生成加油详情报表（与原加注明细导出功能保持一致）
        
        Args:
            device_codes: 设备编码列表
            start_date_str: 开始日期
            end_date_str: 结束日期
            
        Returns:
            (生成的Excel文件路径, 警告信息列表)
        """
        logger.info(f"开始生成加油详情报表: 设备={device_codes}, 日期范围={start_date_str} 至 {end_date_str}")
        
        warnings = []
        
        try:
            # 获取加注明细数据（使用正确的数据源）
            device_codes_tuple = tuple(sorted(device_codes))
            raw_refueling_data = await self.device_repo.get_refueling_details_raw_data(
                device_codes_tuple, start_date_str, end_date_str
            )
            
            if not raw_refueling_data:
                raise ValueError("在指定日期范围内所有选定设备均未找到任何加注数据。")
            
            # 按设备分组数据
            data_by_device = defaultdict(list)
            for row in raw_refueling_data:
                data_by_device[row['device_code']].append(row)
            
            # 处理每个设备的数据（优化版，直接使用原始数据）
            refueling_data_list = []
            failed_devices = []
            
            for device_code in device_codes:
                device_data = data_by_device.get(device_code)
                if not device_data:
                    failed_devices.append(device_code)
                    warnings.append(f"设备 {device_code} 没有加注数据，已跳过。")
                    continue
                
                try:
                    # 直接使用原始数据，避免不必要的字段映射
                    refueling_data_list.extend(device_data)
                        
                except Exception as e:
                    failed_devices.append(device_code)
                    error_msg = f"处理设备 {device_code} 加注数据时发生错误: {str(e)}"
                    warnings.append(error_msg)
                    logger.error(error_msg, exc_info=True)
            
            if not refueling_data_list:
                raise ValueError("所有选定设备均未能成功提取加注数据。")
            
            # 生成加油详情报表
            output_path = self._generate_refueling_details_excel(
                refueling_data_list, start_date_str, end_date_str
            )
            
            logger.info(f"加油详情报表已生成: {output_path}")
            
            # 处理失败设备警告
            if failed_devices:
                warnings.append(f"以下设备处理失败: {', '.join(failed_devices)}")
            
            return output_path, warnings
            
        except Exception as e:
            error_msg = f"生成加油详情报表失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ValueError(error_msg)
    
    def _generate_refueling_details_excel(self, 
                                        refueling_data: List[Dict[str, Any]],
                                        start_date_str: str,
                                        end_date_str: str) -> Path:
        """生成加油详情Excel文件（优化版，基于development-copy分支最佳实践）"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import tempfile
        from pathlib import Path
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "加注明细"
            
            # 数据表格标题（使用查询模板中的原始字段名）
            headers = [
                "订单序号", "加注时间", "油品序号", "油品名称", 
                "水油比：水值", "水油比：油值", "水加注值", "油加注值",
                "原油剩余量", "原油剩余比例", "油加设量", "是否结算：1=待结算 2=待生效 3=已结算", "加注模式：1=近程自动 2=远程自动 3=手动"
            ]
            ws.append(headers)
            
            # 设置表头样式
            for col in range(1, len(headers) + 1):
                header_cell = ws.cell(row=1, column=col)
                header_cell.font = Font(bold=True)
                header_cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # 写入数据（直接使用原始数据，字段名与查询配置一致）
            for record in refueling_data:
                row_data = [
                    record.get("订单序号", ''),
                    record.get("加注时间", ''),
                    record.get("油品序号", ''),
                    record.get("油品名称", '未知油品'),
                    record.get("水油比：水值", 0),
                    record.get("水油比：油值", 0),
                    record.get("水加注值", 0),
                    record.get("油加注值", 0),
                    record.get("原油剩余量", 0),
                    record.get("原油剩余比例", 0),
                    record.get("油加设量", 0),
                    record.get("是否结算：1=待结算 2=待生效 3=已结算", ''),
                    record.get("加注模式：1=近程自动 2=远程自动 3=手动", '')
                ]
                ws.append(row_data)
            
            # 设置列宽（优化版，加宽加注时间列）
            column_widths = {
                'A': 10,  # 订单序号
                'B': 25,  # 加注时间（加宽以完整显示日期时间）
                'C': 10,  # 油品序号
                'D': 15,  # 油品名称
                'E': 12,  # 水油比：水值
                'F': 12,  # 水油比：油值
                'G': 10,  # 水加注值
                'H': 10,  # 油加注值
                'I': 12,  # 原油剩余量
                'J': 12,  # 原油剩余比例
                'K': 10,  # 油加设量
                'L': 15,  # 是否结算（适当加宽）
                'M': 15   # 加注模式（适当加宽）
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 设置数字格式
            for row in range(2, ws.max_row + 1):  # 从第2行开始（数据行）
                for col in range(5, 11):  # E-J列（数字列）
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.00'
            
            # 设置加注时间列的日期时间格式
            for row in range(2, ws.max_row + 1):  # 从第2行开始（数据行）
                time_cell = ws.cell(row=row, column=2)  # B列（加注时间）
                if time_cell.value:
                    # 设置日期时间格式，确保完整显示
                    time_cell.number_format = 'yyyy-mm-dd hh:mm:ss'
            
            # 生成文件名（按照正确格式：公司名称_设备编码_开始日期_结束日期_加注明细.xlsx）
            # 从第一条记录获取设备编码
            if refueling_data:
                first_record = refueling_data[0]
                device_code = first_record.get('device_code', '未知设备')
                
                # 获取客户名称
                try:
                    from src.core.db_handler import DBHandler
                    db_handler = DBHandler()
                    customer_name = db_handler.get_customer_name_by_device_code(device_code)
                except Exception:
                    customer_name = f"客户_{device_code}"
                
                # 替换日期中的非法字符，确保文件名合法（与development-copy分支保持一致）
                safe_start_date = start_date_str.replace("/", "-").replace("\\", "-")
                safe_end_date = end_date_str.replace("/", "-").replace("\\", "-")
                
                filename = f"{customer_name}_{device_code}_{safe_start_date}_{safe_end_date}_加注明细.xlsx"
            else:
                # 如果没有数据，使用默认格式
                filename = f"加注订单明细_{start_date_str}_至_{end_date_str}.xlsx"
            
            output_path = Path(tempfile.gettempdir()) / filename
            wb.save(output_path)
            
            logger.info(f"加注订单明细Excel文件已生成: {output_path}")
            return output_path
            
        except Exception as e:
            error_msg = f"生成加注订单明细Excel文件失败: {str(e)}"
            logger.error(error_msg, exc_info=True)
            raise ValueError(error_msg)