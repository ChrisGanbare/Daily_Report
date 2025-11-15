"""
客户对账单功能一致性测试
对比重构前后的实现，确保功能完全一致
"""
import pytest
from datetime import datetime, date
from pathlib import Path
import tempfile
from openpyxl import load_workbook
from collections import defaultdict

from src.services.report_service import ReportService
from src.repositories.device_repository import DeviceRepository


class TestCustomerStatementConsistency:
    """客户对账单功能一致性测试"""
    
    @pytest.fixture
    def mock_device_repo(self):
        """模拟设备仓库"""
        class MockDeviceRepository:
            async def get_daily_consumption_raw_data(self, device_codes, start_date, end_date):
                # 模拟返回多天的数据
                return [
                    {
                        'device_code': 'DEV001',
                        'customer_name': '测试客户A',
                        'oil_name': '测试油品1',
                        'device_name': '设备1',
                        'report_date': date(2025, 1, 1),
                        'prev_day_inventory': 1000.0,
                        'end_of_day_inventory': 950.0,
                        'daily_refill': 50.0,
                        'daily_order_volume': 100.0
                    },
                    {
                        'device_code': 'DEV001',
                        'customer_name': '测试客户A',
                        'oil_name': '测试油品1',
                        'device_name': '设备1',
                        'report_date': date(2025, 1, 2),
                        'prev_day_inventory': 950.0,
                        'end_of_day_inventory': 900.0,
                        'daily_refill': 0.0,
                        'daily_order_volume': 50.0
                    },
                    {
                        'device_code': 'DEV002',
                        'customer_name': '测试客户A',
                        'oil_name': '测试油品2',
                        'device_name': '设备2',
                        'report_date': date(2025, 1, 1),
                        'prev_day_inventory': 800.0,
                        'end_of_day_inventory': 750.0,
                        'daily_refill': 0.0,
                        'daily_order_volume': 50.0
                    }
                ]
        return MockDeviceRepository()
    
    @pytest.fixture
    def report_service(self, mock_device_repo):
        """创建报表服务实例"""
        return ReportService(mock_device_repo)
    
    @pytest.mark.asyncio
    async def test_calculate_consumption_for_statement_single_day(self, report_service):
        """
        测试单日消耗计算逻辑
        验证计算公式：实际消耗 = (期初库存 - 期末库存 + 加油量) * 桶数
        """
        device_data = [{
            'prev_day_inventory': 1000.0,
            'end_of_day_inventory': 950.0,
            'daily_refill': 50.0,
            'daily_order_volume': 100.0
        }]
        
        barrel_count = 1
        result = report_service._calculate_consumption_for_statement(device_data, barrel_count)
        
        # 验证计算逻辑
        expected_inventory_consumption = max(0, (1000.0 - 950.0) + 50.0)  # = 100.0
        expected_actual_consumption = expected_inventory_consumption * barrel_count  # = 100.0
        expected_error = expected_actual_consumption - 100.0  # = 0.0
        
        assert result['beginning_inventory'] == 1000.0
        assert result['ending_inventory'] == 950.0
        assert result['refill_volume'] == 50.0
        assert result['order_volume'] == 100.0
        assert result['actual_consumption'] == expected_actual_consumption
        assert result['error'] == expected_error
    
    @pytest.mark.asyncio
    async def test_calculate_consumption_for_statement_multiple_days(self, report_service):
        """
        测试多日消耗计算逻辑
        验证：应该汇总所有日期的数据，而不是只使用第一条记录
        """
        device_data = [
            {
                'report_date': date(2025, 1, 1),
                'prev_day_inventory': 1000.0,
                'end_of_day_inventory': 950.0,
                'daily_refill': 50.0,
                'daily_order_volume': 100.0
            },
            {
                'report_date': date(2025, 1, 2),
                'prev_day_inventory': 950.0,
                'end_of_day_inventory': 900.0,
                'daily_refill': 0.0,
                'daily_order_volume': 50.0
            }
        ]
        
        barrel_count = 1
        result = report_service._calculate_consumption_for_statement(device_data, barrel_count)
        
        # 应该汇总所有数据：
        # 期初库存 = 第一条的prev_day_inventory = 1000.0
        # 期末库存 = 最后一条的end_of_day_inventory = 900.0
        # 总加油量 = 50.0 + 0.0 = 50.0
        # 总订单量 = 100.0 + 50.0 = 150.0
        # 实际消耗 = (1000 - 900 + 50) * 1 = 150.0
        # 误差 = 150.0 - 150.0 = 0.0
        
        assert result['beginning_inventory'] == 1000.0
        assert result['ending_inventory'] == 900.0, f"期末库存应该是900.0，实际是{result['ending_inventory']}"
        assert result['refill_volume'] == 50.0, f"总加油量应该是50.0，实际是{result['refill_volume']}"
        assert result['order_volume'] == 150.0, f"总订单量应该是150.0，实际是{result['order_volume']}"
        assert result['actual_consumption'] == 150.0, f"实际消耗应该是150.0，实际是{result['actual_consumption']}"
        assert result['error'] == 0.0, f"误差应该是0.0，实际是{result['error']}"
    
    @pytest.mark.asyncio
    async def test_customer_statement_excel_format(self, report_service, mock_device_repo):
        """
        测试客户对账单Excel格式
        验证表头、数据格式、备注等是否与原分支一致
        """
        # 生成对账单
        device_codes = ['DEV001', 'DEV002']
        start_date = '2025-01-01'
        end_date = '2025-01-02'
        
        output_path, warnings = await report_service._generate_customer_statement_report(
            device_codes, start_date, end_date
        )
        
        # 验证文件存在
        assert output_path.exists()
        
        # 读取Excel文件
        wb = load_workbook(output_path)
        
        # 验证工作表数量（应该为每个客户一个工作表）
        assert len(wb.sheetnames) >= 1
        
        # 验证第一个工作表的内容
        ws = wb.active
        
        # 验证标题
        assert ws.cell(row=1, column=1).value is not None
        assert "对账单" in str(ws.cell(row=1, column=1).value)
        
        # 验证表头
        headers = ["设备编号", "设备名称", "油品名称", "期初库存(L)", "期末库存(L)", 
                  "加油量(L)", "订单量(L)", "实际消耗(L)", "误差(L)"]
        
        # 找到表头行（应该在客户信息之后）
        header_row = None
        for row_idx in range(1, min(10, ws.max_row + 1)):
            first_cell = ws.cell(row=row_idx, column=1).value
            if first_cell == "设备编号":
                header_row = row_idx
                break
        
        assert header_row is not None, "未找到表头行"
        
        # 验证表头内容
        for col_idx, expected_header in enumerate(headers, start=1):
            actual_header = ws.cell(row=header_row, column=col_idx).value
            assert actual_header == expected_header, f"表头第{col_idx}列不匹配: 期望'{expected_header}', 实际'{actual_header}'"
        
        # 验证数据行格式（数字格式）
        data_start_row = header_row + 1
        if data_start_row <= ws.max_row:
            for col_idx in range(4, 10):  # D-I列（数字列）
                cell = ws.cell(row=data_start_row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    # 验证数字格式
                    assert cell.number_format in ['0.00', 'General'], f"第{col_idx}列数字格式不正确: {cell.number_format}"
        
        # 验证备注部分
        notes_found = False
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=1).value
            if cell_value and "备注" in str(cell_value):
                notes_found = True
                break
        
        assert notes_found, "未找到备注部分"
        
        # 清理临时文件（先关闭工作簿）
        wb.close()
        if output_path.exists():
            try:
                output_path.unlink()
            except PermissionError:
                # Windows上文件可能被占用，忽略错误
                pass
    
    @pytest.mark.asyncio
    async def test_customer_statement_date_validation(self, report_service):
        """
        测试日期范围验证
        客户对账单的日期范围不能超过1个月
        """
        device_codes = ['DEV001']
        
        # 测试有效日期范围（1个月内）
        start_date = '2025-01-01'
        end_date = '2025-01-31'
        
        # 应该成功（不抛出异常）
        try:
            output_path, warnings = await report_service._generate_customer_statement_report(
                device_codes, start_date, end_date
            )
            # 如果成功，清理文件
            if output_path.exists():
                output_path.unlink()
        except ValueError as e:
            # 如果因为数据为空而失败，这是正常的
            if "未找到任何数据" not in str(e):
                raise
        
        # 测试无效日期范围（超过1个月）
        start_date = '2025-01-01'
        end_date = '2025-03-01'
        
        with pytest.raises(ValueError, match="日期范围不能超过1个月"):
            await report_service._generate_customer_statement_report(
                device_codes, start_date, end_date
            )
    
    @pytest.mark.asyncio
    async def test_customer_statement_grouping(self, report_service, mock_device_repo):
        """
        测试客户分组逻辑
        验证多个设备是否按客户正确分组
        """
        device_codes = ['DEV001', 'DEV002']
        start_date = '2025-01-01'
        end_date = '2025-01-02'
        
        output_path, warnings = await report_service._generate_customer_statement_report(
            device_codes, start_date, end_date
        )
        
        # 读取Excel文件
        wb = load_workbook(output_path)
        
        # 验证工作表数量（应该为每个客户一个工作表）
        # 由于两个设备都是"测试客户A"，应该只有一个工作表
        assert len(wb.sheetnames) == 1
        
        # 验证工作表名称
        ws_name = wb.sheetnames[0]
        assert "测试客户A" in ws_name or "对账单" in ws_name
        
        # 清理临时文件（先关闭工作簿）
        wb.close()
        if output_path.exists():
            try:
                output_path.unlink()
            except PermissionError:
                # Windows上文件可能被占用，忽略错误
                pass
    
    def test_calculate_consumption_with_barrel_count(self, report_service):
        """
        测试桶数计算
        验证桶数是否正确应用到消耗计算中
        """
        device_data = [{
            'prev_day_inventory': 1000.0,
            'end_of_day_inventory': 950.0,
            'daily_refill': 50.0,
            'daily_order_volume': 100.0
        }]
        
        barrel_count = 2  # 2桶设备
        result = report_service._calculate_consumption_for_statement(device_data, barrel_count)
        
        # 单桶消耗 = (1000 - 950 + 50) = 100.0
        # 实际消耗 = 100.0 * 2 = 200.0
        expected_actual_consumption = 200.0
        expected_error = expected_actual_consumption - 100.0  # = 100.0
        
        assert result['actual_consumption'] == expected_actual_consumption
        assert result['error'] == expected_error
