import pytest
import os
import datetime
import shutil
from openpyxl import load_workbook

# 将src目录添加到Python路径
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.core.data_manager import ReportDataManager
from src.core.consumption_error_handler import DailyConsumptionErrorReportGenerator


def robust_rmtree(path, max_retries=3, delay=0.5):
    """增强的删除函数，支持重试和解除只读属性"""
    import time
    
    def _onerror(func, path, exc_info):
        """处理只读文件或权限问题"""
        if not os.path.exists(path):
            return
        os.chmod(path, 0o777)
        func(path)
    
    for i in range(max_retries):
        try:
            shutil.rmtree(path, onerror=_onerror)
            return
        except PermissionError as e:
            print(f"删除 '{path}' 时遇到权限错误，将在 {delay} 秒后重试... ({i+1}/{max_retries})")
            time.sleep(delay)
    # 最终尝试强制删除
    shutil.rmtree(path, ignore_errors=True)


@pytest.fixture(scope="module")
def temp_output_dir():
    """
    创建一个临时的测试输出目录，并在所有测试结束后清理。
    'scope="module"' 确保这个 fixture 在整个测试模块中只运行一次。
    """
    dir_path = "temp_test_output_daily"
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    yield dir_path  # yield 关键字将目录路径提供给测试函数
    # 测试结束后，清理目录
    if os.path.exists(dir_path):
        robust_rmtree(dir_path)


@pytest.fixture
def mock_db_handler():
    """
    使用 pytest-mock 模拟数据库处理器。
    我们不需要真实的数据库连接，因为我们将直接提供原始数据。
    """
    return True # 返回一个简单的占位符，因为我们会在测试中用 mocker.patch 替换方法


def test_daily_report_with_barrel_count(temp_output_dir, mock_db_handler, mocker):
    """
    测试场景:
    - 模拟一个有3个油桶的设备。
    - 模拟从数据库获取的原始数据。
    - 验证 data_manager 计算出的 `inventory_consumption` 是否正确乘以了桶数。
    - 验证 consumption_error_handler 是否将正确的值写入了Excel。
    """
    # 1. 准备 (Arrange)
    # 模拟设备信息和日期
    device_code = "TEST_DEVICE_01"
    barrel_count = 3
    test_date = datetime.date(2025, 7, 1)

    # 模拟从数据库获取的原始数据 (raw_data_content)
    # 包含前一天和当天的记录
    columns = ["加注时间", "原油剩余量", "油加注值"]
    raw_data_content = [
        (datetime.datetime(2025, 6, 30, 23, 0, 0), 1000.0, 0.0),  # 前一天的最后一条记录
        (datetime.datetime(2025, 7, 1, 8, 0, 0), 990.0, 50.0),   # 当天第一条记录
        (datetime.datetime(2025, 7, 1, 17, 0, 0), 980.0, 126.05) # 当天最后一条记录
    ]
    # 模拟 data_manager.fetch_raw_data 的返回值
    mock_raw_data = ([], columns, raw_data_content)

    # 预期结果
    # 单桶液位消耗 = (前日结束库存 - 当日结束库存) + 当日加油量
    # (1000 - 980) + 0 = 20
    # 真实消耗总量 = 20 * 3 (桶数) = 60
    expected_consumption = 60.0
    expected_order_total = 50.0 + 126.05 # 176.05

    # 2. 执行 (Act)
    # 步骤 A: 使用 data_manager 计算误差数据
    data_manager = ReportDataManager(mock_db_handler)
    # 模拟 fetch_raw_data 方法，使其返回我们准备好的数据
    mocker.patch.object(data_manager, 'fetch_raw_data', return_value=mock_raw_data)
    error_data = data_manager.calculate_daily_errors(mock_raw_data, test_date.strftime('%Y-%m-%d'), test_date.strftime('%Y-%m-%d'), barrel_count)

    # 步骤 B: 使用 consumption_error_handler 生成报表
    report_generator = DailyConsumptionErrorReportGenerator()
    output_file_path = os.path.join(temp_output_dir, "test_daily_report_with_barrels.xlsx")
    report_generator.generate_report(
        inventory_data=[(test_date, 980.0)], # 简化库存数据
        error_data=error_data,
        output_file_path=output_file_path,
        device_code=device_code,
        start_date=test_date,
        end_date=test_date,
        oil_name="Test Oil",
        barrel_count=barrel_count
    )

    # 3. 断言 (Assert)
    # 断言 A: 检查 data_manager 的计算结果是否正确
    calculated_consumption = error_data['daily_consumption'].get(test_date)
    assert calculated_consumption == pytest.approx(expected_consumption), \
        f"data_manager 计算的库存消耗总量应为 {expected_consumption}，但实际为 {calculated_consumption}"

    # 断言 B: 检查生成的Excel文件内容是否正确
    assert os.path.exists(output_file_path), "报表文件未生成"
    wb = load_workbook(output_file_path)
    ws = wb["消耗误差分析"]

    # 数据从第3行开始，"库存消耗总量(L)"是第4列(D)
    consumption_cell_value = ws.cell(row=3, column=4).value
    assert consumption_cell_value == pytest.approx(expected_consumption), \
        f"Excel中的库存消耗总量应为 {expected_consumption}，但实际为 {consumption_cell_value}"

    # 验证计算规则说明
    formula_cell_value = ws.cell(row=35, column=8).value
    assert f"* {barrel_count} (桶数)" in formula_cell_value, "计算规则说明中未正确显示桶数信息"

    # 关键修复：关闭工作簿以释放文件句柄
    wb.close()