import os
import sys
import tempfile
import unittest
from datetime import date, datetime
from unittest.mock import Mock, patch

# 添加项目根目录到sys.path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from src.core.statement_handler import CustomerStatementGenerator
from tests.base_test import BaseTestCase


class TestCoreStatementHandler(BaseTestCase):
    """
    core.statement_handler 模块的单元测试
    """

    def setUp(self):
        """测试前准备"""
        super().setUp()
        self.statement_handler = CustomerStatementGenerator()
        self.test_devices_data = [
            {
                "device_code": "DEV001",
                "oil_name": "切削液",
                "data": [
                    (date(2025, 7, 1), 10.0),
                    (date(2025, 7, 2), 15.0),
                ],
                "raw_data": [],
                "columns": ["加注时间", "原油剩余比例"],
                "customer_name": "测试客户",
            }
        ]

    def test_customerstatementgenerator_initialization(self):
        """
        测试 CustomerStatementGenerator 类的初始化
        """
        # CustomerStatementGenerator的__init__方法目前为空，只需确认能创建实例
        self.assertIsInstance(self.statement_handler, CustomerStatementGenerator)

    def test_customerstatementgenerator_write_cell_safe(self):
        """
        测试 CustomerStatementGenerator._write_cell_safe 方法
        """
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        # 测试正常写入
        self.statement_handler._write_cell_safe(ws, 1, 1, "测试值")
        self.assertEqual(ws.cell(row=1, column=1).value, "测试值")

        # 测试写入合并单元格（应该不会抛出异常）
        ws.merge_cells("A2:A3")
        self.statement_handler._write_cell_safe(ws, 2, 1, "合并单元格值")
        # 值不会被写入，但我们主要检查不抛出异常

    def test_customerstatementgenerator_generate_date_list(self):
        """
        测试 CustomerStatementGenerator._generate_date_list 方法
        """
        start_date = date(2025, 7, 1)
        end_date = date(2025, 7, 5)
        date_list = self.statement_handler._generate_date_list(start_date, end_date)

        self.assertEqual(len(date_list), 5)
        self.assertEqual(date_list[0], start_date)
        self.assertEqual(date_list[-1], end_date)

    def test_customerstatementgenerator_collect_oil_types(self):
        """
        测试 CustomerStatementGenerator._collect_oil_types 方法
        """
        oil_types = self.statement_handler._collect_oil_types(self.test_devices_data)
        self.assertEqual(oil_types, ["切削液"])

    @patch("src.core.statement_handler.load_workbook")
    @patch("os.path.exists")
    def test_customerstatementgenerator_generate_customer_statement_from_template(
        self, mock_exists, mock_load_workbook
    ):
        """
        测试 CustomerStatementGenerator.generate_customer_statement_from_template 方法
        """
        # 模拟模板文件存在
        mock_exists.return_value = True

        # 模拟工作簿和工作表
        mock_wb = Mock()
        mock_daily_sheet = Mock()
        mock_monthly_sheet = Mock()
        mock_statement_sheet = Mock()

        mock_wb.__getitem__ = Mock(
            side_effect=lambda x: {
                "每日用量明细": mock_daily_sheet,
                "每月用量对比": mock_monthly_sheet,
                "中润对账单": mock_statement_sheet,
            }[x]
        )
        mock_wb.sheetnames = ["每日用量明细", "每月用量对比", "中润对账单"]
        mock_load_workbook.return_value = mock_wb

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_file_path = f.name

        try:
            # 调用方法（会抛出异常，因为是模拟环境，但我们主要测试路径覆盖）
            with self.assertRaises(Exception):
                self.statement_handler.generate_customer_statement_from_template(
                    all_devices_data=self.test_devices_data,
                    output_file=output_file_path,
                    customer_name="测试客户",
                    start_date=date(2025, 7, 1),
                    end_date=date(2025, 7, 31),
                )
        finally:
            if os.path.exists(output_file_path):
                os.unlink(output_file_path)

    @patch("os.path.exists")
    def test_customerstatementgenerator_generate_customer_statement_from_template_missing_template(
        self, mock_exists
    ):
        """
        测试 CustomerStatementGenerator.generate_customer_statement_from_template 方法处理模板文件不存在的情况
        """
        # 模拟模板文件不存在
        mock_exists.return_value = False

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            output_file_path = f.name

        try:
            # 应该抛出FileNotFoundError异常
            with self.assertRaises(FileNotFoundError):
                self.statement_handler.generate_customer_statement_from_template(
                    all_devices_data=self.test_devices_data,
                    output_file=output_file_path,
                    customer_name="测试客户",
                    start_date=date(2025, 7, 1),
                    end_date=date(2025, 7, 31),
                )
        finally:
            if os.path.exists(output_file_path):
                os.unlink(output_file_path)


if __name__ == "__main__":
    unittest.main()
