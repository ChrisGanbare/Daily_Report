
import os
import sys
import tempfile
import unittest
from datetime import date

# 将项目根目录添加到Python路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from openpyxl import load_workbook
from src.core.consumption_error_handler import (
    MonthlyConsumptionErrorReportGenerator,
)
from tests.base_test import BaseTestCase


class TestCoreConsumptionErrorHandler(BaseTestCase):
    """
    consumption_error_handler 模块的单元测试。
    """

    def setUp(self):
        """测试前准备。"""
        super().setUp()
        self.monthly_handler = MonthlyConsumptionErrorReportGenerator()

    def test_monthly_report_generation_normal(self):
        """
        TC-FUNC-01: 测试月度报表的正常生成（日期范围 > 31天）。
        验证月度分组、数据点和计算的正确性，特别是针对范围内的最后一个月。
        """
        # 1. 准备：准备模拟数据和预期结果
        device_code = "DEV-M-01"
        start_date = date(2025, 6, 26)
        end_date = date(2025, 10, 25)
        oil_name = "测试油品"

        # 模拟误差数据，模拟 data_manager.calculate_monthly_errors 的输出
        mock_error_data = {
            'monthly_order_totals': {
                '2025-07-25': 1000,
                '2025-08-25': 1100,
                '2025-09-25': 1200,
                '2025-10-25': 1300,  # 关键：确保最后一个月有数据
            },
            'monthly_consumption': {
                '2025-07-25': {'value': 980},
                '2025-08-25': {'value': 1150},
                '2025-09-25': {'value': 1200},
                '2025-10-25': {'value': 1250},  # 关键：确保最后一个月有数据
            },
            'monthly_shortage_errors': {
                '2025-08-25': {'value': 50},  # 消耗量 > 订单量
            },
            'monthly_excess_errors': {
                '2025-07-25': {'value': 20},   # 订单量 > 消耗量
                '2025-10-25': {'value': 50},
            },
        }

        # 在临时目录中创建一个唯一的文件路径，但不创建文件本身
        temp_dir = tempfile.gettempdir()
        # 使用测试用例的ID确保文件名唯一，避免并发测试问题
        output_file_path = os.path.join(temp_dir, f"test_report_{self.id()}.xlsx")
        # 确保在测试开始前，该文件不存在
        if os.path.exists(output_file_path):
            os.remove(output_file_path)

        # 2. 执行：生成报表
        try:
            self.monthly_handler.generate_report(
                inventory_data=[], # 此测试不关注库存数据
                error_data=mock_error_data,
                output_file_path=output_file_path,
                device_code=device_code,
                start_date=start_date,
                end_date=end_date,
                oil_name=oil_name,
            )

            # 3. 断言：验证生成的Excel文件
            self.assertTrue(os.path.exists(output_file_path), "应成功创建Excel报表文件。")
            self.assertGreater(os.path.getsize(output_file_path), 0, "Excel文件不应为空。")

            # 加载工作簿并检查其内容
            wb = load_workbook(output_file_path)
            ws = wb.active

            # 检查标题
            self.assertIn(device_code, ws["A1"].value)
            self.assertIn(oil_name, ws["A1"].value)

            # 检查表头
            self.assertEqual(ws["A2"].value, "月份")
            self.assertEqual(ws["B2"].value, "订单累积总量(L)")

            # *** 针对修复的关键检查 ***
            self.assertEqual(ws["A6"].value, "2025-10-25", "最后一个月的标签应正确。")
            self.assertEqual(ws["B6"].value, 1300, "最后一个月的订单总量应正确。")
            self.assertEqual(ws["C6"].value, 1250, "最后一个月的消耗量应正确。")
            self.assertEqual(ws["E6"].value, 50, "最后一个月的订单误差应正确。")

            # 检查图表
            self.assertGreater(len(ws._charts), 0, "工作表中应存在一个图表。")

        finally:
            # 清理临时文件
            if os.path.exists(output_file_path):
                os.remove(output_file_path)

if __name__ == "__main__":
    unittest.main()
