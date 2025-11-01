import pytest
import os
import datetime
import shutil
from openpyxl import load_workbook

# 将src目录添加到Python路径
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.core.data_manager import ReportDataManager
from src.core.consumption_error_handler import MonthlyConsumptionErrorReportGenerator


@pytest.fixture(scope="module")
def temp_output_dir():
    """
    创建一个临时的测试输出目录，并在所有测试结束后清理。
    """
    dir_path = "temp_test_output_monthly"
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    yield dir_path
    if os.path.exists(dir_path):
        shutil.rmtree(dir_path)


@pytest.fixture
def mock_db_handler():
    """
    模拟数据库处理器。
    """
    return True # 返回一个简单的占位符


def test_monthly_report_with_barrel_count(temp_output_dir, mock_db_handler, mocker):
    """
    测试场景:
    - 模拟一个有2个油桶的设备。
    - 模拟一个月内库存正常下降（无加油）的场景。
    - 验证 data_manager 计算出的月度消耗量是否正确乘以了桶数。
    - 验证 consumption_error_handler 是否将正确的值写入了Excel。
    """
    # 1. 准备 (Arrange)
    device_code = "TEST_DEVICE_02"
    barrel_count = 2
    start_date = datetime.date(2025, 8, 1)
    end_date = datetime.date(2025, 8, 31)
    month_key = "2025-08"

    # 模拟从数据库获取的原始数据
    columns = ["加注时间", "原油剩余量", "油加注值"]
    raw_data_content = [
        (datetime.datetime(2025, 8, 1, 9, 0, 0), 800.0, 150.0),   # 月初库存
        (datetime.datetime(2025, 8, 15, 12, 0, 0), 600.0, 200.0),
        (datetime.datetime(2025, 8, 31, 18, 0, 0), 500.0, 180.0)  # 月末库存
    ]
    mock_raw_data = ([], columns, raw_data_content)

    # 预期结果
    # 单桶液位消耗 = 月初库存 - 月末库存 = 800 - 500 = 300
    # 真实消耗总量 = 300 * 2 (桶数) = 600
    expected_consumption = 600.0

    # 2. 执行 (Act)
    # 步骤 A: 使用 data_manager 计算误差数据
    data_manager = ReportDataManager(mock_db_handler)
    mocker.patch.object(data_manager, 'fetch_raw_data', return_value=mock_raw_data)
    error_data = data_manager.calculate_monthly_errors(mock_raw_data, start_date, end_date, barrel_count)

    # 步骤 B: 使用 consumption_error_handler 生成报表
    report_generator = MonthlyConsumptionErrorReportGenerator()
    output_file_path = os.path.join(temp_output_dir, "test_monthly_report_with_barrels.xlsx")
    report_generator.generate_report(
        inventory_data=[], # 简化
        error_data=error_data,
        output_file_path=output_file_path,
        device_code=device_code,
        start_date=start_date,
        end_date=end_date,
        oil_name="Test Oil",
        barrel_count=barrel_count
    )

    # 3. 断言 (Assert)
    # 断言 A: 检查 data_manager 的计算结果
    calculated_consumption = error_data['monthly_consumption'].get(month_key, {}).get('value')
    assert calculated_consumption == pytest.approx(expected_consumption), \
        f"data_manager 计算的月度消耗总量应为 {expected_consumption}，但实际为 {calculated_consumption}"

    # 断言 B: 检查生成的Excel文件内容
    assert os.path.exists(output_file_path), "报表文件未生成"
    wb = load_workbook(output_file_path)
    ws = wb["消耗误差分析"]

    # 数据从第3行开始，"库存消耗总量(L)"是第3列(C)
    consumption_cell_value = ws.cell(row=3, column=3).value
    assert consumption_cell_value == pytest.approx(expected_consumption), \
        f"Excel中的月度消耗总量应为 {expected_consumption}，但实际为 {consumption_cell_value}"

    # 关键修复：关闭工作簿以释放文件句柄
    wb.close()