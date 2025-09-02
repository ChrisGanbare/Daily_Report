import os
import sys
import tempfile
import unittest
from datetime import date
from unittest.mock import Mock, patch

# 添加项目根目录到sys.path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from openpyxl import load_workbook

from src.core.refueling_details_handler import RefuelingDetailsReportGenerator
from tests.base_test import BaseTestCase


class TestCoreRefuelingDetailsHandler(BaseTestCase):
    """
    core.refueling_details_handler 模块的单元测试
    """

    def setUp(self):
        """测试前准备"""
        super().setUp()
        self.refueling_handler = RefuelingDetailsReportGenerator()
        self.test_data = [
            (1, '2025-07-01 08:00:00', 1, '切削液', 1.0, 1.0, 10.0, 20.0, 800, 0.8, 100, 3, 1),
            (2, '2025-07-02 09:00:00', 1, '切削液', 1.0, 1.0, 10.0, 25.0, 775, 0.775, 100, 3, 2)
        ]
        self.columns = [
            '订单序号', '加注时间', '油品序号', '油品名称', '水油比：水值', '水油比：油值',
            '水加注值', '油加注值', '原油剩余量', '原油剩余比例', '油加设量', 
            '是否结算：1=待结算 2=待生效 3=已结算', '加注模式：1=近程自动 2=远程自动 3=手动'
        ]
        self.device_code = "TEST001"
        self.customer_name = "测试客户"
        self.start_date = date(2025, 7, 1)
        self.end_date = date(2025, 7, 31)

    def test_refuelingdetailsreportgenerator_initialization(self):
        """
        测试 RefuelingDetailsReportGenerator 类的初始化
        """
        self.assertIsInstance(self.refueling_handler, RefuelingDetailsReportGenerator)

    def test_generate_refueling_details_report_success(self):
        """
        测试成功生成加注明细报表
        """
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file_path = tmp_file.name

        try:
            result = self.refueling_handler.generate_refueling_details_report(
                refueling_data=self.test_data,
                output_file_path=tmp_file_path,
                device_code=self.device_code,
                start_date=self.start_date,
                end_date=self.end_date,
                customer_name=self.customer_name,
                columns=self.columns
            )
            
            # 验证返回值
            self.assertTrue(result)
            
            # 验证文件是否创建
            self.assertTrue(os.path.exists(tmp_file_path))
            
            # 验证Excel文件内容
            wb = load_workbook(tmp_file_path)
            ws = wb.active
            
            # 验证列标题（现在从第一行开始）
            for i, column_name in enumerate(self.columns, 1):
                self.assertEqual(ws.cell(row=1, column=i).value, column_name)
            
            # 验证数据行
            for row_idx, test_row in enumerate(self.test_data, 2):  # 从第2行开始是数据
                for col_idx, cell_value in enumerate(test_row, 1):
                    self.assertEqual(ws.cell(row=row_idx, column=col_idx).value, cell_value)
            
            wb.close()
            
            # 强制进行垃圾回收以确保文件句柄被释放
            import gc
            gc.collect()
        finally:
            # 清理临时文件
            if os.path.exists(tmp_file_path):
                # 使用延迟和重试机制确保文件可以被删除
                import time
                for i in range(3):  # 最多重试3次
                    try:
                        os.unlink(tmp_file_path)
                        break
                    except PermissionError:
                        if i < 2:  # 不是最后一次尝试
                            time.sleep(0.1 * (i + 1))  # 等待时间逐渐增加
                        else:
                            raise  # 最后一次尝试仍然失败，则抛出异常

    def test_generate_report_method(self):
        """
        测试generate_report方法
        """
        with patch.object(self.refueling_handler, 'generate_refueling_details_report') as mock_generate:
            mock_generate.return_value = True
            
            result = self.refueling_handler.generate_report(
                refueling_data=self.test_data,
                output_file_path="test.xlsx",
                device_code=self.device_code,
                start_date=self.start_date,
                end_date=self.end_date,
                customer_name=self.customer_name,
                columns=self.columns
            )
            
            # 验证generate_refueling_details_report被正确调用
            mock_generate.assert_called_once_with(
                self.test_data, "test.xlsx", self.device_code, self.start_date,
                self.end_date, self.customer_name, self.columns
            )
            self.assertTrue(result)

    def test_generate_refueling_details_report_with_default_columns(self):
        """
        测试生成加注明细报表时使用默认列标题
        """
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_file_path = tmp_file.name

        try:
            result = self.refueling_handler.generate_refueling_details_report(
                refueling_data=self.test_data,
                output_file_path=tmp_file_path,
                device_code=self.device_code,
                start_date=self.start_date,
                end_date=self.end_date,
                customer_name=self.customer_name
                # 不传入columns参数，使用默认值
            )
            
            # 验证返回值
            self.assertTrue(result)
            
            # 验证文件是否创建
            self.assertTrue(os.path.exists(tmp_file_path))
            
            # 验证Excel文件内容
            wb = load_workbook(tmp_file_path)
            ws = wb.active
            
            # 验证默认列标题（现在从第一行开始）
            default_columns = [
                '订单序号', '加注时间', '油品序号', '油品名称', '水油比：水值', '水油比：油值',
                '水加注值', '油加注值', '原油剩余量', '原油剩余比例', '油加设量', 
                '是否结算：1=待结算 2=待生效 3=已结算', '加注模式：1=近程自动 2=远程自动 3=手动'
            ]
            for i, column_name in enumerate(default_columns, 1):
                self.assertEqual(ws.cell(row=1, column=i).value, column_name)
            
            wb.close()
            
            # 强制进行垃圾回收以确保文件句柄被释放
            import gc
            gc.collect()
        finally:
            # 清理临时文件
            if os.path.exists(tmp_file_path):
                # 使用延迟和重试机制确保文件可以被删除
                import time
                for i in range(3):  # 最多重试3次
                    try:
                        os.unlink(tmp_file_path)
                        break
                    except PermissionError:
                        if i < 2:  # 不是最后一次尝试
                            time.sleep(0.1 * (i + 1))  # 等待时间逐渐增加
                        else:
                            raise  # 最后一次尝试仍然失败，则抛出异常


if __name__ == '__main__':
    unittest.main()
