import os
import sys
import tempfile
import time
import unittest
from datetime import date, timedelta
from unittest.mock import Mock, patch

# 添加项目根目录到sys.path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from openpyxl import load_workbook

from src.core.inventory_handler import InventoryReportGenerator
from tests.base_test import BaseTestCase


class TestCoreInventoryHandler(BaseTestCase):
    """
    core.inventory_handler 模块的单元测试
    """

    def setUp(self):
        """测试前准备"""
        super().setUp()
        self.inventory_handler = InventoryReportGenerator()
        self.test_data = [
            (date(2025, 7, 1), 100.0),
            (date(2025, 7, 2), 95.0),
            (date(2025, 7, 3), 90.0),
        ]

    def test_inventoryreportgenerator_initialization(self):
        """
        测试 InventoryReportGenerator 类的初始化
        """
        # InventoryReportGenerator的__init__方法目前为空，只需确认能创建实例
        self.assertIsInstance(self.inventory_handler, InventoryReportGenerator)

    def test_inventoryreportgenerator_validate_inventory_value_valid(self):
        """
        测试 InventoryReportGenerator._validate_inventory_value 方法验证有效值
        """
        # 测试正常值
        self.assertEqual(self.inventory_handler._validate_inventory_value(100), 100.0)
        self.assertEqual(self.inventory_handler._validate_inventory_value(0), 0.0)
        self.assertEqual(self.inventory_handler._validate_inventory_value("50.5"), 50.5)
        # 测试大于100的值（应该允许）
        self.assertEqual(self.inventory_handler._validate_inventory_value(150), 150.0)

    def test_inventoryreportgenerator_validate_inventory_value_invalid(self):
        """
        测试 InventoryReportGenerator._validate_inventory_value 方法处理无效值
        """
        # 测试负数（应该抛出异常）
        with self.assertRaises(ValueError):
            self.inventory_handler._validate_inventory_value(-10)

        # 测试无效值（应该抛出异常）
        with self.assertRaises(ValueError):
            self.inventory_handler._validate_inventory_value("abc")

        with self.assertRaises(ValueError):
            self.inventory_handler._validate_inventory_value(None)

    def _remove_file_with_retry(self, file_path, max_retries=5, delay=0.1):
        """
        带重试机制的文件删除方法

        Args:
            file_path (str): 要删除的文件路径
            max_retries (int): 最大重试次数
            delay (float): 重试间隔（秒）
        """
        for i in range(max_retries):
            try:
                if os.path.exists(file_path):
                    os.unlink(file_path)
                return True
            except PermissionError:
                if i < max_retries - 1:  # 不是最后一次尝试
                    time.sleep(delay)
                else:
                    # 最后一次尝试仍然失败
                    print(f"警告: 无法删除文件 {file_path}，文件可能仍在被使用")
                    return False
        return False

    def test_inventoryreportgenerator_generate_inventory_report_with_chart_xlsx(self):
        """
        测试 InventoryReportGenerator.generate_inventory_report_with_chart 方法生成XLSX文件
        """
        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        output_file_path = f.name
        f.close()

        wb = None
        try:
            self.inventory_handler.generate_inventory_report_with_chart(
                inventory_data=self.test_data,
                output_file_path=output_file_path,
                device_code="TEST001",
                start_date=date(2025, 7, 1),
                end_date=date(2025, 7, 3),
                oil_name="切削液",
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 验证文件已创建
            self.assertTrue(os.path.exists(output_file_path))

            # 验证文件不为空
            self.assertGreater(os.path.getsize(output_file_path), 0)

            # 验证Excel文件内容
            try:
                wb = load_workbook(output_file_path)
                ws = wb.active

                # 验证标题
                self.assertIn("TEST001", ws["A1"].value)
                self.assertIn("切削液", ws["A1"].value)

                # 验证数据 (注意openpyxl读取时会将date转换为datetime)
                self.assertEqual(ws["A3"].value.date(), date(2025, 7, 1))
                self.assertEqual(ws["B3"].value, 100.0)
                self.assertEqual(ws["A4"].value.date(), date(2025, 7, 2))
                self.assertEqual(ws["B4"].value, 95.0)
                self.assertEqual(ws["A5"].value.date(), date(2025, 7, 3))
                self.assertEqual(ws["B5"].value, 90.0)

                # 验证图表存在
                self.assertGreaterEqual(len(ws._charts), 1)
            finally:
                if wb is not None:
                    wb.close()
        finally:
            # 使用重试机制删除文件
            self._remove_file_with_retry(output_file_path)

    def test_inventoryreportgenerator_generate_inventory_report_with_chart_csv(self):
        """
        测试 InventoryReportGenerator.generate_inventory_report_with_chart 方法生成CSV文件
        """
        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        output_file_path = f.name
        f.close()

        try:
            self.inventory_handler.generate_inventory_report_with_chart(
                inventory_data=self.test_data,
                output_file_path=output_file_path,
                device_code="TEST001",
                start_date=date(2025, 7, 1),
                end_date=date(2025, 7, 3),
                oil_name="切削液",
                export_format="csv",
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # CSV文件应该保存为.csv扩展名
            csv_file_path = output_file_path.replace(".xlsx", ".csv")

            # 验证文件已创建
            self.assertTrue(os.path.exists(csv_file_path))

            # 验证文件不为空
            self.assertGreater(os.path.getsize(csv_file_path), 0)

            # 清理CSV文件
            if os.path.exists(csv_file_path):
                self._remove_file_with_retry(csv_file_path)
        finally:
            # 确保原始xlsx文件也被清理（如果存在）
            self._remove_file_with_retry(output_file_path)

    def test_inventoryreportgenerator_generate_inventory_report_with_invalid_data(self):
        """
        测试 InventoryReportGenerator.generate_inventory_report_with_chart 方法处理无效数据
        """
        # 包含无效数据的测试数据
        test_data = [
            (date(2025, 7, 1), 105.0),  # 有效数据
            (date(2025, 7, 2), "invalid"),  # 无效数据
            (date(2025, 7, 3), -5.0),  # 无效数据
        ]

        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        xlsx_file_path = f.name
        f.close()

        wb = None
        try:
            self.inventory_handler.generate_inventory_report_with_chart(
                test_data, xlsx_file_path, "DEV001", date(2025, 7, 1), date(2025, 7, 3)
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 验证文件已创建
            self.assertTrue(os.path.exists(xlsx_file_path))

            # 验证Excel文件内容
            try:
                wb = load_workbook(xlsx_file_path)
                ws = wb.active

                # 验证标题
                self.assertIn("DEV001", ws["A1"].value)

                # 验证有效数据存在 (注意openpyxl读取时会将date转换为datetime)
                self.assertEqual(ws["A3"].value.date(), date(2025, 7, 1))
                self.assertEqual(ws["B3"].value, 105.0)
            finally:
                if wb is not None:
                    wb.close()
        finally:
            self._remove_file_with_retry(xlsx_file_path)

    def test_inventoryreportgenerator_generate_inventory_report_with_empty_data(self):
        """
        测试 InventoryReportGenerator.generate_inventory_report_with_chart 方法处理空数据
        """
        test_data = []

        f = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        xlsx_file_path = f.name
        f.close()

        wb = None
        try:
            self.inventory_handler.generate_inventory_report_with_chart(
                test_data, xlsx_file_path, "DEV001", date(2025, 7, 1), date(2025, 7, 2)
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 验证文件已创建
            self.assertTrue(os.path.exists(xlsx_file_path))

            # 验证Excel文件内容
            try:
                wb = load_workbook(xlsx_file_path)
                ws = wb.active

                # 验证标题
                self.assertIn("DEV001", ws["A1"].value)

                # 验证默认数据存在 (注意openpyxl读取时会将date转换为datetime)
                self.assertEqual(ws["A3"].value.date(), date(2025, 7, 1))
                self.assertEqual(ws["A4"].value.date(), date(2025, 7, 2))
            finally:
                if wb is not None:
                    wb.close()
        finally:
            self._remove_file_with_retry(xlsx_file_path)

    def test_inventoryreportgenerator_generate_inventory_report_with_chart_csv_format(
        self,
    ):
        """
        测试 InventoryReportGenerator.generate_inventory_report_with_chart 方法生成CSV格式报告
        """
        test_data = [(date(2025, 7, 1), 100.0), (date(2025, 7, 2), 95.0)]

        f = tempfile.NamedTemporaryFile(suffix=".csv", delete=False)
        csv_file_path = f.name
        f.close()

        try:
            self.inventory_handler.generate_inventory_report_with_chart(
                test_data,
                csv_file_path,
                "DEV001",
                date(2025, 7, 1),
                date(2025, 7, 2),
                export_format="csv",
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 验证文件已创建
            self.assertTrue(os.path.exists(csv_file_path))

            # 验证CSV内容
            with open(csv_file_path, "r") as f:
                content = f.read()
                self.assertIn("日期", content)
                self.assertIn("原油剩余量(L)", content)  # 修正列标题检查
        finally:
            self._remove_file_with_retry(csv_file_path)


if __name__ == "__main__":
    unittest.main()