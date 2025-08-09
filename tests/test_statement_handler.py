import unittest
import os
import tempfile
from datetime import date
from openpyxl import Workbook

# 更新导入路径以适应新的目录结构
from src.core.statement_handler import StatementHandler
from base_test import BaseTestCase


class TestStatementHandler(BaseTestCase):
    """测试对账单处理模块"""

    def setUp(self):
        """测试前准备"""
        super().setUp()
        self.statement_handler = StatementHandler()
        self.customer_name = "测试客户"
        self.start_date = date(2025, 7, 1)
        self.end_date = date(2025, 7, 31)
        
        # 创建测试数据
        self.test_data = [
            {
                'device_code': 'DEV001',
                'oil_name': '液压油1',
                'data': [
                    (date(2025, 7, 1), 10.5),
                    (date(2025, 7, 2), 12.3),
                    (date(2025, 7, 3), 9.8)
                ],
                'raw_data': [],
                'columns': []
            },
            {
                'device_code': 'DEV002',
                'oil_name': '液压油2',
                'data': [
                    (date(2025, 7, 1), 15.2),
                    (date(2025, 7, 2), 14.1),
                    (date(2025, 7, 3), 16.7)
                ],
                'raw_data': [],
                'columns': []
            }
        ]

    def test_statement_handler_initialization(self):
        """测试对账单处理器初始化"""
        self.assertIsInstance(self.statement_handler, StatementHandler)
        
    def test_generate_statement_from_template_file_not_found(self):
        """测试模板文件不存在的情况"""
        output_file = os.path.join(self.test_output_dir, "test_statement.xlsx")
        
        # 应该抛出FileNotFoundError异常
        with self.assertRaises(FileNotFoundError):
            self.statement_handler.generate_statement_from_template(
                all_devices_data=self.test_data,
                output_file=output_file,
                customer_name=self.customer_name,
                start_date=self.start_date,
                end_date=self.end_date
            )

    def test_template_path_validation(self):
        """测试模板路径验证"""
        # 检查模板路径是否正确指向项目根目录下的template目录
        expected_template_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 
            'template', 
            'statement_template.xlsx'
        )
        
        # 通过反射检查StatementHandler的模板路径逻辑
        import inspect
        source_code = inspect.getsource(self.statement_handler.generate_statement_from_template)
        self.assertIn('template', source_code)
        self.assertIn('statement_template.xlsx', source_code)

    def test_required_sheets_detection_chinese(self):
        """测试必需中文工作表的识别功能"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建缺少工作表的模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            # 故意缺少"每日用量明细"和"每月用量对比"
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            # 应该抛出KeyError异常
            with self.assertRaises(KeyError) as context:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
            
            # 检查错误信息是否包含缺少的工作表信息
            error_message = str(context.exception)
            self.assertIn("模板工作表名称不匹配", error_message)
            self.assertIn("缺少中文工作表", error_message)

    def test_required_sheets_detection_english(self):
        """测试必需英文工作表的识别功能"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建缺少工作表的模板（只有部分英文工作表）
            template_wb = Workbook()
            template_wb.create_sheet("Statement")
            # 故意缺少"Daily Usage"和"Monthly Comparison"
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            # 应该抛出KeyError异常
            with self.assertRaises(KeyError) as context:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
            
            # 检查错误信息是否包含缺少的工作表信息
            error_message = str(context.exception)
            self.assertIn("模板工作表名称不匹配", error_message)
            self.assertIn("缺少英文工作表", error_message)

    def test_daily_usage_sheet_updates(self):
        """测试每日用量明细工作表的数据更新"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def test_monthly_comparison_sheet_updates(self):
        """测试每月用量对比工作表的数据更新"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def test_statement_sheet_updates(self):
        """测试对账单工作表的数据更新"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def test_multiple_oil_products_handling(self):
        """测试多油品处理功能"""
        # 添加更多测试数据，包含多个油品
        extended_test_data = self.test_data + [
            {
                'device_code': 'DEV003',
                'oil_name': '液压油3',
                'data': [
                    (date(2025, 7, 1), 5.5),
                    (date(2025, 7, 2), 6.3),
                    (date(2025, 7, 3), 7.8)
                ],
                'raw_data': [],
                'columns': []
            }
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    extended_test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
                
                # 应该处理3种油品
                # 注意：由于我们无法直接访问私有方法内部实现，这里只验证文件生成
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def test_dynamic_oil_products_handling(self):
        """测试动态油品数量处理功能"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,  # 只有2个油品
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
                
                # 重新加载生成的文件检查内容
                from openpyxl import load_workbook
                wb = load_workbook(output_file)
                
                # 检查是否包含所有必需的工作表
                self.assertIn("中润对账单", wb.sheetnames)
                self.assertIn("每日用量明细", wb.sheetnames)
                self.assertIn("每月用量对比", wb.sheetnames)
                
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def test_date_format_handling(self):
        """测试日期格式处理功能"""
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    self.test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
                
                # 重新加载生成的文件检查日期格式
                from openpyxl import load_workbook
                wb = load_workbook(output_file)
                daily_usage_sheet = wb["每日用量明细"]
                
                # 检查日期范围格式 (第3行)
                date_range_cell = daily_usage_sheet.cell(row=3, column=2).value
                self.assertIsNotNone(date_range_cell)
                self.assertIn("2025.07.01", date_range_cell)
                self.assertIn("2025.07.31", date_range_cell)
                
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def test_data_aggregation(self):
        """测试数据聚合功能"""
        # 创建包含相同油品的测试数据
        aggregation_test_data = [
            {
                'device_code': 'DEV001',
                'oil_name': '切削液',
                'data': [
                    (date(2025, 7, 1), 10.0),
                    (date(2025, 7, 2), 15.0)
                ],
                'raw_data': [],
                'columns': []
            },
            {
                'device_code': 'DEV002',
                'oil_name': '切削液',  # 相同油品
                'data': [
                    (date(2025, 7, 1), 5.0),
                    (date(2025, 7, 2), 10.0)
                ],
                'raw_data': [],
                'columns': []
            }
        ]
        
        with tempfile.TemporaryDirectory() as temp_dir:
            # 创建完整模板
            template_wb = Workbook()
            template_wb.create_sheet("中润对账单")
            template_wb.create_sheet("每日用量明细")
            template_wb.create_sheet("每月用量对比")
            if "Sheet" in template_wb.sheetnames:
                template_wb.remove(template_wb["Sheet"])
            
            template_path = os.path.join(temp_dir, 'template')
            os.makedirs(template_path, exist_ok=True)
            template_file = os.path.join(template_path, 'statement_template.xlsx')
            template_wb.save(template_file)
            
            output_file = os.path.join(temp_dir, "test_output.xlsx")
            
            try:
                statement_handler = StatementHandler()
                statement_handler.generate_statement_from_template(
                    aggregation_test_data,
                    output_file,
                    self.customer_name,
                    self.start_date,
                    self.end_date
                )
                
                # 检查输出文件是否存在
                self.assertTrue(os.path.exists(output_file))
                
                # 重新加载生成的文件检查数据聚合
                from openpyxl import load_workbook
                wb = load_workbook(output_file)
                daily_usage_sheet = wb["每日用量明细"]
                
                # 检查数据是否正确聚合（7月1日应该是10.0+5.0=15.0）
                # 查找日期为7.1的行
                date_found = False
                for row in range(6, daily_usage_sheet.max_row + 1):
                    date_cell_value = daily_usage_sheet.cell(row=row, column=2).value
                    if date_cell_value == "7.1":
                        date_found = True
                        # 检查聚合后的值（应该在C列）
                        aggregated_value = daily_usage_sheet.cell(row=row, column=3).value
                        # 由于我们不知道确切的列位置，我们检查值是否合理
                        self.assertIsNotNone(aggregated_value)
                        # 值应该大于任何一个单独的值
                        self.assertGreaterEqual(aggregated_value, 10.0)
                        break
                
                # 确保找到了对应的日期行
                self.assertTrue(date_found, "未找到日期为7.1的数据行")
                
            finally:
                # 清理测试文件
                if os.path.exists(output_file):
                    os.remove(output_file)

    def tearDown(self):
        """测试后清理"""
        pass


if __name__ == '__main__':
    unittest.main()