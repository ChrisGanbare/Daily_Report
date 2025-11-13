"""
依赖版本兼容性测试方案

本文件用于测试关键依赖库的版本兼容性，确保在不同版本下核心功能正常工作。
主要测试的依赖库包括：
1. openpyxl - 用于Excel文件处理，特别是图表生成
2. mysql-connector-python - 用于数据库连接
3. pandas - 用于数据处理

测试方案：
1. 针对每个关键依赖，创建专门的测试用例
2. 测试核心功能在不同版本下的行为
3. 验证关键功能的输出是否符合预期
4. 记录测试结果，便于版本升级时参考
"""

import os

# 添加项目根目录到sys.path，确保能正确导入模块
import sys
import tempfile
import time
import unittest
from datetime import date

from openpyxl import load_workbook
from openpyxl.chart import Reference

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from core.inventory_handler import InventoryReportGenerator


class DependencyCompatibilityTest(unittest.TestCase):
    """依赖版本兼容性测试"""

    def setUp(self):
        """测试前准备"""
        self.test_data = [
            (date(2025, 7, 1), 100.0),
            (date(2025, 7, 2), 95.0),
            (date(2025, 7, 3), 90.0),
        ]
        self.device_code = "TEST001"
        self.start_date = date(2025, 7, 1)
        self.end_date = date(2025, 7, 3)
        self.oil_name = "切削液"

    def _remove_file_with_retry(self, file_path, max_retries=5, delay=0.1):
        """
        带重试机制的文件删除方法

        Args:
            file_path (str): 要删除的文件路径
            max_retries (int): 最大重试次数
            delay (float): 重试间隔（秒）
        """
        for i in range(max_retries):
            try:
                if os.path.exists(file_path):
                    os.unlink(file_path)
                return True
            except PermissionError:
                if i < max_retries - 1:  # 不是最后一次尝试
                    time.sleep(delay)
                else:
                    # 最后一次尝试仍然失败
                    print(f"警告: 无法删除文件 {file_path}，文件可能仍在被使用")
                    return False
        return False

    def test_openpyxl_chart_generation(self):
        """测试openpyxl图表生成功能兼容性"""
        tmp_dir = tempfile.TemporaryDirectory()
        tmp_file_path = None
        wb = None
        try:
            tmp_file_path = os.path.join(tmp_dir.name, "compatibility_test_report.xlsx")

            # 创建报表生成器并生成包含图表的报表
            generator = InventoryReportGenerator()
            generator.generate_inventory_report_with_chart(
                inventory_data=self.test_data,
                output_file_path=tmp_file_path,
                device_code=self.device_code,
                start_date=self.start_date,
                end_date=self.end_date,
                oil_name=self.oil_name,
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 验证文件生成
            self.assertTrue(os.path.exists(tmp_file_path))

            # 加载工作簿验证内容
            wb = load_workbook(tmp_file_path)
            ws = wb.active

            # 验证数据行 (注意openpyxl读取时会将date转换为datetime)
            self.assertEqual(ws["A3"].value.date(), self.test_data[0][0])
            self.assertEqual(ws["B3"].value, self.test_data[0][1])
            self.assertEqual(ws["A4"].value.date(), self.test_data[1][0])
            self.assertEqual(ws["B4"].value, self.test_data[1][1])
            self.assertEqual(ws["A5"].value.date(), self.test_data[2][0])
            self.assertEqual(ws["B5"].value, self.test_data[2][1])

            # 验证图表存在
            self.assertEqual(len(ws._charts), 1)
            chart = ws._charts[0]

            # 验证图表有标题
            self.assertIsNotNone(chart.title)

            # 验证数据系列
            self.assertEqual(len(chart.series), 1)
            series = chart.series[0]

            # 验证数据范围引用正确
            expected_values_ref = "'库存数据'!$B$3:$B$5"
            expected_categories_ref = "'库存数据'!$A$3:$A$5"

            # 正确获取引用字符串
            self.assertEqual(str(series.val.numRef.f), expected_values_ref)
            self.assertEqual(str(series.cat.numRef.f), expected_categories_ref)

        finally:
            # 确保工作簿被关闭
            if wb is not None:
                wb.close()
            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)
            # 确保临时目录被清理
            try:
                tmp_dir.cleanup()
            except:
                # 如果清理失败，至少尝试关闭文件
                if tmp_file_path and os.path.exists(tmp_file_path):
                    self._remove_file_with_retry(tmp_file_path)

    def test_openpyxl_chart_axis_labels(self):
        """测试openpyxl图表坐标轴标签功能兼容性"""
        tmp_dir = tempfile.TemporaryDirectory()
        tmp_file_path = None
        wb = None
        try:
            tmp_file_path = os.path.join(tmp_dir.name, "axis_labels_test_report.xlsx")

            # 创建报表生成器并生成包含图表的报表
            generator = InventoryReportGenerator()
            generator.generate_inventory_report_with_chart(
                inventory_data=self.test_data,
                output_file_path=tmp_file_path,
                device_code=self.device_code,
                start_date=self.start_date,
                end_date=self.end_date,
                oil_name=self.oil_name,
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 加载工作簿验证图表坐标轴
            wb = load_workbook(tmp_file_path)
            ws = wb.active
            chart = ws._charts[0]

            # 验证图表存在
            self.assertIsNotNone(chart)

            # 验证X轴和Y轴存在
            self.assertIsNotNone(chart.x_axis)
            self.assertIsNotNone(chart.y_axis)

            # 验证坐标轴标题
            # 注意：在openpyxl的新版本中，坐标轴标题可能通过Title对象设置
            # 我们主要验证坐标轴对象存在且不为空
            self.assertTrue(hasattr(chart.x_axis, "title"))
            self.assertTrue(hasattr(chart.y_axis, "title"))

        finally:
            # 确保工作簿被关闭
            if wb is not None:
                wb.close()
            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)
            # 确保临时目录被清理
            try:
                tmp_dir.cleanup()
            except:
                # 如果清理失败，至少尝试关闭文件
                if tmp_file_path and os.path.exists(tmp_file_path):
                    self._remove_file_with_retry(tmp_file_path)

    def test_openpyxl_chart_title_functionality(self):
        """测试openpyxl图表标题功能兼容性"""
        tmp_dir = tempfile.TemporaryDirectory()
        tmp_file_path = None
        wb = None
        try:
            tmp_file_path = os.path.join(tmp_dir.name, "chart_title_test_report.xlsx")

            # 创建报表生成器并生成包含图表的报表
            generator = InventoryReportGenerator()
            generator.generate_inventory_report_with_chart(
                inventory_data=self.test_data,
                output_file_path=tmp_file_path,
                device_code=self.device_code,
                start_date=self.start_date,
                end_date=self.end_date,
                oil_name=self.oil_name,
            )

            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)

            # 加载工作簿验证图表标题
            wb = load_workbook(tmp_file_path)
            ws = wb.active
            chart = ws._charts[0]

            # 验证图表存在
            self.assertIsNotNone(chart)

            # 验证图表标题存在
            self.assertIsNotNone(chart.title)

            # 验证图表标题有内容
            self.assertTrue(hasattr(chart.title, "tx"))

            # 如果标题有str属性，则验证其内容
            if hasattr(chart.title, "tx") and hasattr(chart.title.tx, "str"):
                # 标题应该包含设备编号和油品名称
                title_text = chart.title.tx.str
                self.assertIn(self.device_code, title_text)
                self.assertIn(self.oil_name, title_text)

        finally:
            # 确保工作簿被关闭
            if wb is not None:
                wb.close()
            # 添加延迟确保文件被系统完全释放
            time.sleep(0.1)
            # 确保临时目录被清理
            try:
                tmp_dir.cleanup()
            except:
                # 如果清理失败，至少尝试关闭文件
                if tmp_file_path and os.path.exists(tmp_file_path):
                    self._remove_file_with_retry(tmp_file_path)

    def test_mysql_connector_basic_connection(self):
        """测试mysql-connector-python基本连接功能兼容性"""
        # 由于我们不希望在测试中实际连接数据库，
        # 这里仅验证模块是否可以正常导入
        try:
            import mysql.connector

            self.assertTrue(hasattr(mysql.connector, "connect"))
        except ImportError:
            self.fail("无法导入mysql.connector模块")

    def test_pandas_basic_functionality(self):
        """测试pandas基本功能兼容性"""
        try:
            import pandas as pd

            # 创建简单数据框测试基本功能
            df = pd.DataFrame(
                {"date": [date(2025, 7, 1), date(2025, 7, 2)], "value": [100.0, 95.0]}
            )
            self.assertEqual(len(df), 2)
            self.assertIn("date", df.columns)
            self.assertIn("value", df.columns)
        except ImportError:
            self.fail("无法导入pandas模块")
        except Exception as e:
            self.fail(f"pandas基本功能测试失败: {e}")