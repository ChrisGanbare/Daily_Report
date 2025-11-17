"""检查模板中的实际颜色值"""
from openpyxl import load_workbook

wb = load_workbook('template/statement_template.xlsx')

# 检查中润对账单数据行填充
ws1 = wb['中润对账单']
cell = ws1.cell(9, 1)
if cell.fill:
    print(f"中润对账单数据行填充:")
    print(f"  start_color: {cell.fill.start_color}")
    if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
        print(f"  RGB: {cell.fill.start_color.rgb}")
    if hasattr(cell.fill.start_color, 'theme'):
        print(f"  Theme: {cell.fill.start_color.theme}")
    if hasattr(cell.fill.start_color, 'tint'):
        print(f"  Tint: {cell.fill.start_color.tint}")

# 检查每月用量对比表头填充
ws3 = wb['每月用量对比']
cell = ws3.cell(6, 1)
if cell.fill:
    print(f"\n每月用量对比表头填充:")
    print(f"  start_color: {cell.fill.start_color}")
    if hasattr(cell.fill.start_color, 'rgb') and cell.fill.start_color.rgb:
        print(f"  RGB: {cell.fill.start_color.rgb}")
    if hasattr(cell.fill.start_color, 'theme'):
        print(f"  Theme: {cell.fill.start_color.theme}")
    if hasattr(cell.fill.start_color, 'tint'):
        print(f"  Tint: {cell.fill.start_color.tint}")

wb.close()

