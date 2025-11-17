"""
分析客户对账单Excel模板的详细结构
包括文本内容、布局和样式
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json

def analyze_cell(cell):
    """分析单元格的详细信息"""
    info = {
        'value': cell.value,
        'data_type': type(cell.value).__name__ if cell.value is not None else None,
    }
    
    # 字体信息
    if cell.font:
        font_info = {}
        if hasattr(cell.font, 'name') and cell.font.name:
            font_info['name'] = cell.font.name
        if hasattr(cell.font, 'size') and cell.font.size:
            font_info['size'] = cell.font.size
        if hasattr(cell.font, 'bold') and cell.font.bold:
            font_info['bold'] = cell.font.bold
        if hasattr(cell.font, 'italic') and cell.font.italic:
            font_info['italic'] = cell.font.italic
        if hasattr(cell.font, 'color') and cell.font.color:
            font_info['color'] = str(cell.font.color)
        if font_info:
            info['font'] = font_info
    
    # 对齐信息
    if cell.alignment:
        alignment_info = {}
        if hasattr(cell.alignment, 'horizontal') and cell.alignment.horizontal:
            alignment_info['horizontal'] = cell.alignment.horizontal
        if hasattr(cell.alignment, 'vertical') and cell.alignment.vertical:
            alignment_info['vertical'] = cell.alignment.vertical
        if hasattr(cell.alignment, 'wrap_text') and cell.alignment.wrap_text:
            alignment_info['wrap_text'] = cell.alignment.wrap_text
        if alignment_info:
            info['alignment'] = alignment_info
    
    # 填充信息
    if cell.fill and hasattr(cell.fill, 'fill_type') and cell.fill.fill_type:
        fill_info = {}
        if hasattr(cell.fill, 'start_color') and cell.fill.start_color:
            fill_info['start_color'] = str(cell.fill.start_color)
        if hasattr(cell.fill, 'end_color') and cell.fill.end_color:
            fill_info['end_color'] = str(cell.fill.end_color)
        if fill_info:
            info['fill'] = fill_info
    
    # 边框信息
    if cell.border:
        border_info = {}
        for side_name in ['left', 'right', 'top', 'bottom']:
            side = getattr(cell.border, side_name, None)
            if side and hasattr(side, 'style') and side.style:
                border_info[side_name] = side.style
        if border_info:
            info['border'] = border_info
    
    # 数字格式
    if cell.number_format and cell.number_format != 'General':
        info['number_format'] = cell.number_format
    
    return info

def analyze_sheet(ws, sheet_name):
    """分析工作表的详细结构"""
    print(f"\n{'='*80}")
    print(f"工作表: {sheet_name}")
    print(f"{'='*80}")
    print(f"最大行数: {ws.max_row}, 最大列数: {ws.max_column}")
    
    # 分析合并单元格
    merged_cells = list(ws.merged_cells.ranges)
    if merged_cells:
        print(f"\n合并单元格 ({len(merged_cells)} 个):")
        for merged_range in merged_cells[:10]:  # 只显示前10个
            print(f"  {merged_range}")
    
    # 分析列宽
    print(f"\n列宽:")
    for col in range(1, min(ws.max_column + 1, 20)):
        col_letter = get_column_letter(col)
        if col_letter in ws.column_dimensions:
            width = ws.column_dimensions[col_letter].width
            if width and width != 8.43:  # 默认宽度
                print(f"  {col_letter}: {width}")
    
    # 分析行高
    print(f"\n行高:")
    for row in range(1, min(ws.max_row + 1, 20)):
        if row in ws.row_dimensions:
            height = ws.row_dimensions[row].height
            if height and height != 15:  # 默认高度
                print(f"  {row}: {height}")
    
    # 分析关键行的内容
    print(f"\n关键行内容分析:")
    key_rows = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    for row in key_rows:
        if row > ws.max_row:
            continue
        print(f"\n  行 {row}:")
        row_data = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row, col)
            if cell.value is not None:
                cell_info = analyze_cell(cell)
                col_letter = get_column_letter(col)
                row_data.append(f"{col_letter}{row}: {cell_info}")
        if row_data:
            for item in row_data[:5]:  # 只显示前5个非空单元格
                print(f"    {item}")
            if len(row_data) > 5:
                print(f"    ... (还有 {len(row_data) - 5} 个单元格)")

def main():
    template_path = "template/statement_template.xlsx"
    print(f"分析模板文件: {template_path}")
    
    try:
        wb = load_workbook(template_path)
        print(f"\n工作表列表: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            analyze_sheet(ws, sheet_name)
        
        # 生成详细报告
        print(f"\n{'='*80}")
        print("详细单元格分析（前20行，前10列）")
        print(f"{'='*80}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n\n工作表: {sheet_name}")
            print("-" * 80)
            
            for row in range(1, min(21, ws.max_row + 1)):
                print(f"\n行 {row}:")
                for col in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row, col)
                    if cell.value is not None:
                        col_letter = get_column_letter(col)
                        cell_info = analyze_cell(cell)
                        print(f"  {col_letter}{row}: {json.dumps(cell_info, ensure_ascii=False, indent=4)}")
        
        wb.close()
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()

