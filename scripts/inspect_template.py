"""查看客户对账单模板结构"""
from openpyxl import load_workbook
from pathlib import Path

template_path = Path("template/statement_template.xlsx")
wb = load_workbook(template_path)

print(f"所有工作表: {wb.sheetnames}\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== 工作表: {sheet_name} ===")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}\n")
    
    # 打印前20行的前15列
    for i in range(1, min(21, ws.max_row + 1)):
        row_data = []
        for j in range(1, min(16, ws.max_column + 1)):
            cell_value = ws.cell(i, j).value
            if cell_value is not None:
                row_data.append(f"{j}:{cell_value}")
        if row_data:
            print(f"行 {i}: {row_data}")
    print("\n" + "="*50 + "\n")

wb.close()

